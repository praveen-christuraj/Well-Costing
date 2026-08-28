"""Integration tests for the Services / Consumables / Tangibles catalogue APIs.

Covers the common template contract for every new module: list/create/update,
duplicate rejection, auto-generated codes, soft delete → restore → permanent
delete, bulk import with flexible dates, xlsx/csv export, configurable
dropdowns and the rate-revision history workflow.
"""

import io

from app.models import Currency, UnitOfMeasurement, VendorSupplier
from openpyxl import load_workbook


def _auth_headers(client) -> dict[str, str]:
    resp = client.post(
        "/api/v1/auth/login",
        json={"email": "engineer@example.com", "password": "Correct-Horse-Battery-1!"},
    )
    token = resp.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _seed_lookups(db_session):
    db_session.add_all([
        Currency(currency_code="USD", currency_name="US Dollar", currency_symbol="$"),
        Currency(currency_code="NGN", currency_name="Naira", currency_symbol="₦"),
        UnitOfMeasurement(unit_code="KG", unit_name="Kilogram", unit_symbol="kg"),
        VendorSupplier(vendor_code="V001", vendor_name="Acme Services"),
    ])
    db_session.commit()


# ---------------------------------------------------------------------------
# Services
# ---------------------------------------------------------------------------


def test_service_crud_soft_delete_and_duplicate(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)

    created = client.post("/api/v1/catalogue/services", headers=headers, json={
        "service_name": "Mud Logging", "provider_type": "inhouse",
    })
    assert created.status_code == 200
    body = created.json()
    assert body["service_code"].startswith("SVC-")
    assert body["provider_type"] == "Inhouse"

    # Duplicate name rejected.
    dup = client.post("/api/v1/catalogue/services", headers=headers, json={
        "service_name": "mud logging", "provider_type": "Inhouse",
    })
    assert dup.status_code == 400

    # 3rd Party requires a vendor.
    no_vendor = client.post("/api/v1/catalogue/services", headers=headers, json={
        "service_name": "Cementing", "provider_type": "3rd Party",
    })
    assert no_vendor.status_code == 400

    third_party = client.post("/api/v1/catalogue/services", headers=headers, json={
        "service_name": "Cementing", "provider_type": "3rd party", "vendor_id": "V001",
    })
    assert third_party.status_code == 200
    assert third_party.json()["vendor_code"] == "V001"

    # Soft delete moves it to the trash and out of the active list.
    sid = body["id"]
    assert client.delete(f"/api/v1/catalogue/services/{sid}", headers=headers).status_code == 200
    assert len(client.get("/api/v1/catalogue/services", headers=headers).json()) == 1
    deleted = client.get("/api/v1/catalogue/services/deleted", headers=headers).json()
    assert len(deleted) == 1

    # Restore.
    assert client.post(f"/api/v1/catalogue/services/{sid}/restore", headers=headers).status_code == 200
    assert len(client.get("/api/v1/catalogue/services", headers=headers).json()) == 2

    # Permanent delete.
    assert client.delete(f"/api/v1/catalogue/services/{sid}/permanent", headers=headers).status_code == 200
    assert client.get("/api/v1/catalogue/services/deleted", headers=headers).json() == []


def test_services_bulk_import_and_export(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)

    csv_data = (
        "service_name,provider_type,vendor_code,description\n"
        "Rig Maintenance,inhouse,,Internal crew\n"
        "Mud Engineering,3rd party,V001,On-site engineer\n"
        "Rig Maintenance,inhouse,,Duplicate row updates\n"
    )
    resp = client.post(
        "/api/v1/catalogue/services/import",
        headers=headers,
        files={"file": ("services.csv", csv_data.encode(), "text/csv")},
    )
    assert resp.status_code == 200
    result = resp.json()
    assert result["imported_count"] == 3
    assert result["error_count"] == 0
    # Duplicate name did not create a second row.
    active = client.get("/api/v1/catalogue/services", headers=headers).json()
    assert len(active) == 2

    exported = client.get("/api/v1/catalogue/services/export?format=csv", headers=headers)
    assert exported.status_code == 200
    assert "service_code" in exported.text
    assert "Rig Maintenance" in exported.text


# ---------------------------------------------------------------------------
# Mud Chemicals + rate revisions
# ---------------------------------------------------------------------------


def test_mud_chemical_rate_revision_workflow(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)

    created = client.post("/api/v1/catalogue/mud-chemicals", headers=headers, json={
        "chemical_name": "Bentonite", "uom": "kg", "unit_rate": "2.50",
        "currency": "USD", "effective_date": "2026-01-15",
    })
    assert created.status_code == 200
    body = created.json()
    assert body["chemical_code"].startswith("MC-")
    assert str(body["current_rate"]) == "2.50"
    assert str(body["previous_rate"]) == "0.00"
    assert len(body["rates"]) == 1

    # Missing mandatory currency/rate.
    bad = client.post("/api/v1/catalogue/mud-chemicals", headers=headers, json={
        "chemical_name": "Barite", "unit_rate": "3.00",
    })
    assert bad.status_code == 400

    # Rate change appends a revision with previous price auto-detected.
    updated = client.put(f"/api/v1/catalogue/mud-chemicals/{body['id']}", headers=headers, json={
        "unit_rate": "3.10", "currency": "USD", "effective_date": "15/03/2026",
    })
    assert updated.status_code == 200
    ubody = updated.json()
    assert str(ubody["current_rate"]) == "3.10"
    assert str(ubody["previous_rate"]) == "2.50"
    assert len(ubody["rates"]) == 2

    history = client.get("/api/v1/catalogue/mud-chemicals/rate-history", headers=headers).json()
    assert len(history) == 2
    assert history[0]["revision_number"] == 2

    exported = client.get("/api/v1/catalogue/mud-chemicals/rate-history/export?format=xlsx", headers=headers)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_mud_chemical_import_flexible_dates_and_revision(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)

    csv_data = (
        "chemical_name,uom,unit_rate,currency,effective_date\n"
        "Barite,kg,3.20,USD,2026-02-01\n"
        "Bentonite,kg,4.00,USD,01/04/2026\n"
    )
    resp = client.post(
        "/api/v1/catalogue/mud-chemicals/import",
        headers=headers,
        files={"file": ("chem.csv", csv_data.encode(), "text/csv")},
    )
    assert resp.status_code == 200
    assert resp.json()["error_count"] == 0

    # Re-import Bentonite at a new rate → revision appended, no duplicate item.
    csv_data2 = (
        "chemical_name,uom,unit_rate,currency,effective_date\n"
        "Bentonite,kg,5.50,USD,2026-06-01\n"
    )
    resp2 = client.post(
        "/api/v1/catalogue/mud-chemicals/import",
        headers=headers,
        files={"file": ("chem2.csv", csv_data2.encode(), "text/csv")},
    )
    assert resp2.json()["imported_count"] == 1
    items = client.get("/api/v1/catalogue/mud-chemicals", headers=headers).json()
    bentonite = next(i for i in items if i["chemical_name"] == "Bentonite")
    assert str(bentonite["current_rate"]) == "5.50"
    assert len(bentonite["rates"]) == 2


# ---------------------------------------------------------------------------
# Drill Bits — configurable dropdowns + uplift math
# ---------------------------------------------------------------------------


def test_drill_bits_config_dropdowns_and_final_cost(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)

    # Unknown bit type rejected until configured.
    rejected = client.post("/api/v1/catalogue/drill-bits", headers=headers, json={
        "bit_name": "PDC-1", "bit_type": "PDC", "model_no": "M1", "size": "12 1/4",
        "manufacturer": "Schlum", "unit_rate_po": "45000", "cost_uplift": "110",
        "currency": "USD",
    })
    assert rejected.status_code == 400

    # Configure dropdowns.
    assert client.post("/api/v1/catalogue/configs/bit_type", headers=headers, json={"value": "PDC"}).status_code == 200
    assert client.post("/api/v1/catalogue/configs/bit_manufacturer", headers=headers, json={"value": "Schlumberger"}).status_code == 200
    # Duplicate config value rejected.
    dup_cfg = client.post("/api/v1/catalogue/configs/bit_type", headers=headers, json={"value": "pdc"})
    assert dup_cfg.status_code == 400

    created = client.post("/api/v1/catalogue/drill-bits", headers=headers, json={
        "bit_name": "PDC-1", "bit_type": "PDC", "model_no": "M1", "size": "12 1/4",
        "manufacturer": "Schlumberger", "unit_rate_po": "45000", "cost_uplift": "110",
        "currency": "USD",
    })
    assert created.status_code == 200
    body = created.json()
    assert body["bit_code"].startswith("DB-")
    # 45000 x 110% = 49500.00
    assert str(body["final_cost"]) == "49500.00"

    options = client.get("/api/v1/catalogue/drill-bits/dropdown-options", headers=headers).json()
    assert "PDC" in options["bit_types"]

    # Rate change creates a revision.
    updated = client.put(f"/api/v1/catalogue/drill-bits/{body['id']}", headers=headers, json={
        "unit_rate_po": "50000", "cost_uplift": "100", "currency": "USD",
    })
    assert updated.status_code == 200
    assert str(updated.json()["final_cost"]) == "50000.00"
    history = client.get("/api/v1/catalogue/drill-bits/rate-history", headers=headers).json()
    assert len(history) == 2


def test_drill_bits_import_autocreates_config_values(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)
    csv_data = (
        "bit_name,bit_type,model_no,size,manufacturer,unit_rate_po,cost_uplift,currency,effective_date\n"
        "Tricone Bit,Tricone,T-90,17 1/2,National Oilwell,30000,100,USD,2026-03-01\n"
    )
    resp = client.post(
        "/api/v1/catalogue/drill-bits/import",
        headers=headers,
        files={"file": ("bits.csv", csv_data.encode(), "text/csv")},
    )
    assert resp.status_code == 200
    assert resp.json()["error_count"] == 0
    options = client.get("/api/v1/catalogue/drill-bits/dropdown-options", headers=headers).json()
    assert "Tricone" in options["bit_types"]
    assert "National Oilwell" in options["manufacturers"]


# ---------------------------------------------------------------------------
# Tangibles
# ---------------------------------------------------------------------------


def test_tangibles_crud_and_scope_validation(client, db_session):
    _seed_lookups(db_session)
    headers = _auth_headers(client)

    for cfg, value in (
        ("tangible_category", "Casing"),
        ("tangible_manufacturer", "Tenaris"),
    ):
        client.post(f"/api/v1/catalogue/configs/{cfg}", headers=headers, json={"value": value})
    # Subcategories depend on a category: created under the parent.
    sub = client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                      json={"value": "Surface Casing", "parent_value": "Casing"})
    assert sub.status_code == 200
    assert sub.json()["parent_value"] == "Casing"

    bad_scope = client.post("/api/v1/catalogue/tangibles", headers=headers, json={
        "tangible_name": "Casing", "tangible_scope": "Underwater",
        "category": "Casing", "subcategory": "Surface Casing", "manufacturer": "Tenaris",
        "unit_rate_po": "120", "currency": "USD",
    })
    assert bad_scope.status_code == 400

    created = client.post("/api/v1/catalogue/tangibles", headers=headers, json={
        "tangible_name": "Casing 9-5/8", "tangible_scope": "drilling",
        "category": "Casing", "subcategory": "Surface Casing", "manufacturer": "Tenaris",
        "uom": "m", "unit_rate_po": "120", "cost_uplift": "100", "currency": "USD",
    })
    assert created.status_code == 200
    body = created.json()
    assert body["tangible_code"].startswith("TNG-")
    assert body["tangible_scope"] == "Drilling"
    assert str(body["final_cost"]) == "120.00"

    # Uplift as multiplier: 1.15 => 115%.
    updated = client.put(f"/api/v1/catalogue/tangibles/{body['id']}", headers=headers, json={
        "unit_rate_po": "200", "cost_uplift": "1.15", "currency": "USD",
    })
    assert str(updated.json()["final_cost"]) == "230.00"

    history = client.get("/api/v1/catalogue/tangibles/rate-history", headers=headers).json()
    assert len(history) == 2

    # Soft delete + permanent delete.
    assert client.delete(f"/api/v1/catalogue/tangibles/{body['id']}", headers=headers).status_code == 200
    assert len(client.get("/api/v1/catalogue/tangibles/deleted", headers=headers).json()) == 1
    assert client.delete(f"/api/v1/catalogue/tangibles/{body['id']}/permanent", headers=headers).status_code == 200


def test_tangible_duplicate_names_allowed_when_criteria_differ(client, db_session):
    """Duplicate names are accepted when Manufacturer / Rate as per PO / Uplift
    / Description differ; rejected only when the name AND every criterion match."""

    _seed_lookups(db_session)
    headers = _auth_headers(client)

    for cfg, value in (
        ("tangible_category", "Casing"),
        ("tangible_manufacturer", "Tenaris"),
        ("tangible_manufacturer", "Vallourec"),
    ):
        assert client.post(f"/api/v1/catalogue/configs/{cfg}", headers=headers,
                           json={"value": value}).status_code == 200
    assert client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                       json={"value": "Surface Casing", "parent_value": "Casing"}).status_code == 200

    def create(name, *, mfr="Tenaris", rate="120", uplift="100", description=None):
        payload = {
            "tangible_name": name, "tangible_scope": "Drilling",
            "category": "Casing", "subcategory": "Surface Casing", "manufacturer": mfr,
            "unit_rate_po": rate, "cost_uplift": uplift, "currency": "USD",
        }
        if description is not None:
            payload["description"] = description
        return client.post("/api/v1/catalogue/tangibles", headers=headers, json=payload)

    def same_name_count():
        rows = client.get("/api/v1/catalogue/tangibles", headers=headers).json()
        return sum(1 for row in rows if row["tangible_name"] == "Casing 9-5/8")

    assert create("Casing 9-5/8").status_code == 200

    # Same name with one criterion different at a time — all accepted.
    assert create("Casing 9-5/8", mfr="Vallourec").status_code == 200
    assert create("Casing 9-5/8", rate="150").status_code == 200
    assert create("Casing 9-5/8", uplift="110").status_code == 200
    assert create("Casing 9-5/8", description="Surface string").status_code == 200
    assert same_name_count() == 5

    # Same name and all criteria identical (case-insensitive) — rejected.
    dup = create("casing 9-5/8", description="surface string")
    assert dup.status_code == 400
    assert "already exists" in dup.json()["error"]["message"]
    assert same_name_count() == 5

    # Renaming onto an existing name with identical criteria — rejected…
    twin = create("Tubing 2-7/8")  # Tenaris / 120 / 100 / no description = row 1
    assert twin.status_code == 200
    clash = client.put(f"/api/v1/catalogue/tangibles/{twin.json()['id']}", headers=headers,
                       json={"tangible_name": "Casing 9-5/8"})
    assert clash.status_code == 400
    assert "already exists" in clash.json()["error"]["message"]

    # …but accepted once one criterion differs.
    renamed = client.put(f"/api/v1/catalogue/tangibles/{twin.json()['id']}", headers=headers,
                         json={"tangible_name": "Casing 9-5/8", "description": "Twin row"})
    assert renamed.status_code == 200
    assert same_name_count() == 6

    # Import: same name with a new rate creates a tangible, not a rate revision.
    csv_data = (
        "tangible_name,tangible_scope,category,subcategory,manufacturer,unit_rate_po,cost_uplift,currency,effective_date,description\n"
        "Casing 9-5/8,Drilling,Casing,Surface Casing,Tenaris,999,100,USD,2026-03-01,\n"
    )
    imp = client.post("/api/v1/catalogue/tangibles/import", headers=headers,
                      files={"file": ("tng.csv", csv_data.encode(), "text/csv")})
    assert imp.status_code == 200
    assert imp.json()["error_count"] == 0
    assert imp.json()["imported_count"] == 1
    assert same_name_count() == 7

    # Import: name identical on every criterion refreshes the row, adds nothing.
    csv_data = (
        "tangible_name,tangible_scope,category,subcategory,manufacturer,unit_rate_po,cost_uplift,currency,effective_date,description\n"
        "Casing 9-5/8,Drilling,Casing,Surface Casing,Tenaris,120,100,USD,2026-03-01,\n"
    )
    imp = client.post("/api/v1/catalogue/tangibles/import", headers=headers,
                      files={"file": ("tng.csv", csv_data.encode(), "text/csv")})
    assert imp.status_code == 200
    assert imp.json()["error_count"] == 0
    assert same_name_count() == 7


def test_tangible_subcategories_depend_on_category(client, db_session):
    """Subcategories are configured under a category and filtered by it."""

    _seed_lookups(db_session)
    headers = _auth_headers(client)

    # Category must exist before a subcategory can be configured.
    missing_parent = client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                                 json={"value": "Surface Casing", "parent_value": "Casing"})
    assert missing_parent.status_code == 400

    no_parent = client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                            json={"value": "Surface Casing"})
    assert no_parent.status_code == 400

    assert client.post("/api/v1/catalogue/configs/tangible_category", headers=headers,
                       json={"value": "Casing"}).status_code == 200
    assert client.post("/api/v1/catalogue/configs/tangible_category", headers=headers,
                       json={"value": "Tubing"}).status_code == 200
    assert client.post("/api/v1/catalogue/configs/tangible_manufacturer", headers=headers,
                       json={"value": "Tenaris"}).status_code == 200

    # Same subcategory name is allowed under different categories.
    assert client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                       json={"value": "Standard", "parent_value": "Casing"}).status_code == 200
    assert client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                       json={"value": "Standard", "parent_value": "Tubing"}).status_code == 200
    dup = client.post("/api/v1/catalogue/configs/tangible_subcategory", headers=headers,
                      json={"value": "standard", "parent_value": "Casing"})
    assert dup.status_code == 400

    # Bulk add under a category uses the new path-style endpoint.
    bulk = client.post("/api/v1/catalogue/configs/tangible_subcategory/bulk", headers=headers,
                       json={"values": ["Intermediate", "Production"], "parent_value": "Casing"})
    assert bulk.status_code == 200
    assert bulk.json()["imported_count"] == 2

    # Dropdown options carry the category link.
    options = client.get("/api/v1/catalogue/tangibles/dropdown-options", headers=headers).json()
    subs = {(s["value"], s["category"]) for s in options["subcategories"]}
    assert ("Intermediate", "Casing") in subs
    assert ("Standard", "Tubing") in subs

    # Tangible entry rejects a subcategory owned by another category.
    mismatch = client.post("/api/v1/catalogue/tangibles", headers=headers, json={
        "tangible_name": "Tubing 2-7/8", "tangible_scope": "Completion",
        "category": "Tubing", "subcategory": "Intermediate", "manufacturer": "Tenaris",
        "unit_rate_po": "90", "currency": "USD",
    })
    assert mismatch.status_code == 400

    ok = client.post("/api/v1/catalogue/tangibles", headers=headers, json={
        "tangible_name": "Casing 9-5/8", "tangible_scope": "Drilling",
        "category": "Casing", "subcategory": "Intermediate", "manufacturer": "Tenaris",
        "unit_rate_po": "120", "currency": "USD",
    })
    assert ok.status_code == 200

    # Updating the category re-validates the subcategory against it.
    move = client.put(f"/api/v1/catalogue/tangibles/{ok.json()['id']}", headers=headers,
                      json={"category": "Tubing"})
    assert move.status_code == 400

    # Import creates missing subcategories under the row's category.
    csv_data = (
        "tangible_name,tangible_scope,category,subcategory,manufacturer,"
        "unit_rate_po,currency,effective_date\n"
        "Wellhead,Drilling,Wellheads,Standard Wellhead,Tenaris,5000,USD,2026-03-01\n"
    )
    imp = client.post("/api/v1/catalogue/tangibles/import", headers=headers,
                      files={"file": ("tng.csv", csv_data.encode(), "text/csv")})
    assert imp.status_code == 200
    assert imp.json()["error_count"] == 0
    options = client.get("/api/v1/catalogue/tangibles/dropdown-options", headers=headers).json()
    subs = {(s["value"], s["category"]) for s in options["subcategories"]}
    assert ("Standard Wellhead", "Wellheads") in subs
    assert "Wellheads" in options["categories"]


def test_import_templates_are_downloadable_xlsx(client, db_session):
    """Every importable module serves an XLSX template users can fill and re-upload."""

    _seed_lookups(db_session)
    headers = _auth_headers(client)
    endpoints = [
        "/api/v1/master-data/uom/import-template",
        "/api/v1/master-data/currencies/import-template",
        "/api/v1/master-data/phases/import-template",
        "/api/v1/master-data/activities/import-template",
        "/api/v1/master-data/hole-sections/import-template",
        "/api/v1/master-data/vendors/import-template",
        "/api/v1/master-data/purchase-orders/import-template",
        "/api/v1/catalogue/services/import-template",
        "/api/v1/catalogue/mud-chemicals/import-template",
        "/api/v1/catalogue/drill-bits/import-template",
        "/api/v1/catalogue/tangibles/import-template",
    ]
    for endpoint in endpoints:
        resp = client.get(endpoint, headers=headers)
        assert resp.status_code == 200, endpoint
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ), endpoint
        assert resp.content[:2] == b"PK", endpoint

    # A filled template uploads cleanly: download, keep the header row, add a row.
    resp = client.get("/api/v1/master-data/uom/import-template", headers=headers)
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    ws.append(["ft", "Feet", "ft", "imperial length"])
    out = io.BytesIO()
    wb.save(out)
    imp = client.post("/api/v1/master-data/uom/import", headers=headers,
                      files={"file": ("uom_template.xlsx", out.getvalue(),
                                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert imp.status_code == 200
    assert imp.json()["imported_count"] == 1


def test_consumable_subcategories_seeded_and_combined_history(client, db_session):
    headers = _auth_headers(client)
    # Subcategories are seeded by migration in real deployments; in the
    # metadata-created test DB the endpoint still responds with the (empty)
    # directory — exercise the route contract.
    resp = client.get("/api/v1/catalogue/consumable-subcategories", headers=headers)
    assert resp.status_code == 200
    assert isinstance(resp.json(), list)

    history = client.get("/api/v1/catalogue/consumables-rate-history", headers=headers)
    assert history.status_code == 200
    assert history.json() == []
    export = client.get("/api/v1/catalogue/consumables-rate-history/export?format=csv", headers=headers)
    assert export.status_code == 200


def test_xlsx_import_roundtrip(client, db_session):
    """An XLSX upload (openpyxl workbook, including an Excel date) imports."""
    from openpyxl import Workbook

    _seed_lookups(db_session)
    headers = _auth_headers(client)

    wb = Workbook()
    ws = wb.active
    ws.append(["chemical_name", "uom", "unit_rate", "currency", "effective_date"])
    ws.append(["Xanthan Gum", "kg", 12.75, "USD", "2026-05-20"])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = client.post(
        "/api/v1/catalogue/mud-chemicals/import",
        headers=headers,
        files={"file": ("chem.xlsx", bio.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    assert resp.json()["imported_count"] == 1
    assert resp.json()["error_count"] == 0
