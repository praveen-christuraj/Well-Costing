"""Phase 8 separate cost-state staging, Excel, posting audit, and lineage tests."""

from io import BytesIO

from app.models.cost_control import CostControlBatch, CostControlPostAttempt, CostTransaction
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from tests.integration.test_estimate_build import auth, create, submitted_requirement

COST_STATES = ["field_estimate", "commitment", "accrual", "actual", "forecast"]


def setup_estimate(client: TestClient, headers: dict[str, str]) -> tuple[dict, dict]:
    requirement, refs = submitted_requirement(client, headers)
    currency = create(
        client,
        "/api/v1/master-data/currencies",
        {"code": "USD", "name": "US Dollar"},
        headers,
    )
    vendor = create(
        client,
        "/api/v1/master-data/vendors",
        {"code": "V-P8", "name": "Phase 8 Vendor"},
        headers,
    )
    estimate = create(
        client,
        "/api/v1/estimates/from-requirement",
        {
            "requirement_id": requirement["id"],
            "code": "EST-P8-CONTROL",
            "title": "Phase 8 cost control",
            "currency_id": currency["id"],
        },
        headers,
    )
    return estimate, {**refs, "currency": currency, "vendor": vendor}


def row() -> dict[str, str]:
    return {
        "transaction_date": "2026-08-13",
        "source_document_type": "field_ticket",
        "source_document_reference": "FT-001",
        "external_transaction_id": "EXT-001",
        "cost_code": "CC-001",
        "vendor_code": "V-P8",
        "description": "Synthetic field cost",
        "quantity": "1.0000",
        "unit_code": "DAY",
        "currency_code": "USD",
        "amount": "1250.0000",
        "correction_kind": "original",
    }


def test_all_cost_states_stage_separately_and_posting_fails_closed(
    client: TestClient, db_session: Session
) -> None:
    headers = auth(client)
    estimate, _refs = setup_estimate(client, headers)
    version_id = estimate["versions"][0]["id"]
    batch_ids: list[str] = []
    for cost_state in COST_STATES:
        response = client.post(
            "/api/v1/cost-control/batches/validate",
            json={
                "estimate_version_id": version_id,
                "cost_state": cost_state,
                "rows": [{**row(), "external_transaction_id": f"EXT-{cost_state}"}],
            },
            headers=headers,
        )
        assert response.status_code == 200, response.text
        assert response.json()["cost_state"] == cost_state
        assert response.json()["status"] == "validated"
        assert response.json()["afe_snapshot_id"] is None
        batch_ids.append(response.json()["id"])

    assert db_session.scalar(select(func.count()).select_from(CostControlBatch)) == 5
    post = client.post(f"/api/v1/cost-control/batches/{batch_ids[0]}/post", headers=headers)
    assert post.status_code == 422
    error = post.json()["error"]
    assert error["code"] == "cost_state_policy_pending"
    assert error["details"]["cost_state_policy_version"] == "pending-all-cost-states"

    blocked = client.get(f"/api/v1/cost-control/batches/{batch_ids[0]}", headers=headers)
    assert blocked.status_code == 200
    assert blocked.json()["status"] == "blocked"
    assert blocked.json()["post_attempts"][0]["status"] == "blocked"
    assert blocked.json()["post_attempts"][0]["created_by"] is not None
    assert db_session.scalar(select(func.count()).select_from(CostControlPostAttempt)) == 1
    assert db_session.scalar(select(func.count()).select_from(CostTransaction)) == 0


def test_excel_preview_template_and_reversal_lineage_validation(client: TestClient) -> None:
    headers = auth(client)
    estimate, _refs = setup_estimate(client, headers)
    version_id = estimate["versions"][0]["id"]

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(list(row()))
    sheet.append(list(row().values()))
    stream = BytesIO()
    workbook.save(stream)
    preview = client.post(
        "/api/v1/cost-control/imports/preview",
        data={"estimate_version_id": version_id, "cost_state": "actual"},
        files={
            "file": (
                "cost-control.xlsx",
                stream.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["batch"]["status"] == "validated"
    assert preview.json()["batch"]["source_type"] == "excel"
    assert preview.json()["applied_mapping"]["amount"] == "amount"

    template = client.get("/api/v1/cost-control/template", headers=headers)
    assert template.status_code == 200
    template_book = load_workbook(BytesIO(template.content), read_only=True)
    template_sheet = template_book.active
    assert template_sheet is not None
    first_row = next(template_sheet.iter_rows(values_only=True))
    assert next(iter(first_row)) == "transaction_date"

    invalid_reversal = client.post(
        "/api/v1/cost-control/batches/validate",
        json={
            "estimate_version_id": version_id,
            "cost_state": "actual",
            "rows": [{**row(), "correction_kind": "reversal"}],
        },
        headers=headers,
    )
    assert invalid_reversal.status_code == 422
