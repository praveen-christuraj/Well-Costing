"""Current AFE Cost Estimate workflow integration coverage.

The released flow is intentionally narrow:

* AFE Lines hold submitted scope only — never consumable usage/day or a planned
  quantity/UOM.
* AFE Cost Estimates is the only place the estimate rate is configured.
* Only submitted AFEs can be priced.
* Daily Cost captures actual consumable quantity and UOM, using the saved AFE
  Cost Estimate rate.
* Pricing, print, and export actions are auditable.
"""

from decimal import Decimal
from io import BytesIO
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.conftest import TEST_PASSWORD

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def headers(client: TestClient) -> dict[str, str]:
    login = client.post(
        "/api/v1/auth/login",
        json={"email": "engineer@example.com", "password": TEST_PASSWORD},
    )
    assert login.status_code == 200, login.text
    return {"Authorization": f"Bearer {login.json()['access_token']}"}


def post(
    client: TestClient, path: str, payload: dict[str, Any], auth: dict[str, str]
) -> dict[str, Any]:
    response = client.post(path, json=payload, headers=auth)
    assert response.status_code == 201, response.text
    return response.json()


def setup_current_scope(client: TestClient, auth: dict[str, str]) -> dict[str, Any]:
    """Create classification-only AFE lines exactly as the current UI does."""
    primary = post(
        client,
        "/api/v1/master-data/primary-categories",
        {"code": "OPERATIONS", "name": "Operations"},
        auth,
    )
    secondary = post(
        client,
        "/api/v1/master-data/secondary-categories",
        {
            "code": "DRILLING",
            "name": "Drilling scope",
            "primary_category_id": primary["id"],
        },
        auth,
    )
    cost_category = post(
        client,
        "/api/v1/master-data/cost-categories",
        {
            "code": "DRILL-COST",
            "name": "Drilling cost",
            "primary_category_id": primary["id"],
            "secondary_category_id": secondary["id"],
        },
        auth,
    )
    cost_code = post(
        client,
        "/api/v1/master-data/cost-codes",
        {"code": "DR-001", "name": "Drilling", "cost_category_id": cost_category["id"]},
        auth,
    )
    barrel = post(
        client,
        "/api/v1/master-data/units",
        {"code": "BBL", "name": "Barrel"},
        auth,
    )
    project = post(
        client,
        "/api/v1/projects",
        {"code": "PRJ-EST-CURRENT", "name": "Current Estimate Project"},
        auth,
    )
    well = post(
        client,
        "/api/v1/wells",
        {"project_id": project["id"], "code": "W-EST-CURRENT", "name": "Current Estimate Well"},
        auth,
    )
    afe = post(
        client,
        "/api/v1/afes",
        {
            "well_id": well["id"],
            "code": "AFE-EST-CURRENT",
            "title": "Current AFE Estimate",
            "budget_amount": "50000.00",
            "total_planned_days": "10.0",
        },
        auth,
    )
    service_line = post(
        client,
        f"/api/v1/afes/{afe['id']}/lines",
        {
            "line_number": 1,
            "secondary_category_id": secondary["id"],
            "cost_code_id": cost_code["id"],
            "service_type": "service",
            "rate_basis": "daily",
        },
        auth,
    )
    consumable_line = post(
        client,
        f"/api/v1/afes/{afe['id']}/lines",
        {
            "line_number": 2,
            "secondary_category_id": secondary["id"],
            "cost_code_id": cost_code["id"],
            "service_type": "consumable",
            "rate_basis": "per_unit",
        },
        auth,
    )
    return {
        "primary": primary,
        "secondary": secondary,
        "cost_code": cost_code,
        "barrel": barrel,
        "project": project,
        "well": well,
        "afe": afe,
        "service_line": service_line,
        "consumable_line": consumable_line,
    }


def submit_and_price(
    client: TestClient, auth: dict[str, str], refs: dict[str, Any]
) -> dict[str, Any]:
    afe_id = refs["afe"]["id"]
    submitted = client.post(f"/api/v1/afes/{afe_id}/submit", headers=auth)
    assert submitted.status_code == 200, submitted.text
    assert submitted.json()["status"] == "submitted"

    priced = client.put(
        f"/api/v1/afes/{afe_id}/cost-estimate/rates",
        json={
            "rates": [
                {"afe_line_id": refs["service_line"]["id"], "unit_rate": "1200.00"},
                {"afe_line_id": refs["consumable_line"]["id"], "unit_rate": "50.00"},
            ]
        },
        headers=auth,
    )
    assert priced.status_code == 200, priced.text
    return priced.json()


def test_scope_only_lines_do_not_require_usage_and_submitted_afe_can_be_priced(
    client: TestClient,
) -> None:
    auth = headers(client)
    refs = setup_current_scope(client, auth)
    afe_id = refs["afe"]["id"]

    # Scope-only line creation succeeds with no daily usage, quantity, or UOM.
    assert refs["consumable_line"]["quantity"] is None
    assert refs["consumable_line"]["unit_id"] is None
    assert refs["consumable_line"]["daily_consumption"] is None

    # The server enforces the submitted workflow gate, not just the UI filter.
    draft_get = client.get(f"/api/v1/afes/{afe_id}/cost-estimate", headers=auth)
    assert draft_get.status_code == 422
    assert "Only submitted AFEs" in draft_get.json()["error"]["message"]
    draft_save = client.put(
        f"/api/v1/afes/{afe_id}/cost-estimate/rates",
        json={"rates": []},
        headers=auth,
    )
    assert draft_save.status_code == 422

    estimate = submit_and_price(client, auth, refs)
    assert estimate["afe_status"] == "submitted"
    assert len(estimate["lines"]) == 2
    # Regression: unit_id is null for scope-only lines and must not cause the
    # estimate serializer to return "An unexpected error occurred".
    assert estimate["lines"][0]["quantity"] is None
    assert estimate["lines"][1]["quantity"] is None
    assert estimate["lines"][0]["unit_id"] is None
    assert estimate["lines"][1]["unit_id"] is None
    assert Decimal(str(estimate["lines"][0]["estimated_amount"])) == Decimal("1200.00")
    assert Decimal(str(estimate["lines"][1]["estimated_amount"])) == Decimal("50.00")
    assert Decimal(str(estimate["estimated_total"])) == Decimal("1250.00")

    page = client.get(
        f"/api/v1/afes?well_id={refs['well']['id']}&status=submitted",
        headers=auth,
    )
    assert page.status_code == 200
    assert [item["id"] for item in page.json()["items"]] == [afe_id]


def test_consumable_actual_quantity_and_uom_are_captured_in_daily_cost(
    client: TestClient,
) -> None:
    auth = headers(client)
    refs = setup_current_scope(client, auth)
    submit_and_price(client, auth, refs)

    reference = client.get(
        f"/api/v1/wells/{refs['well']['id']}/daily-cost/reference-rates",
        headers=auth,
    )
    assert reference.status_code == 200, reference.text
    rates = reference.json()
    assert rates["afe_id"] == refs["afe"]["id"]
    assert rates["consumables"][0]["afe_line_id"] == refs["consumable_line"]["id"]
    assert rates["consumables"][0]["unit_id"] is None

    activity = post(
        client,
        "/api/v1/master-data/activities",
        {"code": "PLANNED", "name": "Planned"},
        auth,
    )
    well_activity = post(
        client,
        "/api/v1/well-activities",
        {
            "well_id": refs["well"]["id"],
            "activity_id": activity["id"],
            "name": "Planned operations",
        },
        auth,
    )
    entry = post(
        client,
        f"/api/v1/wells/{refs['well']['id']}/daily-cost",
        {
            "well_id": refs["well"]["id"],
            "afe_id": refs["afe"]["id"],
            "entry_date": "2026-08-26",
            "sub_activity_id": well_activity["id"],
            "consumables": [
                {
                    "afe_line_id": refs["consumable_line"]["id"],
                    "consumable_id": None,
                    "cost_code_id": refs["cost_code"]["id"],
                    "quantity": "8.5",
                    # The operator supplies the actual UOM here, not on AFE.
                    "unit_id": refs["barrel"]["id"],
                    "unit_rate": "1.00",
                }
            ],
        },
        auth,
    )
    assert entry["consumables"][0]["unit_id"] == refs["barrel"]["id"]
    assert Decimal(str(entry["consumables"][0]["unit_rate"])) == Decimal("50.00")
    assert Decimal(str(entry["total_consumables_cost"])) == Decimal("425.00")


def test_pricing_print_export_and_rate_changes_are_audited(client: TestClient) -> None:
    auth = headers(client)
    refs = setup_current_scope(client, auth)
    estimate = submit_and_price(client, auth, refs)
    afe_id = refs["afe"]["id"]

    afe_printed = client.post(f"/api/v1/afes/{afe_id}/audit/print", headers=auth)
    assert afe_printed.status_code == 204
    printed = client.post(f"/api/v1/afes/{afe_id}/cost-estimate/audit/print", headers=auth)
    assert printed.status_code == 204
    exported = client.get(f"/api/v1/afes/{afe_id}/cost-estimate/export", headers=auth)
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
    workbook = load_workbook(BytesIO(exported.content))
    assert workbook.sheetnames == ["AFE Cost Estimate", "Summaries"]
    assert "Estimated total rate" in [cell.value for cell in workbook["AFE Cost Estimate"][13]]

    audit = client.get(
        "/api/v1/audit-logs?entity_type=afe_cost_estimate&page=1&page_size=50",
        headers=auth,
    )
    assert audit.status_code == 200, audit.text
    actions = {record["action"] for record in audit.json()["items"]}
    assert {"save_rates", "print", "export"} <= actions
    saved = next(record for record in audit.json()["items"] if record["action"] == "save_rates")
    assert "rates_saved" in (saved["details"] or "")

    detail = client.get(f"/api/v1/afes/{afe_id}", headers=auth)
    assert detail.status_code == 200
    local_actions = {record["action"] for record in detail.json()["audit_logs"]}
    assert "printed" in local_actions
    assert "cost_estimate_rates_saved" in local_actions
    assert "cost_estimate_printed" in local_actions
    assert "cost_estimate_exported" in local_actions
    assert Decimal(str(estimate["estimated_total"])) == Decimal("1250.00")
