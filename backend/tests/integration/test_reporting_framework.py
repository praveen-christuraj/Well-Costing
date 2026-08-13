"""Phase 9 shared-dimension reporting and audited Excel shell tests."""

from io import BytesIO

from app.models.reporting import ReportExportAttempt
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from tests.integration.test_estimate_build import auth


def test_pending_financial_metrics_and_audited_excel_export(
    client: TestClient, db_session: Session
) -> None:
    headers = auth(client)
    contract = client.get("/api/v1/reports/contracts/v1", headers=headers)
    assert contract.status_code == 200
    assert contract.json()["contract_version"] == "1.0"
    assert contract.json()["direct_grants_status"] == "not_applied"
    assert contract.json()["transactional_schema_public"] is False
    assert len(contract.json()["views"]) == 9

    response = client.get(
        "/api/v1/reports/cost-overview?cost_state=actual&project_code=PRJ-001",
        headers=headers,
    )
    assert response.status_code == 200, response.text
    report = response.json()
    assert report["policy_version"] == "pending-shared-cost-reporting"
    assert report["metric_status"] == "policy_pending"
    assert report["filters"]["cost_state"] == "actual"
    assert len(report["dimensions"]) == 13
    assert len(report["state_summaries"]) == 5
    assert all(item["amount"] is None for item in report["state_summaries"])
    assert report["variance_to_afe"] is None
    assert report["forecast_at_completion"] is None
    assert report["drill_through"] == []
    assert len(report["pending_metrics"]) == 6

    exported = client.get(
        "/api/v1/reports/cost-overview/export?cost_state=actual",
        headers=headers,
    )
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    assert workbook.sheetnames == ["State Summary", "Drill Through", "Pending Metrics"]
    assert workbook["State Summary"]["C2"].value is None
    assert workbook["Pending Metrics"]["B1"].value == "pending-shared-cost-reporting"

    attempt = db_session.scalar(select(ReportExportAttempt))
    assert attempt is not None
    assert attempt.status == "completed_shell"
    assert attempt.created_by is not None
    assert attempt.file_sha256 is not None
    assert db_session.scalar(select(func.count()).select_from(ReportExportAttempt)) == 1
