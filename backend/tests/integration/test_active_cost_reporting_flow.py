"""End-to-end coverage for the active costing and reporting source chain."""

from decimal import Decimal
from io import BytesIO
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.conftest import TEST_PASSWORD


def test_afe_estimate_daily_cost_reports_and_audit_export(client: TestClient) -> None:
    login = client.post(
        "/api/v1/auth/login",
        json={"email": "engineer@example.com", "password": TEST_PASSWORD},
    )
    headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

    # Retired Cost Builder/version/snapshot/staging boundaries are not part of
    # the active API surface anymore.
    for retired_path in (
        "/api/v1/estimates",
        "/api/v1/cost-control/batches",
        "/api/v1/enterprise-config/summary",
    ):
        assert client.get(retired_path, headers=headers).status_code == 404

    def create(path: str, payload: dict[str, Any]) -> dict[str, Any]:
        response = client.post(path, json=payload, headers=headers)
        assert response.status_code == 201, response.text
        return response.json()

    primary = create(
        "/api/v1/master-data/primary-categories",
        {"code": "OPERATIONS", "name": "Well Operations"},
    )
    secondary = create(
        "/api/v1/master-data/secondary-categories",
        {
            "code": "DRILLING",
            "name": "Drilling Scope",
            "primary_category_id": primary["id"],
        },
    )
    cost_category = create(
        "/api/v1/master-data/cost-categories",
        {
            "code": "DRILL-COST",
            "name": "Drilling Cost",
            "primary_category_id": primary["id"],
            "secondary_category_id": secondary["id"],
        },
    )
    cost_code = create(
        "/api/v1/master-data/cost-codes",
        {"code": "DR-001", "name": "Drilling", "cost_category_id": cost_category["id"]},
    )
    day = create("/api/v1/master-data/units", {"code": "DAY-X", "name": "Day"})
    each = create("/api/v1/master-data/units", {"code": "EA-X", "name": "Each"})
    project = create("/api/v1/projects", {"code": "PRJ-ACTIVE", "name": "Active chain"})
    well = create(
        "/api/v1/wells",
        {"project_id": project["id"], "code": "W-ACTIVE", "name": "Active Well"},
    )
    afe = create(
        "/api/v1/afes",
        {
            "well_id": well["id"],
            "code": "AFE-ACTIVE",
            "title": "Active AFE",
            "budget_amount": "5000",
            "total_planned_days": "2",
        },
    )
    operation_line = create(
        f"/api/v1/afes/{afe['id']}/lines",
        {
            "line_number": 1,
            "secondary_category_id": secondary["id"],
            "cost_code_id": cost_code["id"],
            "quantity": "2",
            "unit_id": day["id"],
            "rate_basis": "daily",
        },
    )
    quantity_line = create(
        f"/api/v1/afes/{afe['id']}/lines",
        {
            "line_number": 2,
            "secondary_category_id": secondary["id"],
            "cost_code_id": cost_code["id"],
            "quantity": "10",
            "unit_id": each["id"],
            "rate_basis": "per_unit",
        },
    )

    priced = client.put(
        f"/api/v1/afes/{afe['id']}/cost-estimate/rates",
        json={
            "rates": [
                {"afe_line_id": operation_line["id"], "unit_rate": "1000"},
                {"afe_line_id": quantity_line["id"], "unit_rate": "50"},
            ]
        },
        headers=headers,
    )
    assert priced.status_code == 200, priced.text
    estimate = priced.json()
    assert "services_total" not in estimate
    assert "totals_by_item_type" not in estimate
    assert estimate["lines"][0]["primary_category_name"] == "Well Operations"
    assert estimate["lines"][0]["secondary_category_name"] == "Drilling Scope"
    assert Decimal(str(estimate["estimated_total"])) == Decimal("2500")

    assert client.post(f"/api/v1/afes/{afe['id']}/submit", headers=headers).status_code == 200
    reference = client.get(
        f"/api/v1/wells/{well['id']}/daily-cost/reference-rates", headers=headers
    ).json()
    assert reference["services"][0]["afe_line_id"] == operation_line["id"]
    assert reference["services"][0]["service_id"] is None
    assert reference["consumables"][0]["afe_line_id"] == quantity_line["id"]
    assert "item_type" not in reference["consumables"][0]

    activity = create(
        "/api/v1/master-data/activities",
        {"code": "PLANNED-X", "name": "Planned", "sequence": 1},
    )
    well_activity = create(
        "/api/v1/well-activities",
        {
            "well_id": well["id"],
            "activity_id": activity["id"],
            "name": "Planned drilling",
            "responsible_party": "Drilling team",
        },
    )
    daily = create(
        f"/api/v1/wells/{well['id']}/daily-cost",
        {
            "well_id": well["id"],
            "afe_id": afe["id"],
            "entry_date": "2026-08-25",
            "phase": "Drilling",
            "sub_activity_id": well_activity["id"],
            "services": [
                {
                    "afe_line_id": operation_line["id"],
                    "service_id": None,
                    "cost_code_id": cost_code["id"],
                    "service_hours": 12,
                    "rate_basis": "daily",
                    # The server must use the estimate rate, not this value.
                    "unit_rate": 1,
                }
            ],
            "consumables": [
                {
                    "afe_line_id": quantity_line["id"],
                    "consumable_id": None,
                    "cost_code_id": cost_code["id"],
                    "quantity": 2,
                    "unit_id": each["id"],
                    "unit_rate": 1,
                }
            ],
        },
    )
    assert Decimal(str(daily["total_daily_cost"])) == Decimal("600")
    assert Decimal(str(daily["services"][0]["unit_rate"])) == Decimal("1000")
    assert Decimal(str(daily["consumables"][0]["unit_rate"])) == Decimal("50")

    for report_type in (
        "afe_register",
        "afe_cost_estimate",
        "daily_cost",
        "cost_performance",
        "well_activities",
    ):
        generated = client.get(
            f"/api/v1/reports/generate?report_type={report_type}&well_id={well['id']}",
            headers=headers,
        )
        assert generated.status_code == 200, generated.text
        assert generated.json()["rows"], report_type

    performance = client.get(
        f"/api/v1/reports/generate?report_type=cost_performance&well_id={well['id']}",
        headers=headers,
    ).json()["rows"][0]
    assert Decimal(str(performance["estimate"])) == Decimal("2500")
    assert Decimal(str(performance["actual"])) == Decimal("600")

    report_export = client.get(
        f"/api/v1/reports/export?report_type=daily_cost&well_id={well['id']}",
        headers=headers,
    )
    assert report_export.status_code == 200
    assert load_workbook(BytesIO(report_export.content)).sheetnames == ["Report", "Summary"]

    audit_export = client.get("/api/v1/audit-logs/export", headers=headers)
    assert audit_export.status_code == 200
    assert load_workbook(BytesIO(audit_export.content)).sheetnames == ["Audit Log"]
