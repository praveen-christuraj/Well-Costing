"""Excel template generation tests."""

from datetime import datetime
from io import BytesIO

import pytest
from app.core.exceptions import BusinessValidationError, NotFoundError
from app.integrations.excel.reader import MAX_CONSECUTIVE_EMPTY_ROWS, ExcelReader
from app.integrations.excel.templates import ExcelTemplateService
from openpyxl import Workbook, load_workbook


def test_requirement_template_has_versioned_profile_headers() -> None:
    content = ExcelTemplateService().create_blank("requirement-items")
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    assert sheet is not None
    headers = [cell.value for cell in sheet[1]]
    assert headers[:6] == [
        "line_number",
        "catalog_item_code",
        "item_type",
        "cost_code",
        "quantity",
        "unit_code",
    ]
    assert sheet.freeze_panes == "A2"


def test_unknown_template_fails_clearly() -> None:
    with pytest.raises(NotFoundError, match="No import template exists"):
        ExcelTemplateService().create_blank("unknown")


def test_reader_rejects_unsupported_file_types() -> None:
    with pytest.raises(BusinessValidationError, match="Unsupported file extension"):
        ExcelReader().read(b"not-a-workbook", "requirements.pdf")


def _xlsx(rows: list[list[object]], *, stray_cell: tuple[int, int, object] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row_index, column_index, value)
    if stray_cell is not None:
        row_index, column_index, value = stray_cell
        sheet.cell(row_index, column_index, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_reader_skips_blank_rows_and_padded_used_range() -> None:
    content = _xlsx(
        [
            ["order_number", "title", None, None],
            ["SO-1", "Directional", None, None],
            [None, None, None, None],
            ["SO-2", "Mud logging", None, None],
        ],
        stray_cell=(MAX_CONSECUTIVE_EMPTY_ROWS + 20, 1, None),
    )

    workbook = ExcelReader().read(content, "service-orders.xlsx")

    assert workbook.columns == ["order_number", "title"]
    assert [row["order_number"] for row in workbook.rows] == ["SO-1", "SO-2"]


def test_reader_keeps_excel_datetimes() -> None:
    content = _xlsx(
        [
            ["order_number", "valid_from"],
            ["SO-1", datetime(2026, 1, 15)],
        ]
    )

    workbook = ExcelReader().read(content, "service-orders.xlsx")

    assert workbook.rows[0]["valid_from"] == datetime(2026, 1, 15)
