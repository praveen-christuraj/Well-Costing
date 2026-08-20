"""Safe tabular workbook reader for Excel and CSV uploads."""

import csv
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from itertools import zip_longest
from pathlib import Path
from typing import Any, cast

import xlrd
from openpyxl import load_workbook

from app.core.exceptions import BusinessValidationError

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
CSV_EXTENSIONS = {".csv"}
ALLOWED_EXTENSIONS = EXCEL_EXTENSIONS | CSV_EXTENSIONS
MAX_WORKBOOK_BYTES = 15 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
MAX_HEADER_COLUMNS = 200
MAX_CONSECUTIVE_EMPTY_ROWS = 200


@dataclass(frozen=True)
class WorkbookRows:
    columns: list[str]
    rows: list[dict[str, Any]]
    sheet_name: str


class ExcelReader:
    """Read one workbook sheet or CSV file into JSON-compatible row dictionaries."""

    def read(self, content: bytes, filename: str, sheet_name: str | None = None) -> WorkbookRows:
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise BusinessValidationError(
                f"Unsupported file extension '{extension}'. Use .xlsx, .xlsm, .xls, or .csv."
            )
        if len(content) > MAX_WORKBOOK_BYTES:
            raise BusinessValidationError("Upload exceeds the 15 MB limit")
        selected_sheet_name = sheet_name or "first sheet"
        try:
            if extension in CSV_EXTENSIONS:
                columns, records = self._read_csv(content)
            elif extension == ".xls":
                columns, records, selected_sheet_name = self._read_xls(content, sheet_name)
            else:
                columns, records, selected_sheet_name = self._read_xlsx_like(content, sheet_name)
        except BusinessValidationError:
            raise
        except Exception as exc:
            raise BusinessValidationError(
                "File could not be read", {"reason": str(exc)}
            ) from exc
        return WorkbookRows(
            columns=columns,
            rows=records,
            sheet_name=selected_sheet_name,
        )

    @staticmethod
    def _read_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
        """Parse CSV bytes, tolerating BOM and common delimiters."""

        text: str | None = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None or not text.strip():
            raise BusinessValidationError("CSV file is empty or uses an unsupported encoding")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

        reader = csv.reader(StringIO(text), delimiter=delimiter)
        rows = list(reader)
        if not rows:
            raise BusinessValidationError("CSV file is empty")

        columns = ExcelReader._headers_from_row(rows[0])
        return columns, ExcelReader._collect_records(columns, rows[1:])

    @staticmethod
    def _read_xlsx_like(
        content: bytes, sheet_name: str | None
    ) -> tuple[list[str], list[dict[str, Any]], str]:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise BusinessValidationError(
                        "Requested sheet not found", {"sheet_name": sheet_name}
                    )
                sheet = workbook[sheet_name]
                selected_sheet_name = sheet_name
            else:
                sheet = workbook[workbook.sheetnames[0]]
                selected_sheet_name = str(sheet.title)

            iterator = sheet.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if header_row is None:
                raise BusinessValidationError("Workbook sheet is empty")

            columns = ExcelReader._headers_from_row(header_row)
            records = ExcelReader._collect_records(columns, iterator)
            return columns, records, selected_sheet_name
        finally:
            workbook.close()

    @staticmethod
    def _read_xls(
        content: bytes, sheet_name: str | None
    ) -> tuple[list[str], list[dict[str, Any]], str]:
        workbook = xlrd.open_workbook(file_contents=content)
        if workbook.nsheets == 0:
            raise BusinessValidationError("Workbook has no sheets")

        if sheet_name:
            try:
                sheet = workbook.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError as exc:
                raise BusinessValidationError(
                    "Requested sheet not found", {"sheet_name": sheet_name}
                ) from exc
            selected_sheet_name = sheet_name
        else:
            sheet = workbook.sheet_by_index(0)
            selected_sheet_name = str(sheet.name)

        if sheet.nrows == 0:
            raise BusinessValidationError("Workbook sheet is empty")

        columns = ExcelReader._headers_from_row(sheet.row_values(0))
        data_rows = (sheet.row_values(row_index) for row_index in range(1, sheet.nrows))
        return columns, ExcelReader._collect_records(columns, data_rows), selected_sheet_name

    @staticmethod
    def _headers_from_row(header_row: Iterable[object]) -> list[str]:
        """Keep named headers only; drop the trailing empty used-range padding Excel adds."""

        columns: list[str] = []
        last_named = -1
        for index, column in enumerate(header_row):
            if index >= MAX_HEADER_COLUMNS:
                break
            label = "" if column is None else str(column).strip()
            columns.append(label)
            if label:
                last_named = index
        if last_named < 0:
            raise BusinessValidationError("Workbook sheet is empty")
        return columns[: last_named + 1]

    @staticmethod
    def _record_from_row(columns: list[str], row: object) -> dict[str, Any]:
        values: Sequence[object] = (
            cast(Sequence[object], row) if isinstance(row, (list, tuple)) else ()
        )
        return {
            key: ExcelReader._python_value(value)
            for key, value in zip_longest(columns, values, fillvalue=None)
            if key
        }

    @staticmethod
    def _collect_records(columns: list[str], rows: Iterable[object]) -> list[dict[str, Any]]:
        """Skip blank rows and stop before Excel's padded used-range can exhaust memory."""

        records: list[dict[str, Any]] = []
        empty_streak = 0
        for row in rows:
            record = ExcelReader._record_from_row(columns, row)
            if not any(value not in (None, "") for value in record.values()):
                empty_streak += 1
                if empty_streak >= MAX_CONSECUTIVE_EMPTY_ROWS:
                    break
                continue
            empty_streak = 0
            if len(records) >= MAX_IMPORT_ROWS:
                raise BusinessValidationError(
                    f"Workbook exceeds the {MAX_IMPORT_ROWS:,} data-row import limit"
                )
            records.append(record)
        return records

    @staticmethod
    def _python_value(value: object) -> object:
        if isinstance(value, datetime):
            return value
        if value == "":
            return None
        if hasattr(value, "to_pydatetime"):
            return value.to_pydatetime()  # type: ignore[union-attr]
        if hasattr(value, "item"):
            try:
                return value.item()  # type: ignore[union-attr]
            except (TypeError, ValueError):
                return value
        return value
