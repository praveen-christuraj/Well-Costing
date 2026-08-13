"""Excel exporter sharing Phase 2 template headers."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.exceptions import NotFoundError
from app.integrations.excel.mapper import PROFILE_REGISTRY


class ExcelExporter:
    def export(self, entity: str, rows: list[dict[str, Any]]) -> bytes:
        try:
            profile = PROFILE_REGISTRY[entity]
        except KeyError as exc:
            raise NotFoundError(f"No Excel export profile exists for '{entity}'") from exc
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = entity[:31]
        headers = profile.headers
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for row_index, row in enumerate(rows, start=2):
            for column, header in enumerate(headers, start=1):
                sheet.cell(row_index, column, self._value(row.get(header)))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(max(1, len(rows) + 1), len(headers)).coordinate}"
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _value(value: object) -> str | int | float | bool | None:
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return str(value)
