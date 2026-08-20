"""Blank, versioned Excel and CSV templates for cost-library entities."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.exceptions import NotFoundError
from app.integrations.excel.mapper import PROFILE_REGISTRY, MappingProfile


class ExcelTemplateService:
    @staticmethod
    def _profile(entity: str) -> MappingProfile:
        try:
            return PROFILE_REGISTRY[entity]
        except KeyError as exc:
            raise NotFoundError(f"No import template exists for '{entity}'") from exc

    def create_blank(self, entity: str) -> bytes:
        profile = self._profile(entity)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = entity[:31]
        for column, header in enumerate(profile.headers, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
            sheet.column_dimensions[cell.column_letter].width = max(16, len(header) + 3)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(profile.headers)).coordinate}"
        if "is_active" in profile.headers:
            column = profile.headers.index("is_active") + 1
            validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(  # pyright: ignore[reportUnknownMemberType]
                f"{sheet.cell(2, column).coordinate}:{sheet.cell(5000, column).coordinate}"
            )
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def create_blank_csv(self, entity: str) -> bytes:
        """UTF-8 (BOM) CSV header row so Excel opens it with correct columns."""

        profile = self._profile(entity)
        header_line = ",".join(profile.headers)
        return ("\ufeff" + header_line + "\r\n").encode("utf-8")
