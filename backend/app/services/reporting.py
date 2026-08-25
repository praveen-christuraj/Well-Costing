"""Reports over Master Data → AFE → AFE Cost Estimate → Daily Cost.

The retired Cost Builder estimate/version/snapshot and cost-control staging
models are intentionally absent from this read model.
"""

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.afe import Afe, AfeLine, Well
from app.models.afe_estimates import AfeCostEstimateLine
from app.models.categories import WellActivity
from app.models.daily_cost import DailyCostEntry
from app.models.user import User
from app.schemas.reporting import (
    GeneratedReport,
    ReportColumn,
    ReportFilters,
    ReportingContractRead,
    ReportingContractView,
    ReportSummary,
)
from app.services.afe_estimates import AfeEstimateService
from app.services.audit import log_entity_action

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)

REPORT_TITLES = {
    "afe_register": "AFE Register",
    "afe_cost_estimate": "AFE Cost Estimate Detail",
    "daily_cost": "Daily Cost Register",
    "cost_performance": "Cost Performance",
    "well_activities": "Well Activities & Cost Accountability",
}


class ReportingService:
    def __init__(self, session: Session, actor: User) -> None:
        self.session, self.actor = session, actor

    @staticmethod
    def contract() -> ReportingContractRead:
        """Describe the active reporting sources, not retired staging tables."""
        views = [
            ("master_data", "source", "Configured categories, codes, units and vendors"),
            ("afe", "source", "Projects, wells, AFE headers, sections, phases and lines"),
            ("afe_cost_estimate", "source", "Well-scoped rates for configured AFE lines"),
            ("daily_cost", "source", "Daily operational cost headers and charge lines"),
            ("well_activities", "source", "Well activity and responsible-party allocation"),
        ]
        return ReportingContractRead(
            contract_version="2.0",
            contract_status="active",
            schema_name="application_api",
            direct_grants_status="api_only",
            transactional_schema_public=False,
            views=[
                ReportingContractView(name=name, kind=kind, description=description)
                for name, kind, description in views
            ],
            pending_metrics=[],
        )

    def generate(self, filters: ReportFilters) -> GeneratedReport:
        builders = {
            "afe_register": self._afe_register,
            "afe_cost_estimate": self._afe_cost_estimate,
            "daily_cost": self._daily_cost,
            "cost_performance": self._cost_performance,
            "well_activities": self._well_activities,
        }
        columns, rows, summaries, description = builders[filters.report_type](filters)
        return GeneratedReport(
            report_type=filters.report_type,
            title=REPORT_TITLES[filters.report_type],
            description=description,
            generated_at=datetime.now(UTC),
            filters=filters,
            columns=columns,
            rows=rows,
            summaries=summaries,
        )

    def export(self, filters: ReportFilters) -> bytes:
        report = self.generate(filters)
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise RuntimeError("Workbook did not create a worksheet")
        sheet.title = "Report"
        sheet["A1"] = report.title.upper()
        sheet["A1"].font = TITLE_FONT
        sheet["A2"] = report.description
        sheet["A3"] = f"Generated {report.generated_at.strftime('%Y-%m-%d %H:%M')} UTC"

        header_row = 5
        for index, column in enumerate(report.columns, start=1):
            cell = sheet.cell(header_row, index, column.label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for row_index, row in enumerate(report.rows, start=header_row + 1):
            for column_index, column in enumerate(report.columns, start=1):
                value = row.get(column.key)
                cell = sheet.cell(row_index, column_index, value)
                if column.format == "money":
                    cell.number_format = "#,##0.00;[Red]-#,##0.00"
                elif column.format == "number":
                    cell.number_format = "#,##0.00"
        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = (
            f"A5:{get_column_letter(max(1, len(report.columns)))}{max(5, 5 + len(report.rows))}"
        )
        for index, column in enumerate(report.columns, start=1):
            width = max(12, min(38, len(column.label) + 4))
            for row in report.rows[:200]:
                width = max(width, min(38, len(str(row.get(column.key) or "")) + 2))
            sheet.column_dimensions[get_column_letter(index)].width = width

        summary = workbook.create_sheet("Summary")
        summary.append(["Measure", "Value"])
        for cell in summary[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for item in report.summaries:
            summary.append([item.label, item.value])
            if item.format == "money":
                summary.cell(summary.max_row, 2).number_format = "#,##0.00;[Red]-#,##0.00"
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 20

        stream = BytesIO()
        workbook.save(stream)
        content = stream.getvalue()
        log_entity_action(
            self.session,
            self.actor.id,
            "export",
            "report",
            entity_code=filters.report_type,
            details={
                "report_type": filters.report_type,
                "row_count": len(report.rows),
                "filters": filters.model_dump(mode="json"),
            },
        )
        self.session.commit()
        return content

    # ---------------------------------------------------------------- sources
    def _afes(self, filters: ReportFilters) -> list[Afe]:
        afes = list(
            self.session.scalars(
                select(Afe).where(Afe.is_active.is_(True)).order_by(Afe.created_at.desc())
            ).unique()
        )
        return [afe for afe in afes if self._matches_afe(afe, filters)]

    @staticmethod
    def _matches_afe(afe: Afe, filters: ReportFilters) -> bool:
        if filters.afe_id and afe.id != filters.afe_id:
            return False
        if filters.well_id and afe.well_id != filters.well_id:
            return False
        return not (filters.project_id and afe.well.project_id != filters.project_id)

    def _entries(self, filters: ReportFilters) -> list[DailyCostEntry]:
        entries = list(
            self.session.scalars(
                select(DailyCostEntry)
                .where(DailyCostEntry.is_active.is_(True))
                .order_by(DailyCostEntry.entry_date.desc())
            ).unique()
        )
        result: list[DailyCostEntry] = []
        for entry in entries:
            if filters.well_id and entry.well_id != filters.well_id:
                continue
            if filters.project_id and entry.well.project_id != filters.project_id:
                continue
            if filters.afe_id and entry.afe_id != filters.afe_id:
                continue
            if filters.date_from and entry.entry_date < filters.date_from:
                continue
            if filters.date_to and entry.entry_date > filters.date_to:
                continue
            result.append(entry)
        return result

    def _rate_map(self, afe_ids: set[UUID]) -> dict[UUID, AfeCostEstimateLine]:
        if not afe_ids:
            return {}
        rates = self.session.scalars(
            select(AfeCostEstimateLine).where(
                AfeCostEstimateLine.afe_id.in_(afe_ids),
                AfeCostEstimateLine.is_active.is_(True),
            )
        ).all()
        return {rate.afe_line_id: rate for rate in rates}

    @staticmethod
    def _amount(line: AfeLine, rates: dict[UUID, AfeCostEstimateLine]) -> Decimal:
        rate = rates.get(line.id)
        return AfeEstimateService.effective_quantity(line) * (
            Decimal(rate.unit_rate) if rate else Decimal("0")
        )

    @staticmethod
    def _column(key: str, label: str, format: str = "text") -> ReportColumn:
        return ReportColumn(key=key, label=label, format=format)  # type: ignore[arg-type]

    # --------------------------------------------------------------- builders
    def _afe_register(
        self, filters: ReportFilters
    ) -> tuple[list[ReportColumn], list[dict[str, Any]], list[ReportSummary], str]:
        afes = self._afes(filters)
        rates = self._rate_map({afe.id for afe in afes})
        rows: list[dict[str, Any]] = []
        total_budget = Decimal("0")
        total_estimate = Decimal("0")
        for afe in afes:
            active_lines = [line for line in afe.items if line.is_active]
            estimate = sum((self._amount(line, rates) for line in active_lines), Decimal("0"))
            budget = Decimal(afe.budget_amount or 0)
            total_budget += budget
            total_estimate += estimate
            rows.append(
                {
                    "project": afe.well.project.code,
                    "well": afe.well.code,
                    "afe": afe.code,
                    "title": afe.title,
                    "revision": afe.revision_number,
                    "status": afe.status,
                    "budget": budget,
                    "estimate": estimate,
                    "variance": budget - estimate,
                    "planned_days": afe.total_planned_days,
                    "line_count": len(active_lines),
                    "priced_lines": sum(
                        1
                        for line in active_lines
                        if rates.get(line.id) and rates[line.id].unit_rate > 0
                    ),
                }
            )
        columns = [
            self._column("project", "Project"),
            self._column("well", "Well"),
            self._column("afe", "AFE"),
            self._column("title", "Title"),
            self._column("revision", "Revision", "number"),
            self._column("status", "Status", "status"),
            self._column("budget", "AFE budget", "money"),
            self._column("estimate", "Cost estimate", "money"),
            self._column("variance", "Variance", "money"),
            self._column("planned_days", "Planned days", "number"),
            self._column("line_count", "Lines", "number"),
            self._column("priced_lines", "Priced lines", "number"),
        ]
        summaries = [
            ReportSummary(key="afes", label="AFEs", value=len(rows), format="number"),
            ReportSummary(
                key="budget", label="Total AFE budget", value=total_budget, format="money"
            ),
            ReportSummary(
                key="estimate", label="Total cost estimate", value=total_estimate, format="money"
            ),
            ReportSummary(
                key="variance",
                label="Budget less estimate",
                value=total_budget - total_estimate,
                format="money",
            ),
        ]
        return columns, rows, summaries, "AFE headers and priced totals from AFE Cost Estimates."

    def _afe_cost_estimate(
        self, filters: ReportFilters
    ) -> tuple[list[ReportColumn], list[dict[str, Any]], list[ReportSummary], str]:
        afes = self._afes(filters)
        rates = self._rate_map({afe.id for afe in afes})
        rows: list[dict[str, Any]] = []
        total = Decimal("0")
        for afe in afes:
            for line in afe.items:
                if not line.is_active:
                    continue
                rate = rates.get(line.id)
                secondary = line.secondary_category
                primary = secondary.primary_category if secondary else None
                amount = self._amount(line, rates)
                total += amount
                rows.append(
                    {
                        "project": afe.well.project.code,
                        "well": afe.well.code,
                        "afe": afe.code,
                        "line": line.line_number,
                        "primary_category": primary.name if primary else None,
                        "secondary_category": secondary.name if secondary else None,
                        "cost_code": line.cost_code.code if line.cost_code else None,
                        "section": "All sections"
                        if line.applies_to_all_sections
                        else (line.hole_section.code if line.hole_section else None),
                        "rate_basis": line.rate_basis.replace("_", " "),
                        "quantity": AfeEstimateService.effective_quantity(line),
                        "unit": line.unit.code if line.unit else None,
                        "unit_rate": Decimal(rate.unit_rate) if rate else Decimal("0"),
                        "amount": amount,
                        "vendor": rate.vendor.name if rate and rate.vendor else None,
                        "remarks": rate.remarks if rate else None,
                    }
                )
        columns = [
            self._column("project", "Project"),
            self._column("well", "Well"),
            self._column("afe", "AFE"),
            self._column("line", "Line", "number"),
            self._column("primary_category", "Primary category"),
            self._column("secondary_category", "Secondary category"),
            self._column("cost_code", "Cost code"),
            self._column("section", "Section"),
            self._column("rate_basis", "Rate basis"),
            self._column("quantity", "Quantity", "number"),
            self._column("unit", "Unit"),
            self._column("unit_rate", "Unit rate", "money"),
            self._column("amount", "Estimated amount", "money"),
            self._column("vendor", "Vendor"),
            self._column("remarks", "Remarks"),
        ]
        summaries = [
            ReportSummary(
                key="lines", label="AFE estimate lines", value=len(rows), format="number"
            ),
            ReportSummary(key="total", label="Estimated total", value=total, format="money"),
            ReportSummary(
                key="priced",
                label="Priced lines",
                value=sum(1 for row in rows if Decimal(row["unit_rate"] or 0) > 0),
                format="number",
            ),
        ]
        return (
            columns,
            rows,
            summaries,
            "Rates and amounts for classifications selected on each AFE line.",
        )

    def _daily_cost(
        self, filters: ReportFilters
    ) -> tuple[list[ReportColumn], list[dict[str, Any]], list[ReportSummary], str]:
        entries = self._entries(filters)
        rows = [
            {
                "date": entry.entry_date,
                "project": entry.well.project.code,
                "well": entry.well.code,
                "afe": entry.afe.code if entry.afe else None,
                "section": entry.hole_section.code if entry.hole_section else None,
                "phase": entry.phase,
                "activity": entry.sub_activity.activity.name if entry.sub_activity else None,
                "sub_activity": entry.sub_activity.name if entry.sub_activity else None,
                "responsible_party": entry.sub_activity.responsible_party
                if entry.sub_activity
                else None,
                "operational_charges": entry.total_services_cost,
                "quantity_charges": entry.total_consumables_cost,
                "daily_total": entry.total_daily_cost,
                "cumulative": entry.cumulative_cost,
                "depth": entry.current_depth,
                "progress": entry.daily_progress,
                "summary": entry.operational_summary,
            }
            for entry in entries
        ]
        total = sum((Decimal(entry.total_daily_cost) for entry in entries), Decimal("0"))
        columns = [
            self._column("date", "Date", "date"),
            self._column("project", "Project"),
            self._column("well", "Well"),
            self._column("afe", "AFE"),
            self._column("section", "Section"),
            self._column("phase", "Phase"),
            self._column("activity", "Activity"),
            self._column("sub_activity", "Sub-activity"),
            self._column("responsible_party", "Responsible party"),
            self._column("operational_charges", "Operational charges", "money"),
            self._column("quantity_charges", "Quantity charges", "money"),
            self._column("daily_total", "Daily total", "money"),
            self._column("cumulative", "Cumulative", "money"),
            self._column("depth", "Current depth", "number"),
            self._column("progress", "Daily progress", "number"),
            self._column("summary", "Operational summary"),
        ]
        summaries = [
            ReportSummary(key="days", label="Daily cost entries", value=len(rows), format="number"),
            ReportSummary(key="actual", label="Total daily cost", value=total, format="money"),
            ReportSummary(
                key="average",
                label="Average per entry",
                value=(total / len(rows) if rows else Decimal("0")),
                format="money",
            ),
        ]
        return columns, rows, summaries, "Actual operational costs recorded in Daily Cost."

    def _cost_performance(
        self, filters: ReportFilters
    ) -> tuple[list[ReportColumn], list[dict[str, Any]], list[ReportSummary], str]:
        wells = list(self.session.scalars(select(Well).where(Well.is_active.is_(True))).unique())
        wells = [well for well in wells if not filters.well_id or well.id == filters.well_id]
        wells = [
            well
            for well in wells
            if not filters.project_id or well.project_id == filters.project_id
        ]
        afes = self._afes(filters)
        afe_by_well: dict[UUID, Afe] = {}
        for afe in sorted(
            afes, key=lambda item: (item.status == "submitted", item.revision_number), reverse=True
        ):
            afe_by_well.setdefault(afe.well_id, afe)
        rates = self._rate_map({afe.id for afe in afes})
        entries = self._entries(filters)
        actual_by_well: dict[UUID, Decimal] = {}
        days_by_well: dict[UUID, set[Any]] = {}
        for entry in entries:
            actual_by_well[entry.well_id] = actual_by_well.get(
                entry.well_id, Decimal("0")
            ) + Decimal(entry.total_daily_cost)
            days_by_well.setdefault(entry.well_id, set()).add(entry.entry_date)
        rows: list[dict[str, Any]] = []
        for well in wells:
            afe = afe_by_well.get(well.id)
            if filters.afe_id and (afe is None or afe.id != filters.afe_id):
                continue
            budget = Decimal(afe.budget_amount or 0) if afe else Decimal("0")
            estimate = sum(
                (
                    self._amount(line, rates)
                    for line in (afe.items if afe else [])
                    if line.is_active
                ),
                Decimal("0"),
            )
            actual = actual_by_well.get(well.id, Decimal("0"))
            rows.append(
                {
                    "project": well.project.code,
                    "well": well.code,
                    "afe": afe.code if afe else None,
                    "status": afe.status if afe else "No AFE",
                    "budget": budget,
                    "estimate": estimate,
                    "actual": actual,
                    "budget_remaining": budget - actual,
                    "estimate_remaining": estimate - actual,
                    "budget_used_pct": (actual / budget * 100) if budget > 0 else Decimal("0"),
                    "days_logged": len(days_by_well.get(well.id, set())),
                }
            )
        total_budget = sum((Decimal(row["budget"]) for row in rows), Decimal("0"))
        total_actual = sum((Decimal(row["actual"]) for row in rows), Decimal("0"))
        columns = [
            self._column("project", "Project"),
            self._column("well", "Well"),
            self._column("afe", "AFE"),
            self._column("status", "AFE status", "status"),
            self._column("budget", "AFE budget", "money"),
            self._column("estimate", "AFE cost estimate", "money"),
            self._column("actual", "Daily Cost actual", "money"),
            self._column("budget_remaining", "Budget remaining", "money"),
            self._column("estimate_remaining", "Estimate remaining", "money"),
            self._column("budget_used_pct", "Budget used %", "number"),
            self._column("days_logged", "Days logged", "number"),
        ]
        summaries = [
            ReportSummary(key="wells", label="Wells", value=len(rows), format="number"),
            ReportSummary(
                key="budget", label="Total AFE budget", value=total_budget, format="money"
            ),
            ReportSummary(
                key="actual", label="Total Daily Cost actual", value=total_actual, format="money"
            ),
            ReportSummary(
                key="remaining",
                label="Budget remaining",
                value=total_budget - total_actual,
                format="money",
            ),
        ]
        return columns, rows, summaries, "AFE budget and estimate compared with Daily Cost actuals."

    def _well_activities(
        self, filters: ReportFilters
    ) -> tuple[list[ReportColumn], list[dict[str, Any]], list[ReportSummary], str]:
        activities = list(
            self.session.scalars(
                select(WellActivity)
                .where(WellActivity.is_active.is_(True))
                .order_by(WellActivity.name)
            ).unique()
        )
        well_map = {
            well.id: well
            for well in self.session.scalars(select(Well).where(Well.is_active.is_(True))).unique()
        }
        activities = [item for item in activities if item.well_id in well_map]
        activities = [
            item for item in activities if not filters.well_id or item.well_id == filters.well_id
        ]
        activities = [
            item
            for item in activities
            if not filters.project_id or well_map[item.well_id].project_id == filters.project_id
        ]
        entries = self._entries(filters)
        cost_by_activity: dict[UUID, Decimal] = {}
        entry_count: dict[UUID, set[UUID]] = {}
        for entry in entries:
            for line in [*entry.services, *entry.consumables]:
                activity_id = line.sub_activity_id or entry.sub_activity_id
                if activity_id is None:
                    continue
                cost_by_activity[activity_id] = cost_by_activity.get(
                    activity_id, Decimal("0")
                ) + Decimal(line.amount)
                entry_count.setdefault(activity_id, set()).add(entry.id)
        rows = [
            {
                "project": well_map[item.well_id].project.code,
                "well": well_map[item.well_id].code,
                "activity": item.activity.name if item.activity else None,
                "activity_code": item.activity.code if item.activity else None,
                "sub_activity": item.name,
                "responsible_party": item.responsible_party,
                "daily_entries": len(entry_count.get(item.id, set())),
                "actual_cost": cost_by_activity.get(item.id, Decimal("0")),
                "description": item.description,
            }
            for item in activities
        ]
        total = sum((Decimal(row["actual_cost"] or 0) for row in rows), Decimal("0"))
        columns = [
            self._column("project", "Project"),
            self._column("well", "Well"),
            self._column("activity", "Activity"),
            self._column("activity_code", "Activity code"),
            self._column("sub_activity", "Sub-activity"),
            self._column("responsible_party", "Responsible party"),
            self._column("daily_entries", "Daily entries", "number"),
            self._column("actual_cost", "Attributed actual cost", "money"),
            self._column("description", "Description"),
        ]
        summaries = [
            ReportSummary(
                key="activities",
                label="Configured well activities",
                value=len(rows),
                format="number",
            ),
            ReportSummary(key="actual", label="Attributed Daily Cost", value=total, format="money"),
        ]
        return (
            columns,
            rows,
            summaries,
            "Well Activities with Daily Cost attributed to the responsible party.",
        )
