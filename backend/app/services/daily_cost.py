"""Daily cost entry processing, rate calculations, and AFE comparative analytics."""

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessValidationError, ConflictError, NotFoundError
from app.models.afe import Afe, AfeLine, Well
from app.models.afe_estimates import AfeCostEstimateLine
from app.models.categories import WellActivity
from app.models.daily_cost import (
    DailyCostConsumableLine,
    DailyCostEntry,
    DailyCostServiceLine,
)
from app.schemas.daily_cost import (
    ComparisonBucket,
    ConsumableBreakdownItem,
    DailyCostAnalyticsRead,
    DailyCostComparisonRead,
    DailyCostConsumableLineRead,
    DailyCostEntryCreate,
    DailyCostEntryRead,
    DailyCostServiceLineRead,
    DailyTrendPoint,
    DateComparisonPoint,
    ReferenceConsumableRate,
    ReferenceServiceRate,
    ServiceBreakdownItem,
)
from app.services.afe_estimates import AfeEstimateService
from app.services.audit import log_entity_action

REPORT_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
REPORT_HEADER_FONT = Font(bold=True, color="FFFFFF")
REPORT_TITLE_FONT = Font(bold=True, size=14)
REPORT_LABEL_FONT = Font(bold=True)


class DailyCostService:
    def __init__(self, session: Session, actor_id: UUID) -> None:
        self.session, self.actor_id = session, actor_id

    @staticmethod
    def _entry_snapshot(
        entry: DailyCostEntry,
        *,
        service_line_ids: list[str] | None = None,
        consumable_line_ids: list[str] | None = None,
    ) -> dict[str, object]:
        """Return a serialisable before/after snapshot for the global audit log."""
        return {
            "entry_date": str(entry.entry_date),
            "sub_activity_id": str(entry.sub_activity_id) if entry.sub_activity_id else None,
            "phase": entry.phase,
            "hole_section_id": str(entry.hole_section_id) if entry.hole_section_id else None,
            "current_depth": str(entry.current_depth) if entry.current_depth is not None else None,
            "daily_progress": (
                str(entry.daily_progress) if entry.daily_progress is not None else None
            ),
            "operational_summary": entry.operational_summary,
            "total_services_cost": str(entry.total_services_cost),
            "total_consumables_cost": str(entry.total_consumables_cost),
            "total_daily_cost": str(entry.total_daily_cost),
            "service_line_ids": (
                service_line_ids
                if service_line_ids is not None
                else [str(line.id) for line in entry.services]
            ),
            "consumable_line_ids": (
                consumable_line_ids
                if consumable_line_ids is not None
                else [str(line.id) for line in entry.consumables]
            ),
        }

    def list_entries(
        self, well_id: UUID, *, include_deleted: bool = False
    ) -> list[DailyCostEntryRead]:
        """List active day logs, optionally including deleted logs for recovery."""
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")
        statement = select(DailyCostEntry).where(DailyCostEntry.well_id == well_id)
        if not include_deleted:
            statement = statement.where(DailyCostEntry.is_active.is_(True))
        entries = self.session.scalars(
            statement.order_by(DailyCostEntry.entry_date.desc(), DailyCostEntry.updated_at.desc())
        ).all()
        return [self._read_entry(e) for e in entries]

    def get_entry(self, well_id: UUID, entry_date: date) -> DailyCostEntryRead | None:
        # A well may have one saved entry per configured sub-activity on a date.
        # The date editor displays the most recently updated one rather than
        # failing with MultipleResultsFound when more than one is present.
        entry = self.session.scalars(
            select(DailyCostEntry)
            .where(
                DailyCostEntry.well_id == well_id,
                DailyCostEntry.entry_date == entry_date,
                DailyCostEntry.is_active.is_(True),
            )
            .order_by(DailyCostEntry.updated_at.desc())
        ).first()
        return self._read_entry(entry) if entry else None

    def save_entry(self, well_id: UUID, payload: DailyCostEntryCreate) -> DailyCostEntryRead:
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")

        # The day's activity type (Planned, NPT-1, UPA-1, …) is mandatory so
        # that every cost is accounted to Planned / NPT / UPA correctly.
        if payload.sub_activity_id is None:
            raise BusinessValidationError(
                "Select the day's activity type (Planned, NPT-1, UPA-1, …) before saving. "
                "If the list is empty, configure the Well Activities page for this well first."
            )
        sub_activity = self.session.get(WellActivity, payload.sub_activity_id)
        if not sub_activity or sub_activity.well_id != well_id or not sub_activity.is_active:
            raise BusinessValidationError(
                "The selected activity type is not configured for this well. "
                "Configure it on the Well Activities page first."
            )
        line_activity_ids = {
            s.sub_activity_id for s in payload.services if s.sub_activity_id is not None
        } | {c.sub_activity_id for c in payload.consumables if c.sub_activity_id is not None}
        if line_activity_ids:
            valid_ids = set(
                self.session.scalars(
                    select(WellActivity.id).where(
                        WellActivity.well_id == well_id,
                        WellActivity.id.in_(line_activity_ids),
                        WellActivity.is_active.is_(True),
                    )
                ).all()
            )
            if line_activity_ids - valid_ids:
                raise BusinessValidationError(
                    "A cost line references an activity type that is not configured for this well."
                )

        afe_id = payload.afe_id
        if not afe_id:
            # Pick active/submitted AFE for the well, or latest draft AFE
            active_afe = self.session.scalar(
                select(Afe)
                .where(Afe.well_id == well_id, Afe.is_active.is_(True))
                .order_by(Afe.status.desc(), Afe.revision_number.desc())
            )
            afe_id = active_afe.id if active_afe else None
        if afe_id is not None:
            selected_afe = self.session.get(Afe, afe_id)
            if (
                selected_afe is None
                or not selected_afe.is_active
                or selected_afe.well_id != well_id
            ):
                raise BusinessValidationError(
                    "The selected AFE is not active or does not belong to this well."
                )

        # The database allows one log per well/date/sub-activity. Match on the
        # activity as well; otherwise saving a second activity for the same day
        # would overwrite the first one or fail nondeterministically.
        entry = self.session.scalar(
            select(DailyCostEntry).where(
                DailyCostEntry.well_id == well_id,
                DailyCostEntry.entry_date == payload.entry_date,
                DailyCostEntry.sub_activity_id == payload.sub_activity_id,
            )
        )

        if entry is not None and not entry.is_active:
            raise BusinessValidationError(
                "This daily cost entry was deleted. Recover it from Deleted Day Logs "
                "before editing or saving it again."
            )

        previous_snapshot: dict[str, object] | None = None
        replaced_service_line_ids: list[str] = []
        replaced_consumable_line_ids: list[str] = []
        if not entry:
            created = True
            entry = DailyCostEntry(
                well_id=well_id,
                afe_id=afe_id,
                entry_date=payload.entry_date,
                hole_section_id=payload.hole_section_id,
                phase=payload.phase,
                sub_activity_id=payload.sub_activity_id,
                current_depth=payload.current_depth,
                daily_progress=payload.daily_progress,
                operational_summary=payload.operational_summary,
                created_by=self.actor_id,
                updated_by=self.actor_id,
            )
            self.session.add(entry)
            self.session.flush()
        else:
            created = False
            previous_snapshot = self._entry_snapshot(entry)
            replaced_service_line_ids = [str(line.id) for line in entry.services]
            replaced_consumable_line_ids = [str(line.id) for line in entry.consumables]
            entry.afe_id = afe_id
            entry.hole_section_id = payload.hole_section_id
            entry.phase = payload.phase
            entry.sub_activity_id = payload.sub_activity_id
            entry.current_depth = payload.current_depth
            entry.daily_progress = payload.daily_progress
            entry.operational_summary = payload.operational_summary
            entry.updated_by = self.actor_id

            # Clear existing lines for this entry
            self.session.execute(
                delete(DailyCostServiceLine).where(
                    DailyCostServiceLine.daily_cost_entry_id == entry.id
                )
            )
            self.session.execute(
                delete(DailyCostConsumableLine).where(
                    DailyCostConsumableLine.daily_cost_entry_id == entry.id
                )
            )
            self.session.flush()

        # Resolve every modern Daily Cost line back to the selected AFE line and
        # its saved estimate rate. Client-supplied rates are accepted only for
        # historical catalogue-only rows; current rows cannot bypass the AFE
        # Cost Estimate source of truth.
        source_ids = {
            line.afe_line_id
            for line in [*payload.services, *payload.consumables]
            if line.afe_line_id is not None
        }
        source_lines = (
            {
                line.id: line
                for line in self.session.scalars(
                    select(AfeLine).where(AfeLine.id.in_(source_ids), AfeLine.is_active.is_(True))
                ).all()
            }
            if source_ids
            else {}
        )
        if source_ids - set(source_lines):
            raise BusinessValidationError(
                "A Daily Cost line references an inactive or missing AFE line."
            )
        if any(line.afe_id != afe_id for line in source_lines.values()):
            raise BusinessValidationError(
                "Every Daily Cost line must belong to the governing AFE for this well."
            )
        estimate_rates = (
            {
                rate.afe_line_id: rate
                for rate in self.session.scalars(
                    select(AfeCostEstimateLine).where(
                        AfeCostEstimateLine.afe_line_id.in_(source_ids),
                        AfeCostEstimateLine.is_active.is_(True),
                    )
                ).all()
            }
            if source_ids
            else {}
        )

        total_services = Decimal("0")
        for s_input in payload.services:
            hours = Decimal(str(s_input.service_hours))
            if hours < 0 or hours > 24:
                raise BusinessValidationError("Operational charge hours must be between 0 and 24")
            operating_days = hours / Decimal("24.0")
            source = source_lines.get(s_input.afe_line_id) if s_input.afe_line_id else None
            if source is not None and source.rate_basis in {"per_unit", "daily_consumption"}:
                raise BusinessValidationError(
                    "Per-unit and daily-consumption AFE lines must be entered as quantity charges."
                )
            estimate_rate = estimate_rates.get(s_input.afe_line_id) if s_input.afe_line_id else None
            base_rate = (
                Decimal(estimate_rate.unit_rate)
                if estimate_rate is not None
                else (Decimal("0") if source is not None else Decimal(str(s_input.unit_rate)))
            )
            rate_basis = source.rate_basis if source is not None else s_input.rate_basis
            override = s_input.override_rate
            # Override rate takes precedence if present and positive
            effective_rate = (
                Decimal(str(override))
                if override is not None and Decimal(str(override)) > 0
                else base_rate
            )
            if rate_basis == "daily":
                line_amount = operating_days * effective_rate
            else:
                line_amount = effective_rate
            total_services += line_amount

            s_line = DailyCostServiceLine(
                daily_cost_entry_id=entry.id,
                afe_line_id=s_input.afe_line_id,
                service_id=s_input.service_id,
                cost_code_id=source.cost_code_id if source is not None else s_input.cost_code_id,
                vendor_id=(
                    estimate_rate.vendor_id
                    if estimate_rate is not None and estimate_rate.vendor_id is not None
                    else s_input.vendor_id
                ),
                hole_section_id=s_input.hole_section_id or payload.hole_section_id,
                sub_activity_id=s_input.sub_activity_id,
                service_type=s_input.service_type,
                service_hours=hours,
                operating_days=operating_days,
                rate_basis=rate_basis,
                unit_rate=base_rate,
                override_rate=Decimal(str(override)) if override is not None else None,
                amount=line_amount,
                remarks=s_input.remarks,
                created_by=self.actor_id,
                updated_by=self.actor_id,
            )
            self.session.add(s_line)

        total_consumables = Decimal("0")
        for c_input in payload.consumables:
            qty = Decimal(str(c_input.quantity))
            source = source_lines.get(c_input.afe_line_id) if c_input.afe_line_id else None
            if source is not None and source.rate_basis not in {"per_unit", "daily_consumption"}:
                raise BusinessValidationError(
                    "Daily, section, service and fixed AFE lines must be entered "
                    "as operational charges."
                )
            estimate_rate = estimate_rates.get(c_input.afe_line_id) if c_input.afe_line_id else None
            base_rate = (
                Decimal(estimate_rate.unit_rate)
                if estimate_rate is not None
                else (Decimal("0") if source is not None else Decimal(str(c_input.unit_rate)))
            )
            override = c_input.override_rate
            effective_rate = (
                Decimal(str(override))
                if override is not None and Decimal(str(override)) > 0
                else base_rate
            )
            line_amount = qty * effective_rate
            total_consumables += line_amount

            c_line = DailyCostConsumableLine(
                daily_cost_entry_id=entry.id,
                afe_line_id=c_input.afe_line_id,
                consumable_id=c_input.consumable_id,
                cost_code_id=source.cost_code_id if source is not None else c_input.cost_code_id,
                vendor_id=(
                    estimate_rate.vendor_id
                    if estimate_rate is not None and estimate_rate.vendor_id is not None
                    else c_input.vendor_id
                ),
                sub_activity_id=c_input.sub_activity_id,
                quantity=qty,
                unit_id=source.unit_id if source is not None else c_input.unit_id,
                unit_rate=base_rate,
                override_rate=Decimal(str(override)) if override is not None else None,
                amount=line_amount,
                remarks=c_input.remarks,
                created_by=self.actor_id,
                updated_by=self.actor_id,
            )
            self.session.add(c_line)

        entry.total_services_cost = total_services
        entry.total_consumables_cost = total_consumables
        entry.total_daily_cost = total_services + total_consumables
        self.session.flush()
        current_service_line_ids = [
            str(line_id)
            for line_id in self.session.scalars(
                select(DailyCostServiceLine.id).where(
                    DailyCostServiceLine.daily_cost_entry_id == entry.id
                )
            ).all()
        ]
        current_consumable_line_ids = [
            str(line_id)
            for line_id in self.session.scalars(
                select(DailyCostConsumableLine.id).where(
                    DailyCostConsumableLine.daily_cost_entry_id == entry.id
                )
            ).all()
        ]

        log_entity_action(
            self.session,
            self.actor_id,
            "create" if created else "update",
            "daily_cost_entry",
            entity_id=entry.id,
            entity_code=str(entry.entry_date),
            details={
                "well_id": str(well_id),
                "before": previous_snapshot,
                "after": self._entry_snapshot(
                    entry,
                    service_line_ids=current_service_line_ids,
                    consumable_line_ids=current_consumable_line_ids,
                ),
                "replaced_service_line_ids": replaced_service_line_ids,
                "replaced_consumable_line_ids": replaced_consumable_line_ids,
            },
        )
        self._recompute_cumulative_costs(well_id)
        self.session.commit()
        self.session.refresh(entry)
        return self._read_entry(entry)

    def delete_entry(self, well_id: UUID, entry_id: UUID) -> None:
        """Soft-delete a daily cost entry and recompute the cumulative costs."""
        entry = self.session.get(DailyCostEntry, entry_id)
        if not entry or entry.well_id != well_id:
            raise NotFoundError("Daily cost entry not found")
        if not entry.is_active:
            raise BusinessValidationError("Daily cost entry is already deleted")
        snapshot = self._entry_snapshot(entry)
        entry.is_active = False
        entry.updated_by = self.actor_id
        self.session.flush()
        log_entity_action(
            self.session,
            self.actor_id,
            "soft_delete",
            "daily_cost_entry",
            entity_id=entry.id,
            entity_code=str(entry.entry_date),
            details={"well_id": str(well_id), "before": snapshot},
        )
        self._recompute_cumulative_costs(well_id)
        self.session.commit()

    def recover_entry(self, well_id: UUID, entry_id: UUID) -> DailyCostEntryRead:
        """Recover a soft-deleted daily cost entry."""
        entry = self.session.get(DailyCostEntry, entry_id)
        if not entry or entry.well_id != well_id:
            raise NotFoundError("Daily cost entry not found")
        if entry.is_active:
            raise BusinessValidationError("Daily cost entry is not deleted and cannot be recovered")
        existing = self.session.scalar(
            select(DailyCostEntry).where(
                DailyCostEntry.well_id == well_id,
                DailyCostEntry.entry_date == entry.entry_date,
                DailyCostEntry.sub_activity_id == entry.sub_activity_id,
                DailyCostEntry.is_active.is_(True),
                DailyCostEntry.id != entry.id,
            )
        )
        if existing:
            raise ConflictError(
                "An active entry for this date already exists. Delete it before recovering."
            )
        entry.is_active = True
        entry.updated_by = self.actor_id
        self.session.flush()
        log_entity_action(
            self.session,
            self.actor_id,
            "recover",
            "daily_cost_entry",
            entity_id=entry.id,
            entity_code=str(entry.entry_date),
            details={"well_id": str(well_id), "after": self._entry_snapshot(entry)},
        )
        self._recompute_cumulative_costs(well_id)
        self.session.commit()
        self.session.refresh(entry)
        return self._read_entry(entry)

    def _active_afe(self, well_id: UUID) -> Afe | None:
        """The well's governing AFE: submitted preferred, then latest revision."""
        return self.session.scalar(
            select(Afe)
            .where(Afe.well_id == well_id, Afe.is_active.is_(True))
            .order_by(Afe.status.desc(), Afe.revision_number.desc())
        )

    def get_reference_rates(self, well_id: UUID) -> dict[str, Any]:
        """Return Daily Cost choices from the governing AFE Cost Estimate.

        The user-selected rate basis controls which entry grid and calculation
        applies: per-unit/daily-consumption lines are quantity charges; all
        other bases are operational/time or fixed charges. No item type is
        inferred from category names.
        """
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")

        afe = self._active_afe(well_id)
        if afe is None:
            return {
                "afe_id": None,
                "afe_code": None,
                "afe_title": None,
                "rates_source": "afe_cost_estimate",
                "priced_line_count": 0,
                "unpriced_line_count": 0,
                "services": [],
                "consumables": [],
            }

        rate_rows = {
            row.afe_line_id: row
            for row in self.session.scalars(
                select(AfeCostEstimateLine).where(
                    AfeCostEstimateLine.afe_id == afe.id,
                    AfeCostEstimateLine.is_active.is_(True),
                )
            )
        }
        operational: list[ReferenceServiceRate] = []
        quantity: list[ReferenceConsumableRate] = []
        priced = 0
        unpriced = 0

        for line in afe.items:
            if not line.is_active:
                continue
            rate_row = rate_rows.get(line.id)
            unit_rate = Decimal(rate_row.unit_rate) if rate_row else Decimal("0")
            priced += int(unit_rate > 0)
            unpriced += int(unit_rate <= 0)
            secondary = line.secondary_category
            primary = secondary.primary_category if secondary else None
            catalog = line.catalog_item
            code = (
                catalog.code
                if catalog
                else (secondary.code if secondary else f"LINE-{line.line_number}")
            )
            name = (
                catalog.name
                if catalog
                else (secondary.name if secondary else f"AFE line {line.line_number}")
            )
            if line.rate_basis in {"per_unit", "daily_consumption"}:
                quantity.append(
                    ReferenceConsumableRate(
                        afe_line_id=line.id,
                        consumable_id=catalog.id if catalog else None,
                        consumable_code=code,
                        consumable_name=name,
                        primary_category_name=primary.name if primary else None,
                        secondary_category_name=secondary.name if secondary else None,
                        cost_code_id=line.cost_code_id,
                        cost_code=line.cost_code.code if line.cost_code else "UNKNOWN",
                        vendor_id=rate_row.vendor_id if rate_row else None,
                        vendor_name=(
                            rate_row.vendor.name if rate_row and rate_row.vendor else None
                        ),
                        rate_basis=line.rate_basis,
                        unit_id=line.unit_id,
                        unit_code=line.unit.code if line.unit else "EA",
                        unit_rate=unit_rate,
                    )
                )
            else:
                operational.append(
                    ReferenceServiceRate(
                        afe_line_id=line.id,
                        service_id=catalog.id if catalog else None,
                        service_code=code,
                        service_name=name,
                        primary_category_name=primary.name if primary else None,
                        secondary_category_name=secondary.name if secondary else None,
                        cost_code_id=line.cost_code_id,
                        cost_code=line.cost_code.code if line.cost_code else "UNKNOWN",
                        vendor_id=rate_row.vendor_id if rate_row else None,
                        vendor_name=(
                            rate_row.vendor.name if rate_row and rate_row.vendor else None
                        ),
                        rate_basis=line.rate_basis,
                        unit_id=line.unit_id,
                        unit_code=line.unit.code if line.unit else "EA",
                        operating_rate=unit_rate,
                    )
                )

        return {
            "afe_id": str(afe.id),
            "afe_code": afe.code,
            "afe_title": afe.title,
            "rates_source": "afe_cost_estimate",
            "priced_line_count": priced,
            "unpriced_line_count": unpriced,
            "services": [item.model_dump() for item in operational],
            "consumables": [item.model_dump() for item in quantity],
        }

    def get_analytics(self, well_id: UUID) -> DailyCostAnalyticsRead:
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")

        afe = self.session.scalar(
            select(Afe)
            .where(Afe.well_id == well_id, Afe.is_active.is_(True))
            .order_by(Afe.status.desc(), Afe.revision_number.desc())
        )

        afe_budget = Decimal("0")
        total_planned_days = Decimal("0")
        if afe:
            afe_budget = afe.budget_amount if afe.budget_amount > 0 else Decimal("0")
            total_planned_days = (
                afe.total_planned_days if afe.total_planned_days > 0 else Decimal("0")
            )

        entries = self.session.scalars(
            select(DailyCostEntry)
            .where(DailyCostEntry.well_id == well_id, DailyCostEntry.is_active.is_(True))
            .order_by(DailyCostEntry.entry_date.asc())
        ).all()

        cumulative_cost = Decimal("0")
        trend_all: list[DailyTrendPoint] = []
        services_dict: dict[UUID, dict[str, Any]] = {}
        consumables_dict: dict[UUID, dict[str, Any]] = {}

        for e in entries:
            cumulative_cost += e.total_daily_cost
            trend_all.append(
                DailyTrendPoint(
                    entry_date=e.entry_date,
                    daily_cost=e.total_daily_cost,
                    cumulative_cost=cumulative_cost,
                    services_cost=e.total_services_cost,
                    consumables_cost=e.total_consumables_cost,
                    phase=e.phase,
                    current_depth=e.current_depth,
                )
            )
            for service_line in e.services:
                key = service_line.afe_line_id or service_line.service_id
                if key is None:
                    continue
                afe_line = service_line.afe_line
                secondary = afe_line.secondary_category if afe_line else None
                if key not in services_dict:
                    services_dict[key] = {
                        "service_id": key,
                        "service_code": (
                            service_line.service.code
                            if service_line.service
                            else (secondary.code if secondary else "AFE-LINE")
                        ),
                        "service_name": (
                            service_line.service.name
                            if service_line.service
                            else (secondary.name if secondary else "Configured AFE charge")
                        ),
                        "total_hours": Decimal("0"),
                        "total_days": Decimal("0"),
                        "total_cost": Decimal("0"),
                    }
                services_dict[key]["total_hours"] += service_line.service_hours
                services_dict[key]["total_days"] += service_line.operating_days
                services_dict[key]["total_cost"] += service_line.amount

            for quantity_line in e.consumables:
                key = quantity_line.afe_line_id or quantity_line.consumable_id
                if key is None:
                    continue
                afe_line = quantity_line.afe_line
                secondary = afe_line.secondary_category if afe_line else None
                if key not in consumables_dict:
                    consumables_dict[key] = {
                        "consumable_id": key,
                        "consumable_code": (
                            quantity_line.consumable.code
                            if quantity_line.consumable
                            else (secondary.code if secondary else "AFE-LINE")
                        ),
                        "consumable_name": (
                            quantity_line.consumable.name
                            if quantity_line.consumable
                            else (secondary.name if secondary else "Configured AFE quantity charge")
                        ),
                        "unit_code": quantity_line.unit.code if quantity_line.unit else "UOM",
                        "total_quantity": Decimal("0"),
                        "total_cost": Decimal("0"),
                    }
                consumables_dict[key]["total_quantity"] += quantity_line.quantity
                consumables_dict[key]["total_cost"] += quantity_line.amount

        days_elapsed = len(entries)
        burn_rate = (
            (cumulative_cost / Decimal(str(days_elapsed))) if days_elapsed > 0 else Decimal("0")
        )
        remaining_days = max(Decimal("0"), total_planned_days - Decimal(str(days_elapsed)))
        forecast_cost = cumulative_cost + (remaining_days * burn_rate)
        balance = afe_budget - cumulative_cost
        variance = afe_budget - forecast_cost

        # Services breakdown list with percentage
        tot_svc = sum((item["total_cost"] for item in services_dict.values()), Decimal("0"))
        services_breakdown = [
            ServiceBreakdownItem(
                service_id=v["service_id"],
                service_code=v["service_code"],
                service_name=v["service_name"],
                total_hours=v["total_hours"],
                total_days=v["total_days"],
                total_cost=v["total_cost"],
                percentage=((v["total_cost"] / tot_svc) * 100) if tot_svc > 0 else Decimal("0"),
            )
            for v in sorted(services_dict.values(), key=lambda x: x["total_cost"], reverse=True)
        ]

        # Consumables breakdown list with percentage
        tot_con = sum((item["total_cost"] for item in consumables_dict.values()), Decimal("0"))
        consumables_breakdown = [
            ConsumableBreakdownItem(
                consumable_id=v["consumable_id"],
                consumable_code=v["consumable_code"],
                consumable_name=v["consumable_name"],
                unit_code=v["unit_code"],
                total_quantity=v["total_quantity"],
                total_cost=v["total_cost"],
                percentage=((v["total_cost"] / tot_con) * 100) if tot_con > 0 else Decimal("0"),
            )
            for v in sorted(consumables_dict.values(), key=lambda x: x["total_cost"], reverse=True)
        ]

        return DailyCostAnalyticsRead(
            well_id=well_id,
            well_code=well.code,
            afe_id=afe.id if afe else None,
            afe_code=afe.code if afe else None,
            afe_budget=afe_budget,
            total_planned_days=total_planned_days,
            cumulative_actual_cost=cumulative_cost,
            balance_amount=balance,
            days_elapsed=days_elapsed,
            burn_rate_daily_avg=burn_rate,
            remaining_planned_days=remaining_days,
            forecast_at_end_of_well=forecast_cost,
            variance_to_afe=variance,
            trend_last_5_days=trend_all[-5:],
            trend_last_7_days=trend_all[-7:],
            trend_all_days=trend_all,
            services_breakdown=services_breakdown,
            consumables_breakdown=consumables_breakdown,
        )

    # ------------------------------------------------------------- comparison
    def get_comparison(self, well_id: UUID) -> DailyCostComparisonRead:
        """Well-scoped planned-versus-actual comparison across every dimension.

        Planned figures come from the AFE (budget, planned days, section/phase
        plan) and the AFE Cost Estimates (priced AFE lines). Actual figures
        come from the saved daily cost entries.
        """
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")

        afe = self._active_afe(well_id)

        estimate_total = Decimal("0")
        planned_by_section: dict[str, Decimal] = {}
        planned_days_by_section: dict[str, Decimal] = {}
        planned_days_by_phase: dict[str, Decimal] = {}
        afe_budget = Decimal("0")
        planned_days = Decimal("0")

        if afe:
            afe_budget = Decimal(afe.budget_amount or 0)
            planned_days = Decimal(afe.total_planned_days or 0)
            rate_rows = {
                row.afe_line_id: row
                for row in self.session.scalars(
                    select(AfeCostEstimateLine).where(
                        AfeCostEstimateLine.afe_id == afe.id,
                        AfeCostEstimateLine.is_active.is_(True),
                    )
                )
            }
            for line in afe.items:
                if not line.is_active:
                    continue
                rate_row = rate_rows.get(line.id)
                amount = AfeEstimateService.effective_quantity(line) * (
                    Decimal(rate_row.unit_rate) if rate_row else Decimal("0")
                )
                estimate_total += amount
                section_key = (
                    "All sections"
                    if line.applies_to_all_sections
                    else (line.hole_section.code if line.hole_section else "Unassigned")
                )
                planned_by_section[section_key] = (
                    planned_by_section.get(section_key, Decimal("0")) + amount
                )
            for section in afe.sections:
                if not section.is_active:
                    continue
                section_key = section.hole_section.code if section.hole_section else "Unassigned"
                planned_days_by_section[section_key] = planned_days_by_section.get(
                    section_key, Decimal("0")
                ) + Decimal(section.planned_days or 0)
                # Roll the section's phase plan: prefer the child phase rows,
                # falling back to the legacy single phase value.
                phase_plan = (
                    [(ph.phase, ph.planned_days) for ph in (section.phases or []) if ph.is_active]
                    if (section.phases or [])
                    else [(section.phase, section.planned_days)]
                )
                for phase_name, phase_days in phase_plan:
                    planned_days_by_phase[phase_name or "Unassigned"] = planned_days_by_phase.get(
                        phase_name or "Unassigned", Decimal("0")
                    ) + Decimal(phase_days or 0)

        entries = self.session.scalars(
            select(DailyCostEntry)
            .where(DailyCostEntry.well_id == well_id, DailyCostEntry.is_active.is_(True))
            .order_by(DailyCostEntry.entry_date.asc())
        ).all()

        planned_daily = (afe_budget / planned_days) if planned_days > 0 and afe_budget > 0 else None

        by_date: list[DateComparisonPoint] = []
        weeks: dict[str, ComparisonBucket] = {}
        months: dict[str, ComparisonBucket] = {}
        sections: dict[str, ComparisonBucket] = {}
        phases: dict[str, ComparisonBucket] = {}
        activities: dict[str, ComparisonBucket] = {}
        sub_activities: dict[str, ComparisonBucket] = {}

        def bucket(store: dict[str, ComparisonBucket], key: str, label: str) -> ComparisonBucket:
            item = store.get(key)
            if item is None:
                item = ComparisonBucket(key=key, label=label)
                store[key] = item
            return item

        def add_amount(item: ComparisonBucket, services: Decimal, consumables: Decimal) -> None:
            item.services_cost += services
            item.consumables_cost += consumables
            item.total_cost += services + consumables

        cumulative = Decimal("0")
        for day_number, entry in enumerate(entries, start=1):
            cumulative += entry.total_daily_cost
            by_date.append(
                DateComparisonPoint(
                    entry_date=entry.entry_date,
                    day_number=day_number,
                    phase=entry.phase,
                    hole_section_code=(entry.hole_section.code if entry.hole_section else None),
                    activity_name=(entry.sub_activity.name if entry.sub_activity else None),
                    services_cost=entry.total_services_cost,
                    consumables_cost=entry.total_consumables_cost,
                    daily_cost=entry.total_daily_cost,
                    cumulative_cost=cumulative,
                    planned_cumulative=(
                        planned_daily * day_number if planned_daily is not None else None
                    ),
                    current_depth=entry.current_depth,
                    daily_progress=entry.daily_progress,
                )
            )

            iso = entry.entry_date.isocalendar()
            week_key = f"{iso[0]}-W{iso[1]:02d}"
            week = bucket(weeks, week_key, f"Week {iso[1]:02d}, {iso[0]}")
            week.entry_count += 1
            add_amount(week, entry.total_services_cost, entry.total_consumables_cost)

            month_key = entry.entry_date.strftime("%Y-%m")
            month = bucket(months, month_key, entry.entry_date.strftime("%B %Y"))
            month.entry_count += 1
            add_amount(month, entry.total_services_cost, entry.total_consumables_cost)

            phase_key = entry.phase or "Unassigned"
            phase = bucket(phases, phase_key, phase_key)
            phase.entry_count += 1
            phase.actual_days = (phase.actual_days or Decimal("0")) + Decimal("1")
            add_amount(phase, entry.total_services_cost, entry.total_consumables_cost)

            entry_section_key = entry.hole_section.code if entry.hole_section else "Unassigned"
            entry_sub = entry.sub_activity

            for line in entry.services or []:
                line_section_key = (
                    line.hole_section.code if line.hole_section else entry_section_key
                )
                section = bucket(sections, line_section_key, line_section_key)
                add_amount(section, Decimal(line.amount), Decimal("0"))
                self._attribute_activity(
                    activities,
                    sub_activities,
                    line.sub_activity or entry_sub,
                    Decimal(line.amount),
                    Decimal("0"),
                    bucket,
                    add_amount,
                )
            for line in entry.consumables or []:
                section = bucket(sections, entry_section_key, entry_section_key)
                add_amount(section, Decimal("0"), Decimal(line.amount))
                self._attribute_activity(
                    activities,
                    sub_activities,
                    line.sub_activity or entry_sub,
                    Decimal("0"),
                    Decimal(line.amount),
                    bucket,
                    add_amount,
                )

        # Sections planned but not yet spent still appear in the comparison.
        for section_key in set(planned_by_section) | set(planned_days_by_section):
            bucket(sections, section_key, section_key)
        for section_key, item in sections.items():
            planned = planned_by_section.get(section_key)
            item.planned_cost = planned
            item.planned_days = planned_days_by_section.get(section_key)
            if planned is not None:
                item.variance = planned - item.total_cost

        for phase_key in planned_days_by_phase:
            bucket(phases, phase_key, phase_key)
        for phase_key, item in phases.items():
            item.planned_days = planned_days_by_phase.get(phase_key)

        return DailyCostComparisonRead(
            well_id=well_id,
            well_code=well.code,
            well_name=well.name,
            afe_id=afe.id if afe else None,
            afe_code=afe.code if afe else None,
            afe_title=afe.title if afe else None,
            afe_budget=afe_budget,
            estimate_total=estimate_total,
            cumulative_actual_cost=cumulative,
            variance_to_budget=afe_budget - cumulative,
            variance_to_estimate=estimate_total - cumulative,
            total_planned_days=planned_days,
            days_elapsed=len(entries),
            by_date=by_date,
            by_week=[weeks[key] for key in sorted(weeks)],
            by_month=[months[key] for key in sorted(months)],
            by_section=sorted(sections.values(), key=lambda b: b.total_cost, reverse=True),
            by_phase=sorted(phases.values(), key=lambda b: b.total_cost, reverse=True),
            by_activity=sorted(activities.values(), key=lambda b: b.total_cost, reverse=True),
            by_sub_activity=sorted(
                sub_activities.values(), key=lambda b: b.total_cost, reverse=True
            ),
        )

    @staticmethod
    def _attribute_activity(
        activities: dict[str, ComparisonBucket],
        sub_activities: dict[str, ComparisonBucket],
        sub_activity: WellActivity | None,
        services: Decimal,
        consumables: Decimal,
        bucket: Any,
        add_amount: Any,
    ) -> None:
        if sub_activity is None:
            unassigned = bucket(activities, "UNASSIGNED", "Unassigned")
            add_amount(unassigned, services, consumables)
            return
        activity = sub_activity.activity
        activity_key = activity.code if activity else "UNASSIGNED"
        activity_bucket = bucket(
            activities, activity_key, activity.name if activity else "Unassigned"
        )
        activity_bucket.activity_code = activity_key
        activity_bucket.activity_name = activity.name if activity else None
        add_amount(activity_bucket, services, consumables)

        sub_bucket = bucket(sub_activities, sub_activity.name, sub_activity.name)
        sub_bucket.activity_code = activity_key
        sub_bucket.activity_name = activity.name if activity else None
        sub_bucket.responsible_party = sub_activity.responsible_party
        add_amount(sub_bucket, services, consumables)

    # ---------------------------------------------------------------- reports
    def export_day_report(self, well_id: UUID, entry_date: date) -> bytes:
        """Printable daily cost report for one operational day."""
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")
        entry = self.session.scalar(
            select(DailyCostEntry).where(
                DailyCostEntry.well_id == well_id,
                DailyCostEntry.entry_date == entry_date,
                DailyCostEntry.is_active.is_(True),
            )
        )
        if entry is None:
            raise NotFoundError(f"No daily cost entry exists for {entry_date}")

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Daily Cost Report"

        sheet["A1"] = "DAILY COST REPORT"
        sheet["A1"].font = REPORT_TITLE_FONT

        header_pairs = [
            ("Project", well.project.code if well.project else ""),
            ("Well", f"{well.code} — {well.name}"),
            ("Rig", well.rig_name or ""),
            ("AFE", entry.afe.code if entry.afe else ""),
            ("Report date", str(entry.entry_date)),
            ("Phase", entry.phase or ""),
            ("Hole section", entry.hole_section.code if entry.hole_section else ""),
            ("Activity type", entry.sub_activity.name if entry.sub_activity else ""),
            (
                "Current depth",
                float(entry.current_depth) if entry.current_depth is not None else "",
            ),
            (
                "Daily progress",
                float(entry.daily_progress) if entry.daily_progress is not None else "",
            ),
            ("Operational summary", entry.operational_summary or ""),
        ]
        row_index = 3
        for label, value in header_pairs:
            sheet.cell(row_index, 1, label).font = REPORT_LABEL_FONT
            sheet.cell(row_index, 2, value)
            row_index += 1

        row_index += 1
        sheet.cell(row_index, 1, "Operational / time charges").font = REPORT_TITLE_FONT
        row_index = self._write_table(
            sheet,
            row_index + 1,
            [
                "AFE charge",
                "Type",
                "Hours",
                "Days",
                "Rate basis",
                "Unit rate",
                "Override",
                "Amount",
                "Section",
                "Activity",
                "Remarks",
            ],
            [
                [
                    (
                        f"{s.service.code} — {s.service.name}"
                        if s.service
                        else (
                            f"{s.afe_line.secondary_category.code} — "
                            f"{s.afe_line.secondary_category.name}"
                            if s.afe_line
                            else ""
                        )
                    ),
                    s.service_type,
                    float(s.service_hours),
                    float(s.operating_days),
                    s.rate_basis,
                    float(s.unit_rate),
                    float(s.override_rate) if s.override_rate is not None else "",
                    float(s.amount),
                    s.hole_section.code if s.hole_section else "",
                    s.sub_activity.name if s.sub_activity else "",
                    s.remarks or "",
                ]
                for s in entry.services or []
            ],
        )

        row_index += 1
        sheet.cell(row_index, 1, "Quantity charges").font = REPORT_TITLE_FONT
        row_index = self._write_table(
            sheet,
            row_index + 1,
            [
                "AFE quantity charge",
                "Quantity",
                "Unit",
                "Unit rate",
                "Override",
                "Amount",
                "Activity",
                "Remarks",
            ],
            [
                [
                    (
                        f"{c.consumable.code} — {c.consumable.name}"
                        if c.consumable
                        else (
                            f"{c.afe_line.secondary_category.code} — "
                            f"{c.afe_line.secondary_category.name}"
                            if c.afe_line
                            else ""
                        )
                    ),
                    float(c.quantity),
                    c.unit.code if c.unit else "",
                    float(c.unit_rate),
                    float(c.override_rate) if c.override_rate is not None else "",
                    float(c.amount),
                    c.sub_activity.name if c.sub_activity else "",
                    c.remarks or "",
                ]
                for c in entry.consumables or []
            ],
        )

        row_index += 1
        for label, value in [
            ("Total operational charges", float(entry.total_services_cost)),
            ("Total quantity charges", float(entry.total_consumables_cost)),
            ("Total daily cost", float(entry.total_daily_cost)),
            ("Cumulative well cost", float(entry.cumulative_cost)),
        ]:
            sheet.cell(row_index, 1, label).font = REPORT_LABEL_FONT
            sheet.cell(row_index, 2, value)
            row_index += 1

        sheet.column_dimensions["A"].width = 38
        for letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            sheet.column_dimensions[letter].width = 14

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_entries_workbook(self, well_id: UUID) -> bytes:
        """The full daily cost register for a well, for filing and audits."""
        well = self.session.get(Well, well_id)
        if not well or not well.is_active:
            raise NotFoundError("Well not found")
        entries = self.session.scalars(
            select(DailyCostEntry)
            .where(DailyCostEntry.well_id == well_id, DailyCostEntry.is_active.is_(True))
            .order_by(DailyCostEntry.entry_date.asc())
        ).all()

        workbook = Workbook()
        register = workbook.active
        assert register is not None
        register.title = "Daily register"
        register["A1"] = f"DAILY COST REGISTER — {well.code} {well.name}"
        register["A1"].font = REPORT_TITLE_FONT
        register["A2"] = f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M')} UTC"
        self._write_table(
            register,
            4,
            [
                "Date",
                "Phase",
                "Hole section",
                "Activity type",
                "Depth",
                "Progress",
                "Operational charges",
                "Quantity charges",
                "Total daily cost",
                "Cumulative cost",
                "Summary",
            ],
            [
                [
                    str(e.entry_date),
                    e.phase or "",
                    e.hole_section.code if e.hole_section else "",
                    e.sub_activity.name if e.sub_activity else "",
                    float(e.current_depth) if e.current_depth is not None else "",
                    float(e.daily_progress) if e.daily_progress is not None else "",
                    float(e.total_services_cost),
                    float(e.total_consumables_cost),
                    float(e.total_daily_cost),
                    float(e.cumulative_cost),
                    e.operational_summary or "",
                ]
                for e in entries
            ],
        )

        services_sheet = workbook.create_sheet("Operational charges")
        self._write_table(
            services_sheet,
            1,
            [
                "Date",
                "AFE charge code",
                "AFE charge name",
                "Type",
                "Hours",
                "Days",
                "Rate basis",
                "Unit rate",
                "Override",
                "Amount",
                "Section",
                "Activity",
                "Vendor",
                "Remarks",
            ],
            [
                [
                    str(e.entry_date),
                    (
                        s.service.code
                        if s.service
                        else (s.afe_line.secondary_category.code if s.afe_line else "")
                    ),
                    (
                        s.service.name
                        if s.service
                        else (s.afe_line.secondary_category.name if s.afe_line else "")
                    ),
                    s.service_type,
                    float(s.service_hours),
                    float(s.operating_days),
                    s.rate_basis,
                    float(s.unit_rate),
                    float(s.override_rate) if s.override_rate is not None else "",
                    float(s.amount),
                    s.hole_section.code if s.hole_section else "",
                    s.sub_activity.name if s.sub_activity else "",
                    s.vendor.name if s.vendor else "",
                    s.remarks or "",
                ]
                for e in entries
                for s in e.services or []
            ],
        )

        consumables_sheet = workbook.create_sheet("Quantity charges")
        self._write_table(
            consumables_sheet,
            1,
            [
                "Date",
                "Code",
                "Name",
                "Quantity",
                "Unit",
                "Unit rate",
                "Override",
                "Amount",
                "Activity",
                "Vendor",
                "Remarks",
            ],
            [
                [
                    str(e.entry_date),
                    (
                        c.consumable.code
                        if c.consumable
                        else (c.afe_line.secondary_category.code if c.afe_line else "")
                    ),
                    (
                        c.consumable.name
                        if c.consumable
                        else (c.afe_line.secondary_category.name if c.afe_line else "")
                    ),
                    float(c.quantity),
                    c.unit.code if c.unit else "",
                    float(c.unit_rate),
                    float(c.override_rate) if c.override_rate is not None else "",
                    float(c.amount),
                    c.sub_activity.name if c.sub_activity else "",
                    c.vendor.name if c.vendor else "",
                    c.remarks or "",
                ]
                for e in entries
                for c in e.consumables or []
            ],
        )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_comparison_workbook(self, well_id: UUID) -> bytes:
        """Every comparison dimension exported to one workbook."""
        comparison = self.get_comparison(well_id)
        workbook = Workbook()

        date_sheet = workbook.active
        assert date_sheet is not None
        date_sheet.title = "By date"
        self._write_table(
            date_sheet,
            1,
            [
                "Date",
                "Day",
                "Phase",
                "Section",
                "Activity",
                "Operational charges",
                "Quantity charges",
                "Daily cost",
                "Cumulative",
                "Planned cumulative",
                "Depth",
                "Progress",
            ],
            [
                [
                    str(p.entry_date),
                    p.day_number,
                    p.phase or "",
                    p.hole_section_code or "",
                    p.activity_name or "",
                    float(p.services_cost),
                    float(p.consumables_cost),
                    float(p.daily_cost),
                    float(p.cumulative_cost),
                    float(p.planned_cumulative) if p.planned_cumulative is not None else "",
                    float(p.current_depth) if p.current_depth is not None else "",
                    float(p.daily_progress) if p.daily_progress is not None else "",
                ]
                for p in comparison.by_date
            ],
        )

        bucket_sheets: list[tuple[str, list[ComparisonBucket]]] = [
            ("By week", comparison.by_week),
            ("By month", comparison.by_month),
            ("By section", comparison.by_section),
            ("By phase", comparison.by_phase),
            ("By activity", comparison.by_activity),
            ("By sub-activity", comparison.by_sub_activity),
        ]
        for title, buckets in bucket_sheets:
            sheet = workbook.create_sheet(title)
            self._write_table(
                sheet,
                1,
                [
                    "Group",
                    "Activity",
                    "Responsible party",
                    "Entries",
                    "Operational charges",
                    "Quantity charges",
                    "Total cost",
                    "Planned cost",
                    "Variance",
                    "Planned days",
                    "Actual days",
                ],
                [
                    [
                        b.label,
                        b.activity_name or "",
                        b.responsible_party or "",
                        b.entry_count,
                        float(b.services_cost),
                        float(b.consumables_cost),
                        float(b.total_cost),
                        float(b.planned_cost) if b.planned_cost is not None else "",
                        float(b.variance) if b.variance is not None else "",
                        float(b.planned_days) if b.planned_days is not None else "",
                        float(b.actual_days) if b.actual_days is not None else "",
                    ]
                    for b in buckets
                ],
            )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _write_table(
        sheet: Worksheet, start_row: int, headers: list[str], rows: list[list[Any]]
    ) -> int:
        """Write a styled table; returns the last row index used."""
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(start_row, column, header)
            cell.font = REPORT_HEADER_FONT
            cell.fill = REPORT_HEADER_FILL
        for offset, row in enumerate(rows, start=1):
            for column, value in enumerate(row, start=1):
                sheet.cell(start_row + offset, column, value)
        return start_row + len(rows)

    def _recompute_cumulative_costs(self, well_id: UUID) -> None:
        entries = self.session.scalars(
            select(DailyCostEntry)
            .where(DailyCostEntry.well_id == well_id, DailyCostEntry.is_active.is_(True))
            .order_by(DailyCostEntry.entry_date.asc())
        ).all()
        running_total = Decimal("0")
        for e in entries:
            running_total += e.total_daily_cost
            e.cumulative_cost = running_total

    @staticmethod
    def _read_entry(entry: DailyCostEntry) -> DailyCostEntryRead:
        services_read = [
            DailyCostServiceLineRead(
                id=s.id,
                daily_cost_entry_id=s.daily_cost_entry_id,
                afe_line_id=s.afe_line_id,
                service_id=s.service_id,
                service_code=(
                    s.service.code
                    if s.service
                    else (s.afe_line.secondary_category.code if s.afe_line else None)
                ),
                service_name=(
                    s.service.name
                    if s.service
                    else (s.afe_line.secondary_category.name if s.afe_line else None)
                ),
                primary_category_name=(
                    s.afe_line.secondary_category.primary_category.name if s.afe_line else None
                ),
                secondary_category_name=(
                    s.afe_line.secondary_category.name if s.afe_line else None
                ),
                cost_code_id=s.cost_code_id,
                cost_code=s.cost_code.code if s.cost_code else None,
                vendor_id=s.vendor_id,
                vendor_name=s.vendor.name if s.vendor else None,
                hole_section_id=s.hole_section_id,
                hole_section_code=s.hole_section.code if s.hole_section else None,
                sub_activity_id=s.sub_activity_id,
                sub_activity_name=s.sub_activity.name if s.sub_activity else None,
                service_type=getattr(s, "service_type", "operation"),
                service_hours=s.service_hours,
                operating_days=s.operating_days,
                rate_basis=s.rate_basis,
                unit_rate=s.unit_rate,
                override_rate=s.override_rate,
                amount=s.amount,
                remarks=s.remarks,
                created_at=s.created_at,
                updated_at=s.updated_at,
            )
            for s in (entry.services or [])
        ]
        consumables_read = [
            DailyCostConsumableLineRead(
                id=c.id,
                daily_cost_entry_id=c.daily_cost_entry_id,
                afe_line_id=c.afe_line_id,
                consumable_id=c.consumable_id,
                consumable_code=(
                    c.consumable.code
                    if c.consumable
                    else (c.afe_line.secondary_category.code if c.afe_line else None)
                ),
                consumable_name=(
                    c.consumable.name
                    if c.consumable
                    else (c.afe_line.secondary_category.name if c.afe_line else None)
                ),
                primary_category_name=(
                    c.afe_line.secondary_category.primary_category.name if c.afe_line else None
                ),
                secondary_category_name=(
                    c.afe_line.secondary_category.name if c.afe_line else None
                ),
                cost_code_id=c.cost_code_id,
                cost_code=c.cost_code.code if c.cost_code else None,
                vendor_id=c.vendor_id,
                vendor_name=c.vendor.name if c.vendor else None,
                sub_activity_id=c.sub_activity_id,
                sub_activity_name=c.sub_activity.name if c.sub_activity else None,
                quantity=c.quantity,
                unit_id=c.unit_id,
                unit_code=c.unit.code if c.unit else None,
                unit_rate=c.unit_rate,
                override_rate=c.override_rate,
                amount=c.amount,
                remarks=c.remarks,
                created_at=c.created_at,
                updated_at=c.updated_at,
            )
            for c in (entry.consumables or [])
        ]
        return DailyCostEntryRead(
            id=entry.id,
            well_id=entry.well_id,
            well_code=entry.well.code if entry.well else None,
            afe_id=entry.afe_id,
            afe_code=entry.afe.code if entry.afe else None,
            entry_date=entry.entry_date,
            hole_section_id=entry.hole_section_id,
            hole_section_code=entry.hole_section.code if entry.hole_section else None,
            phase=entry.phase,
            sub_activity_id=entry.sub_activity_id,
            sub_activity_name=entry.sub_activity.name if entry.sub_activity else None,
            current_depth=entry.current_depth,
            daily_progress=entry.daily_progress,
            operational_summary=entry.operational_summary,
            total_services_cost=entry.total_services_cost,
            total_consumables_cost=entry.total_consumables_cost,
            total_daily_cost=entry.total_daily_cost,
            cumulative_cost=entry.cumulative_cost,
            is_active=entry.is_active,
            services=services_read,
            consumables=consumables_read,
            created_at=entry.created_at,
            updated_at=entry.updated_at,
        )
