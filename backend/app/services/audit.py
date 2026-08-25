"""Global audit logging service."""

import json
import logging
from io import BytesIO
from math import ceil
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.audit import AuditLog
from app.repositories.audit import AuditRepository
from app.schemas.audit import AuditLogRead
from app.schemas.master_data import PageResponse

logger = logging.getLogger("app.audit")


def log_entity_action(
    session: Session,
    actor_id: UUID | None,
    action: str,
    entity_type: str,
    entity_id: UUID | None = None,
    entity_code: str | None = None,
    details: Any | None = None,
) -> None:
    """Record a global audit entry for an entity action.

    Shared by every service so create/update/soft_delete/recover/hard_delete all
    follow the same audited procedure. Never raises: an audit failure must not
    block the business operation the entry describes.
    """

    try:
        actor_email = None
        try:
            from app.models.user import User

            user = session.get(User, actor_id) if actor_id is not None else None
            if user:
                actor_email = user.email
        except Exception:
            pass
        AuditService(session, actor_id, actor_email).log(
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_code=entity_code,
            details=details,
            commit=False,
        )
    except Exception:
        logger.exception(
            "Audit log write failed",
            extra={"action": action, "entity_type": entity_type, "entity_id": str(entity_id)},
        )


class AuditService:
    def __init__(
        self, session: Session, actor_id: UUID | None = None, actor_email: str | None = None
    ) -> None:
        self.session = session
        self.actor_id = actor_id
        self.actor_email = actor_email
        self.repository = AuditRepository(session)

    def log(
        self,
        *,
        action: str,
        entity_type: str,
        entity_id: UUID | None = None,
        entity_code: str | None = None,
        details: dict[str, Any] | str | None = None,
        ip_address: str | None = None,
        user_agent: str | None = None,
        commit: bool = False,
    ) -> AuditLog:
        details_str: str | None = None
        if isinstance(details, dict):
            try:
                details_str = json.dumps(details, default=str)
            except Exception:
                details_str = str(details)
        elif isinstance(details, str):
            details_str = details

        # JSON logger for observability
        logger.info(
            action,
            extra={
                "actor_id": str(self.actor_id) if self.actor_id else None,
                "actor_email": self.actor_email,
                "entity_type": entity_type,
                "entity_id": str(entity_id) if entity_id else None,
                "entity_code": entity_code,
                "details": details_str,
            },
        )

        entry = self.repository.create(
            actor_id=self.actor_id,
            actor_email=self.actor_email,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_code=entity_code,
            details=details_str,
            ip_address=ip_address,
            user_agent=user_agent,
        )
        if commit:
            self.session.commit()
            self.session.refresh(entry)
        else:
            self.session.flush()
        return entry

    def list_page(
        self,
        *,
        page: int,
        page_size: int,
        search: str | None,
        action: str | None,
        entity_type: str | None,
        actor_id: UUID | None,
    ) -> PageResponse:
        records, total = self.repository.list(
            page=page,
            page_size=page_size,
            search=search,
            action=action,
            entity_type=entity_type,
            actor_id=actor_id,
        )
        return PageResponse(
            items=[AuditLogRead.model_validate(r) for r in records],
            page=page,
            page_size=page_size,
            total=total,
            pages=ceil(total / page_size) if total else 0,
        )

    def export_workbook(
        self,
        *,
        search: str | None,
        action: str | None,
        entity_type: str | None,
        actor_id: UUID | None,
    ) -> bytes:
        """Export every row matching the current audit filters."""
        records, _ = self.repository.list(
            page=1,
            page_size=1_000_000,
            search=search,
            action=action,
            entity_type=entity_type,
            actor_id=actor_id,
        )
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise RuntimeError("Workbook did not create a worksheet")
        sheet.title = "Audit Log"
        headers = [
            "Timestamp",
            "Actor",
            "Action",
            "Entity",
            "Entity code",
            "Entity ID",
            "Details",
            "IP address",
            "User agent",
        ]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for row_index, record in enumerate(records, start=2):
            values = [
                record.created_at,
                record.actor_email,
                record.action,
                record.entity_type,
                record.entity_code,
                str(record.entity_id) if record.entity_id else None,
                record.details,
                record.ip_address,
                record.user_agent,
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row_index, column, value)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:I{max(1, len(records) + 1)}"
        for column, width in {
            "A": 22,
            "B": 28,
            "C": 20,
            "D": 24,
            "E": 24,
            "F": 38,
            "G": 70,
            "H": 18,
            "I": 42,
        }.items():
            sheet.column_dimensions[column].width = width
        stream = BytesIO()
        workbook.save(stream)
        content = stream.getvalue()
        self.log(
            action="export",
            entity_type="audit_log",
            entity_code="filtered-audit-log",
            details={
                "row_count": len(records),
                "filters": {"search": search, "action": action, "entity_type": entity_type},
            },
            commit=True,
        )
        return content

    def get(self, log_id: UUID) -> AuditLogRead:
        from app.core.exceptions import NotFoundError

        record = self.repository.get(log_id)
        if record is None:
            raise NotFoundError("Audit log not found")
        return AuditLogRead.model_validate(record)
