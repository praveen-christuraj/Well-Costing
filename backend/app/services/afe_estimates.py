"""AFE Cost Estimate services.

The AFE Cost Estimates page prices every configured AFE line with a
well-scoped unit rate. User-configured primary/secondary classifications are
shown as entered; no service/tangible/other type is inferred. The line's rate
basis is the explicit calculation method used downstream.
"""

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessValidationError, NotFoundError
from app.models.afe import Afe, AfeAuditLog, AfeLine
from app.models.afe_estimates import AfeCostEstimateLine
from app.models.master_data import Vendor
from app.schemas.afe_estimates import (
    AfeCostEstimateGroupTotal,
    AfeCostEstimateLineRead,
    AfeCostEstimateRead,
    AfeCostEstimateSaveRequest,
)
from app.services.audit import log_entity_action

CONSUMABLE_ITEM_TYPES = {"mud_chemical", "cement_additive", "material"}

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class AfeEstimateService:
    """Well-scoped pricing of AFE lines."""

    def __init__(self, session: Session, actor_id: UUID) -> None:
        self.session, self.actor_id = session, actor_id

    # ------------------------------------------------------------------ reads
    def get_estimate(self, afe_id: UUID) -> AfeCostEstimateRead:
        afe = self._get_afe(afe_id)
        rate_rows = self._rate_rows(afe_id)
        lines = [
            self._read_line(item, rate_rows.get(item.id)) for item in afe.items if item.is_active
        ]
        return self._build_read(afe, lines)

    def save_rates(self, afe_id: UUID, payload: AfeCostEstimateSaveRequest) -> AfeCostEstimateRead:
        """Save the one current estimate rate for each submitted AFE scope line.

        The released screen deliberately exposes one estimated rate rather than
        the retired multi-rate/usage planner. The additional persisted rate
        columns remain populated for historical Daily Cost rows, but the audit
        snapshot records exactly what the user changed on this screen.
        """
        afe = self._get_afe(afe_id)
        line_ids = {item.id for item in afe.items if item.is_active}
        rate_rows = self._rate_rows(afe_id)

        saved_rates: list[dict[str, object]] = []
        for entry in payload.rates:
            if entry.afe_line_id not in line_ids:
                raise BusinessValidationError(
                    "A rate referenced an AFE line that does not belong to this AFE. "
                    "Reload the AFE Cost Estimates page and try again."
                )
            if entry.vendor_id is not None:
                vendor = self.session.get(Vendor, entry.vendor_id)
                if vendor is None or not vendor.is_active:
                    raise BusinessValidationError("vendor_id must reference an active vendor")
            row = rate_rows.get(entry.afe_line_id)
            previous = self._rate_snapshot(row) if row is not None else None
            if row is None:
                row = AfeCostEstimateLine(
                    afe_id=afe_id,
                    afe_line_id=entry.afe_line_id,
                    created_by=self.actor_id,
                )
                self.session.add(row)
                rate_rows[entry.afe_line_id] = row

            # ``unit_rate`` remains the compatibility alias consumed by Daily
            # Cost. The current screen's sole rate is mirrored to operating.
            rate = Decimal(str(entry.operating_rate or entry.unit_rate))
            row.operating_rate = rate
            row.unit_rate = rate
            for field in (
                "standby_rate",
                "mobilization_rate",
                "demobilization_rate",
                "fixed_charges",
                "personnel_operating_rate",
                "personnel_standby_rate",
                "other_rate",
            ):
                setattr(row, field, Decimal(str(getattr(entry, field))))
            row.multiply_by_input = entry.multiply_by_input
            row.vendor_id = entry.vendor_id
            row.remarks = entry.remarks.strip() if entry.remarks else None
            row.is_active = True
            row.updated_by = self.actor_id
            saved_rates.append(
                {
                    "afe_line_id": str(entry.afe_line_id),
                    "before": previous,
                    "after": self._rate_snapshot(row),
                }
            )

        self.session.flush()
        self._append_afe_history(
            afe,
            "cost_estimate_rates_saved",
            f"Saved estimate rates for {len(saved_rates)} AFE line(s)",
        )
        log_entity_action(
            self.session,
            self.actor_id,
            "save_rates",
            "afe_cost_estimate",
            entity_id=afe_id,
            entity_code=afe.code,
            details={"rates_saved": len(saved_rates), "rates": saved_rates},
        )
        self.session.commit()
        return self.get_estimate(afe_id)

    def record_print(self, afe_id: UUID) -> None:
        """Audit a browser print of the current submitted cost estimate."""
        estimate = self.get_estimate(afe_id)
        afe = self._get_afe(afe_id)
        self._append_afe_history(afe, "cost_estimate_printed", "AFE Cost Estimate printed")
        log_entity_action(
            self.session,
            self.actor_id,
            "print",
            "afe_cost_estimate",
            entity_id=afe.id,
            entity_code=afe.code,
            details={
                "line_count": estimate.line_count,
                "estimated_total": str(estimate.estimated_total),
                "source": "afe_cost_estimate_page",
            },
        )
        self.session.commit()

    # ----------------------------------------------------------------- export
    def export_workbook(self, afe_id: UUID) -> bytes:
        """Full AFE Cost Estimate as a printable Excel record."""
        estimate = self.get_estimate(afe_id)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "AFE Cost Estimate"

        sheet["A1"] = "AFE COST ESTIMATE"
        sheet["A1"].font = TITLE_FONT
        sheet["A2"] = f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M')} UTC"

        header_pairs = [
            ("Project", f"{estimate.project_code or ''} — {estimate.project_name or ''}"),
            ("Well", f"{estimate.well_code or ''} — {estimate.well_name or ''}"),
            ("Rig", estimate.rig_name or ""),
            ("AFE", f"{estimate.afe_code} (rev {estimate.revision_number})"),
            ("Title", estimate.afe_title),
            ("Status", estimate.afe_status),
            ("Planned days", float(estimate.total_planned_days)),
            ("Estimated total", float(estimate.estimated_total)),
        ]
        row_index = 4
        for label, value in header_pairs:
            sheet.cell(row_index, 1, label).font = LABEL_FONT
            sheet.cell(row_index, 2, value)
            row_index += 1

        row_index += 1
        # Current AFE scope does not carry planned usage/day or quantity. The
        # estimate workbook therefore contains only current scope, rate, and
        # total fields rather than the retired usage-planning columns.
        headers = [
            "#",
            "Primary category",
            "Secondary category",
            "Cost code",
            "Hole section",
            "Rate basis",
            "Estimated total rate",
            "Estimated amount",
            "Vendor",
            "Remarks",
        ]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row_index, column, header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        first_data_row = row_index + 1
        for line in estimate.lines:
            row_index += 1
            values: list[Any] = [
                line.line_number,
                line.primary_category_name or line.primary_category_code,
                line.secondary_category_name or line.secondary_category_code,
                line.cost_code,
                "All sections" if line.applies_to_all_sections else line.hole_section_code,
                line.rate_basis.replace("_", " "),
                float(line.unit_rate),
                float(line.estimated_amount),
                line.vendor_name,
                line.remarks,
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column, value)
                cell.border = THIN_BORDER
        row_index += 1
        total_cell = sheet.cell(row_index, 7, "Estimated total")
        total_cell.font = LABEL_FONT
        amount_cell = sheet.cell(row_index, 8, float(estimate.estimated_total))
        amount_cell.font = LABEL_FONT
        if estimate.lines:
            sheet.freeze_panes = sheet.cell(first_data_row, 1).coordinate
        for column, width in enumerate([6, 24, 30, 14, 14, 16, 18, 18, 20, 24], start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width

        summary = workbook.create_sheet("Summaries")
        self._write_summary_block(summary, 1, "By hole section", estimate.totals_by_section)
        offset = 4 + len(estimate.totals_by_section)
        self._write_summary_block(
            summary, offset, "By primary category", estimate.totals_by_primary_category
        )
        offset += 3 + len(estimate.totals_by_primary_category)
        self._write_summary_block(
            summary, offset, "By secondary category", estimate.totals_by_secondary_category
        )
        offset += 3 + len(estimate.totals_by_secondary_category)
        self._write_summary_block(summary, offset, "By cost code", estimate.totals_by_cost_code)
        offset += 3 + len(estimate.totals_by_cost_code)
        self._write_summary_block(summary, offset, "By rate basis", estimate.totals_by_rate_basis)
        summary.column_dimensions["A"].width = 30
        summary.column_dimensions["B"].width = 12
        summary.column_dimensions["C"].width = 18

        output = BytesIO()
        workbook.save(output)
        content = output.getvalue()
        afe = self._get_afe(estimate.afe_id)
        self._append_afe_history(
            afe,
            "cost_estimate_exported",
            "AFE Cost Estimate exported to Excel",
        )
        log_entity_action(
            self.session,
            self.actor_id,
            "export",
            "afe_cost_estimate",
            entity_id=estimate.afe_id,
            entity_code=estimate.afe_code,
            details={
                "line_count": estimate.line_count,
                "estimated_total": str(estimate.estimated_total),
                "format": "xlsx",
            },
        )
        self.session.commit()
        return content

    @staticmethod
    def _write_summary_block(
        sheet: Worksheet,
        start_row: int,
        title: str,
        totals: list[AfeCostEstimateGroupTotal],
    ) -> None:
        sheet.cell(start_row, 1, title).font = TITLE_FONT
        header_row = start_row + 1
        for column, header in enumerate(["Group", "Lines", "Estimated total"], start=1):
            cell = sheet.cell(header_row, column, header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for index, total in enumerate(totals, start=1):
            sheet.cell(header_row + index, 1, total.label)
            sheet.cell(header_row + index, 2, total.line_count)
            cell = sheet.cell(header_row + index, 3, float(total.estimated_total))
            cell.alignment = Alignment(horizontal="right")

    # ---------------------------------------------------------------- helpers
    def _get_afe(self, afe_id: UUID) -> Afe:
        afe = self.session.get(Afe, afe_id)
        if afe is None or not afe.is_active:
            raise NotFoundError("AFE not found")
        # Pricing is the next workflow step after an AFE is submitted. This is
        # enforced server-side as well as in the selector so a stale URL or API
        # client can never price a draft scope.
        if afe.status != "submitted":
            raise BusinessValidationError(
                "Only submitted AFEs are available in AFE Cost Estimates. "
                "Submit the AFE scope first."
            )
        return afe

    def _rate_rows(self, afe_id: UUID) -> dict[UUID, AfeCostEstimateLine]:
        rows = self.session.scalars(
            select(AfeCostEstimateLine).where(
                AfeCostEstimateLine.afe_id == afe_id,
                AfeCostEstimateLine.is_active.is_(True),
            )
        ).all()
        return {row.afe_line_id: row for row in rows}

    @staticmethod
    def effective_quantity(item: AfeLine) -> Decimal:
        """Historical planned quantity, retained only for older AFE records."""
        if item.quantity and item.quantity > 0:
            return Decimal(item.quantity)
        if item.computed_quantity and item.computed_quantity > 0:
            return Decimal(item.computed_quantity)
        return Decimal("0")

    @classmethod
    def estimate_multiplier(cls, item: AfeLine) -> Decimal:
        """Return the multiplier for a line's estimated amount.

        Current AFE Lines are scope-only: no usage/day, quantity, or UOM is
        entered there. A saved estimate rate is therefore the line's estimated
        total and uses a multiplier of one. Older records that explicitly have
        a positive planned quantity remain readable with their historic
        quantity-based calculation.
        """
        historical_quantity = cls.effective_quantity(item)
        return historical_quantity if historical_quantity > 0 else Decimal("1")

    @staticmethod
    def _rate_snapshot(rate: AfeCostEstimateLine | None) -> dict[str, object] | None:
        if rate is None:
            return None
        return {
            "unit_rate": str(rate.unit_rate),
            "operating_rate": str(rate.operating_rate),
            "standby_rate": str(rate.standby_rate),
            "mobilization_rate": str(rate.mobilization_rate),
            "demobilization_rate": str(rate.demobilization_rate),
            "fixed_charges": str(rate.fixed_charges),
            "personnel_operating_rate": str(rate.personnel_operating_rate),
            "personnel_standby_rate": str(rate.personnel_standby_rate),
            "other_rate": str(rate.other_rate),
            "multiply_by_input": rate.multiply_by_input,
            "vendor_id": str(rate.vendor_id) if rate.vendor_id else None,
            "remarks": rate.remarks,
        }

    def _append_afe_history(self, afe: Afe, action: str, remarks: str) -> None:
        """Keep the AFE-local history useful alongside the global audit log."""
        self.session.add(
            AfeAuditLog(
                afe_id=afe.id,
                action=action,
                previous_status=afe.status,
                new_status=afe.status,
                remarks=remarks,
                actor_id=self.actor_id,
            )
        )

    @staticmethod
    def _classification(
        item: AfeLine,
    ) -> tuple[UUID | None, str | None, str | None, UUID | None, str | None, str | None]:
        """Return exactly the classification configured in Master Data/AFE.

        There is intentionally no fallback that guesses a type from category
        names or rate bases. Historical catalogue identity remains available as
        a display fallback only.
        """
        secondary = item.secondary_category
        primary = secondary.primary_category if secondary else None
        return (
            primary.id if primary else None,
            primary.code if primary else None,
            primary.name if primary else None,
            secondary.id if secondary else None,
            secondary.code if secondary else None,
            secondary.name if secondary else None,
        )

    def _read_line(
        self, item: AfeLine, rate: AfeCostEstimateLine | None
    ) -> AfeCostEstimateLineRead:
        historical_quantity = self.effective_quantity(item)
        quantity = historical_quantity if historical_quantity > 0 else None
        estimate_multiplier = self.estimate_multiplier(item)
        unit_rate = Decimal(rate.operating_rate or rate.unit_rate) if rate else Decimal("0")
        estimated_amount = unit_rate * estimate_multiplier
        (
            primary_id,
            primary_code,
            primary_name,
            secondary_id,
            secondary_code,
            secondary_name,
        ) = self._classification(item)
        catalog_item = item.catalog_item
        return AfeCostEstimateLineRead(
            afe_line_id=item.id,
            estimate_line_id=rate.id if rate else None,
            line_number=item.line_number,
            catalog_item_id=item.catalog_item_id,
            catalog_item_code=catalog_item.code if catalog_item else None,
            catalog_item_name=catalog_item.name if catalog_item else None,
            primary_category_id=primary_id,
            primary_category_code=primary_code,
            primary_category_name=primary_name,
            secondary_category_id=secondary_id,
            secondary_category_code=secondary_code,
            secondary_category_name=secondary_name,
            cost_code_id=item.cost_code_id,
            cost_code=item.cost_code.code if item.cost_code else None,
            hole_section_id=item.hole_section_id,
            hole_section_code=item.hole_section.code if item.hole_section else None,
            applies_to_all_sections=item.applies_to_all_sections,
            rate_basis=item.rate_basis,
            quantity=quantity,
            unit_id=item.unit_id,
            unit_code=item.unit.code if item.unit else None,
            unit_rate=unit_rate,
            operating_rate=unit_rate,
            standby_rate=Decimal(rate.standby_rate) if rate else Decimal("0"),
            mobilization_rate=Decimal(rate.mobilization_rate) if rate else Decimal("0"),
            demobilization_rate=Decimal(rate.demobilization_rate) if rate else Decimal("0"),
            fixed_charges=Decimal(rate.fixed_charges) if rate else Decimal("0"),
            personnel_operating_rate=(
                Decimal(rate.personnel_operating_rate) if rate else Decimal("0")
            ),
            personnel_standby_rate=Decimal(rate.personnel_standby_rate) if rate else Decimal("0"),
            other_rate=Decimal(rate.other_rate) if rate else Decimal("0"),
            multiply_by_input=rate.multiply_by_input if rate else item.rate_basis == "daily",
            estimate_multiplier=estimate_multiplier,
            estimated_amount=estimated_amount,
            vendor_id=rate.vendor_id if rate else None,
            vendor_name=rate.vendor.name if rate and rate.vendor else None,
            remarks=rate.remarks if rate else None,
            notes=item.notes,
            rate_saved_at=rate.updated_at if rate else None,
        )

    def _build_read(self, afe: Afe, lines: list[AfeCostEstimateLineRead]) -> AfeCostEstimateRead:
        estimated_total = sum((line.estimated_amount for line in lines), Decimal("0"))
        priced = sum(1 for line in lines if line.unit_rate > 0)
        well = afe.well
        return AfeCostEstimateRead(
            afe_id=afe.id,
            afe_code=afe.code,
            afe_title=afe.title,
            afe_status=afe.status,
            revision_number=afe.revision_number,
            project_code=well.project.code if well and well.project else None,
            project_name=well.project.name if well and well.project else None,
            well_id=afe.well_id,
            well_code=well.code if well else None,
            well_name=well.name if well else None,
            rig_name=well.rig_name if well else None,
            budget_amount=Decimal(afe.budget_amount or 0),
            total_planned_days=Decimal(afe.total_planned_days or 0),
            total_planned_depth=Decimal(afe.total_planned_depth or 0),
            depth_unit_code=afe.depth_unit.code if afe.depth_unit else None,
            line_count=len(lines),
            priced_line_count=priced,
            unpriced_line_count=len(lines) - priced,
            estimated_total=estimated_total,
            variance_to_budget=Decimal(afe.budget_amount or 0) - estimated_total,
            lines=lines,
            totals_by_section=self._group(lines, "section"),
            totals_by_primary_category=self._group(lines, "primary_category"),
            totals_by_secondary_category=self._group(lines, "secondary_category"),
            totals_by_cost_code=self._group(lines, "cost_code"),
            totals_by_rate_basis=self._group(lines, "rate_basis"),
        )

    @staticmethod
    def _group(
        lines: list[AfeCostEstimateLineRead], dimension: str
    ) -> list[AfeCostEstimateGroupTotal]:
        buckets: dict[str, AfeCostEstimateGroupTotal] = {}
        for line in lines:
            if dimension == "section":
                key = (
                    "All sections"
                    if line.applies_to_all_sections
                    else (line.hole_section_code or "Unassigned")
                )
            elif dimension == "primary_category":
                key = line.primary_category_name or line.primary_category_code or "Unassigned"
            elif dimension == "secondary_category":
                key = line.secondary_category_name or line.secondary_category_code or "Unassigned"
            elif dimension == "cost_code":
                key = line.cost_code or "Unassigned"
            else:
                key = line.rate_basis.replace("_", " ").title()
            bucket = buckets.get(key)
            if bucket is None:
                bucket = AfeCostEstimateGroupTotal(key=key, label=key, line_count=0)
                buckets[key] = bucket
            bucket.line_count += 1
            bucket.estimated_total += line.estimated_amount
        return sorted(buckets.values(), key=lambda b: b.estimated_total, reverse=True)
