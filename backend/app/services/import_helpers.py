"""Shared helpers for bulk import / export across catalogue modules.

Kept dependency-light so both API routes and tests can reuse the same flexible
parsing: dates accept a wide range of formats (including Excel serial numbers),
numbers strip currency symbols/thousands separators, percentages accept both
``110`` and ``1.1`` styles, and workbook rows are normalised into trimmed,
snake-cased dicts.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

_DATE_FORMATS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%m-%d-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y/%m/%d",
    "%d.%m.%Y",
    "%d-%b-%Y",
    "%d-%B-%Y",
    "%b %d, %Y",
    "%B %d, %Y",
    "%Y-%m-%d %H:%M:%S",
    "%d-%m-%Y %H:%M",
    "%m/%d/%Y %H:%M",
]


def norm_header(value: Any) -> str:
    """Normalise a column header: lower-case, trimmed, underscores for spaces."""

    return str(value or "").strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


def parse_date_flexible(value: Any) -> date | None:
    """Parse a date from strings, Excel serials, datetime/date objects.

    Raises ValueError with a helpful message when nothing matches.
    """

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    val = str(value).strip()
    if not val:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    # Excel serial date (days since 1899-12-30)
    try:
        if re.match(r"^\d+(\.\d+)?$", val):
            serial = float(val)
            if 1 < serial < 100000:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    except Exception:  # pragma: no cover - defensive
        pass
    # Last resort: split on - / . and guess Y/M/D vs D/M/Y
    try:
        parts = re.split(r"[-/. ]", val.split()[0])
        if len(parts) == 3:
            if len(parts[0]) == 4:
                y, m, d = parts
            else:
                d, m, y = parts
                if len(y) == 2:
                    y = "20" + y if int(y) < 50 else "19" + y
            return date(int(y), int(m), int(d))
    except Exception:
        pass
    raise ValueError(f"Unrecognized date format: '{value}'. Use YYYY-MM-DD or DD/MM/YYYY.")


def parse_decimal(value: Any, *, field: str = "value", allow_blank: bool = True) -> Decimal | None:
    """Parse a number, stripping currency symbols, commas and spaces."""

    if value is None:
        if allow_blank:
            return None
        raise ValueError(f"{field} is required")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    val = str(value).strip()
    if not val:
        if allow_blank:
            return None
        raise ValueError(f"{field} is required")
    cleaned = re.sub(r"[^0-9.\-]", "", val)
    if cleaned in ("", "-", ".", "-."):
        if allow_blank:
            return None
        raise ValueError(f"{field} is required")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {field}: '{value}'") from exc


def parse_uplift(value: Any, *, default: Decimal = Decimal("100")) -> Decimal:
    """Parse a cost-uplift percentage.

    Accepts ``110`` / ``110%`` / ``1.1`` (values <= 2 are treated as a
    multiplier). Defaults to 100% when blank.
    """

    if value is None or str(value).strip() == "":
        return default
    dec = parse_decimal(value, field="cost uplift", allow_blank=False)
    assert dec is not None
    if dec <= Decimal("2"):
        dec = dec * Decimal("100")
    return dec.quantize(Decimal("0.01"))


def final_cost(unit_rate: Decimal | None, uplift: Decimal | None) -> Decimal:
    """Final cost = unit rate as per PO x cost uplift percentage."""

    rate = unit_rate or Decimal("0")
    pct = uplift if uplift is not None else Decimal("100")
    return (rate * pct / Decimal("100")).quantize(Decimal("0.01"))


def read_tabular_file(contents: bytes, filename: str) -> list[tuple[int, dict[str, Any]]]:
    """Parse CSV/XLSX upload into (row_number, normalised_row_dict) pairs."""

    rows: list[tuple[int, dict[str, Any]]] = []
    if filename.lower().endswith(".csv"):
        text_data = contents.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text_data))
        for i, row in enumerate(reader, start=2):
            norm = {norm_header(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
            if any(str(v).strip() for v in norm.values() if v is not None):
                rows.append((i, norm))
    elif filename.lower().endswith((".xlsx", ".xls")):
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel workbook has no active sheet")
        header_row = [norm_header(cell.value) for cell in ws[1]]
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v is not None and str(v).strip() != "" for v in row):
                continue
            row_dict: dict[str, Any] = {}
            for h_name, val in zip(header_row, row, strict=False):
                if h_name:
                    row_dict[h_name] = val
            rows.append((r_idx, row_dict))
    else:
        raise ValueError("Unsupported file format. Please upload a CSV or XLSX file.")
    return rows


def row_get(row: dict[str, Any], *keys: str) -> Any:
    """First non-blank value among the given header aliases."""

    for key in keys:
        if key in row:
            val = row[key]
            if val is not None and str(val).strip() != "":
                return val
    return None


def template_xlsx_response(
    filename: str,
    headers: list[str],
    sample_rows: list[list[Any]] | None = None,
    dropdowns: dict[int, list[str]] | None = None,
    note: str | None = None,
) -> Any:
    """Build a ready-to-fill XLSX import template.

    The 'Import' sheet mirrors the upload parser exactly (same header names,
    nothing else), so the user can download it, fill it with data and upload
    the very same file — an untouched template simply imports zero rows.
    Header cells are styled, the header row is frozen, columns are sized and
    enum columns get in-cell dropdown validation where practical. Example rows
    and guidance live on a separate 'Instructions' sheet so they can never be
    mistaken for data by the parser.
    """

    from fastapi import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    ws.append(headers)

    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(34, len(headers[col_idx - 1]) + 8))
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Guidance + examples live on a separate sheet so a filled template can be
    # uploaded as-is without the parser treating anything as extra data rows.
    guide = wb.create_sheet("Instructions")
    guide.column_dimensions["A"].width = 110
    title_cell = guide.cell(row=1, column=1, value="How to use this template")
    title_cell.font = Font(bold=True, size=12, color="0F766E")
    lines = [
        "1. Fill the 'Import' sheet with your data — one record per row, starting at row 2.",
        "2. Keep the header row exactly as it is; the import matches columns by header name.",
        "3. Upload this same file from the Import dialog on the page.",
    ]
    if note:
        lines += ["", f"Note: {note}"]
    for offset, line in enumerate(lines, start=3):
        cell = guide.cell(row=offset, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(color="334155", size=10)
    if sample_rows:
        example_title = guide.cell(row=len(lines) + 5, column=1, value="Example rows (for reference — do not paste into the Import sheet unchanged)")
        example_title.font = Font(bold=True, size=11, color="0F766E")
        start_row = len(lines) + 6
        example_header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        for col_idx, header in enumerate(headers, start=1):
            cell = guide.cell(row=start_row + 1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = example_header_fill
            guide.column_dimensions[get_column_letter(col_idx)].width = max(
                guide.column_dimensions[get_column_letter(col_idx)].width or 0,
                min(30, len(header) + 6),
            )
        for row_offset, row in enumerate(sample_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                guide.cell(row=start_row + row_offset, column=col_idx, value=value).font = Font(size=10)

    # Data rows start at row 2; validation covers a generous fill range.
    for col_idx, choices in (dropdowns or {}).items():
        if not choices:
            continue
        formula = '"' + ",".join(str(choice) for choice in choices[:60]) + '"'
        if len(formula) > 255:  # Excel list-validation formula limit
            continue
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Pick a value from the dropdown list."
        validation.errorTitle = "Invalid value"
        column = get_column_letter(col_idx)
        validation.add(f"{column}2:{column}1000")
        ws.add_data_validation(validation)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


def spreadsheet_response(rows: list[list[Any]], headers: list[str], filename: str, fmt: str) -> Any:
    """Build a CSV or XLSX download Response from header + row lists."""

    from fastapi import Response

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = filename[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )
