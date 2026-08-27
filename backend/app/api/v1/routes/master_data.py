"""Master data API routes for UOM, Currencies, Phases, Activities, and Hole Sections."""

import csv
import io
import logging
from datetime import UTC, datetime
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.dependencies.auth import get_current_user
from app.db.session import get_db
from app.models.master_data import Activity, Currency, HoleSection, Phase, UnitOfMeasurement
from app.models.user import User
from app.schemas.master_data import (
    ActivityOut,
    BulkImportResponse,
    CurrencyOut,
    HoleSectionOut,
    PhaseOut,
    UOMOut,
)
from app.services.audit import log_audit

logger = logging.getLogger("app")

router = APIRouter(prefix="/master-data", tags=["master-data"])

MODULE_CONFIG = {
    "uom": {
        "model": UnitOfMeasurement,
        "name": "Unit of Measurements",
        "code_field": "unit_code",
        "name_field": "unit_name",
        "symbol_field": "unit_symbol",
        "out_schema": UOMOut,
    },
    "currencies": {
        "model": Currency,
        "name": "Currency",
        "code_field": "currency_code",
        "name_field": "currency_name",
        "symbol_field": "currency_symbol",
        "out_schema": CurrencyOut,
    },
    "phases": {
        "model": Phase,
        "name": "Phases",
        "code_field": "phase_code",
        "name_field": "phase_name",
        "symbol_field": None,
        "out_schema": PhaseOut,
    },
    "activities": {
        "model": Activity,
        "name": "Activities",
        "code_field": "activity_code",
        "name_field": "activity_name",
        "symbol_field": None,
        "out_schema": ActivityOut,
    },
    "hole-sections": {
        "model": HoleSection,
        "name": "Hole Sections",
        "code_field": "section_code",
        "name_field": "section_name",
        "symbol_field": None,
        "out_schema": HoleSectionOut,
    },
}


PROTECTED_FIELDS = {"id", "created_at", "updated_at", "created_by", "updated_by"}


def _get_config(module: str) -> dict[str, Any]:
    cfg = MODULE_CONFIG.get(module)
    if not cfg:
        raise HTTPException(status_code=404, detail=f"Module '{module}' not found")
    return cfg


def _allowed_fields(model: Any) -> set[str]:
    return set(model.__table__.columns.keys()) - PROTECTED_FIELDS


def _clean_payload(model: Any, payload: dict[str, Any]) -> dict[str, Any]:
    allowed = _allowed_fields(model)
    return {key: value for key, value in payload.items() if key in allowed}


def _prepare_payload(cfg: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    """Keep only model columns and default a missing symbol to the code."""

    cleaned = _clean_payload(cfg["model"], payload)
    code_field = cfg["code_field"]
    name_field = cfg["name_field"]
    symbol_field = cfg["symbol_field"]
    if code_field in cleaned and cleaned[code_field] is not None:
        cleaned[code_field] = str(cleaned[code_field]).strip()
    if name_field in cleaned and cleaned[name_field] is not None:
        cleaned[name_field] = str(cleaned[name_field]).strip()
    if symbol_field:
        symbol = cleaned.get(symbol_field)
        if symbol is None or (isinstance(symbol, str) and not symbol.strip()):
            cleaned[symbol_field] = cleaned.get(code_field) or ""
        else:
            cleaned[symbol_field] = str(symbol).strip()
    return cleaned


def _to_out(cfg: dict[str, Any], record: Any) -> Any | None:
    schema = cfg["out_schema"]
    try:
        return schema.model_validate(record)
    except ValidationError:
        data = {name: getattr(record, name, None) for name in schema.model_fields}
        try:
            return schema.model_validate(data)
        except ValidationError:
            logger.warning(
                "Skipping unreadable %s record %s",
                cfg["name"],
                getattr(record, "id", None),
            )
            return None


@router.get("/{module}")
def list_records(
    module: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
) -> list[Any]:
    """List all active records for a master data module."""
    cfg = _get_config(module)
    model = cfg["model"]
    stmt = select(model).where(model.is_deleted == False).order_by(model.id.desc())
    records = db.scalars(stmt).all()
    return [item for item in (_to_out(cfg, r) for r in records) if item is not None]


@router.get("/{module}/deleted")
def list_deleted_records(
    module: str,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
) -> list[Any]:
    """List soft-deleted records for a master data module."""
    cfg = _get_config(module)
    model = cfg["model"]
    stmt = select(model).where(model.is_deleted == True).order_by(model.deleted_at.desc())
    records = db.scalars(stmt).all()
    return [item for item in (_to_out(cfg, r) for r in records) if item is not None]


@router.post("/{module}")
def create_record(
    module: str,
    payload: dict[str, Any],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
) -> Any:
    """Create a new master data record."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]
    name_field = cfg["name_field"]
    payload = _prepare_payload(cfg, payload)

    code_val = payload.get(code_field)
    if not code_val:
        raise HTTPException(status_code=400, detail=f"Field '{code_field}' is required")
    if not payload.get(name_field):
        raise HTTPException(status_code=400, detail=f"Field '{name_field}' is required")

    existing = db.scalar(select(model).where(getattr(model, code_field) == code_val))
    if existing and not existing.is_deleted:
        raise HTTPException(status_code=400, detail=f"Record with code '{code_val}' already exists")
    if existing and existing.is_deleted:
        for key, val in payload.items():
            setattr(existing, key, val)
        existing.is_deleted = False
        existing.deleted_at = None
        existing.updated_by = current_user.id
        db.commit()
        db.refresh(existing)
        log_audit(
            db,
            user=current_user,
            action="RESTORE",
            module=cfg["name"],
            entity_id=existing.id,
            entity_code=str(code_val),
            details=f"Restored existing deleted {cfg['name']} record {code_val} on create",
            request=request,
        )
        return _to_out(cfg, existing)

    instance = model(**payload, created_by=current_user.id, updated_by=current_user.id)
    db.add(instance)
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="CREATE",
        module=cfg["name"],
        entity_id=instance.id,
        entity_code=str(code_val),
        details=f"Created {cfg['name']} record {code_val}",
        request=request,
    )
    return _to_out(cfg, instance)


@router.put("/{module}/{record_id}")
def update_record(
    module: str,
    record_id: int,
    payload: dict[str, Any],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
) -> Any:
    """Update an existing master data record."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]
    out_schema = cfg["out_schema"]

    instance = db.get(model, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="Record not found")

    new_code = payload.get(code_field)
    if new_code and new_code != getattr(instance, code_field):
        existing = db.scalar(select(model).where(getattr(model, code_field) == new_code))
        if existing:
            raise HTTPException(status_code=400, detail=f"Record with code '{new_code}' already exists")

    for key, val in payload.items():
        if hasattr(instance, key):
            setattr(instance, key, val)
    instance.updated_by = current_user.id
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="UPDATE",
        module=cfg["name"],
        entity_id=instance.id,
        entity_code=str(getattr(instance, code_field)),
        details=f"Updated {cfg['name']} record {record_id}",
        request=request,
    )
    return out_schema.model_validate(instance)


@router.delete("/{module}/{record_id}")
def soft_delete_record(
    module: str,
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
) -> dict[str, Any]:
    """Soft delete a record (move to deleted entries tab)."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]

    instance = db.get(model, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="Record not found")

    instance.is_deleted = True
    instance.deleted_at = datetime.now(UTC)
    db.commit()

    code_val = str(getattr(instance, code_field))
    log_audit(
        db,
        user=current_user,
        action="SOFT_DELETE",
        module=cfg["name"],
        entity_id=instance.id,
        entity_code=code_val,
        details=f"Soft deleted {cfg['name']} record {code_val}",
        request=request,
    )
    return {"status": "success", "message": "Record moved to deleted entries"}


@router.post("/{module}/{record_id}/restore")
def restore_record(
    module: str,
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
) -> dict[str, Any]:
    """Restore a soft-deleted record."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]

    instance = db.get(model, record_id)
    if not instance or not instance.is_deleted:
        raise HTTPException(status_code=404, detail="Deleted record not found")

    instance.is_deleted = False
    instance.deleted_at = None
    db.commit()

    code_val = str(getattr(instance, code_field))
    log_audit(
        db,
        user=current_user,
        action="RESTORE",
        module=cfg["name"],
        entity_id=instance.id,
        entity_code=code_val,
        details=f"Restored {cfg['name']} record {code_val}",
        request=request,
    )
    return {"status": "success", "message": "Record restored successfully"}


@router.delete("/{module}/{record_id}/permanent")
def permanent_delete_record(
    module: str,
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
) -> dict[str, Any]:
    """Permanently delete a record from deleted entries."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]

    instance = db.get(model, record_id)
    if not instance:
        raise HTTPException(status_code=404, detail="Record not found")

    code_val = str(getattr(instance, code_field))
    db.delete(instance)
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="PERMANENT_DELETE",
        module=cfg["name"],
        entity_id=record_id,
        entity_code=code_val,
        details=f"Permanently deleted {cfg['name']} record {code_val}",
        request=request,
    )
    return {"status": "success", "message": "Record permanently deleted"}


@router.post("/{module}/import", response_model=BulkImportResponse)
async def bulk_import(
    module: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
) -> BulkImportResponse:
    """Bulk import records from CSV or Excel file."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]
    name_field = cfg["name_field"]
    symbol_field = cfg["symbol_field"]

    contents = await file.read()
    filename = file.filename or ""

    rows: list[tuple[int, dict[str, Any]]] = []
    errors: list[str] = []

    try:
        if filename.endswith(".csv"):
            text_data = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text_data))
            for i, row in enumerate(reader, start=1):
                norm_row = {str(k).strip().lower().replace(" ", "_"): v for k, v in row.items() if k}
                rows.append((i, norm_row))
        elif filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Excel workbook has no active sheet")
            header_row = [str(cell.value).strip().lower().replace(" ", "_") for cell in ws[1] if cell.value is not None]
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                row_dict = {}
                for h_name, val in zip(header_row, row, strict=False):
                    if h_name:
                        row_dict[h_name] = str(val).strip() if val is not None else ""
                rows.append((r_idx, row_dict))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload a CSV or XLSX file.")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(exc)}") from exc

    imported_count = 0
    error_count = 0

    code_keys = {code_field, "code", code_field.replace("_", "")}
    name_keys = {name_field, "name", name_field.replace("_", "")}
    symbol_keys = {symbol_field, "symbol", symbol_field.replace("_", "")} if symbol_field else set()

    for r_num, row in rows:
        code_val = None
        for k in code_keys:
            if k in row and row[k]:
                code_val = row[k]
                break
        if not code_val:
            keys_list = list(row.keys())
            if keys_list and row[keys_list[0]]:
                code_val = row[keys_list[0]]

        name_val = None
        for k in name_keys:
            if k in row and row[k]:
                name_val = row[k]
                break
        if not name_val and len(row) > 1:
            keys_list = list(row.keys())
            if len(keys_list) > 1 and row[keys_list[1]]:
                name_val = row[keys_list[1]]

        symbol_val = None
        if symbol_field:
            for k in symbol_keys:
                if k in row and row[k]:
                    symbol_val = row[k]
                    break
            if not symbol_val and len(row) > 2:
                keys_list = list(row.keys())
                if len(keys_list) > 2 and row[keys_list[2]]:
                    symbol_val = row[keys_list[2]]

        desc_val = row.get("description", row.get("desc", ""))

        if not code_val:
            error_count += 1
            errors.append(f"Row {r_num}: Missing code")
            continue
        if not name_val:
            name_val = code_val

        existing = db.scalar(select(model).where(getattr(model, code_field) == code_val))
        if existing:
            setattr(existing, name_field, name_val)
            if symbol_field and symbol_val:
                setattr(existing, symbol_field, symbol_val)
            if desc_val:
                existing.description = desc_val
            existing.is_deleted = False
            existing.deleted_at = None
            existing.updated_by = current_user.id
            imported_count += 1
        else:
            create_data = {
                code_field: code_val,
                name_field: name_val,
                "description": desc_val if desc_val else None,
            }
            if symbol_field:
                create_data[symbol_field] = symbol_val if symbol_val else code_val
            try:
                inst = model(**create_data, created_by=current_user.id, updated_by=current_user.id)
                db.add(inst)
                imported_count += 1
            except Exception as exc:
                error_count += 1
                errors.append(f"Row {r_num} ({code_val}): {str(exc)}")

    db.commit()

    log_audit(
        db,
        user=current_user,
        action="BULK_IMPORT",
        module=cfg["name"],
        details=f"Imported {imported_count} records with {error_count} errors from {filename}",
        request=request,
    )

    return BulkImportResponse(
        imported_count=imported_count,
        error_count=error_count,
        errors=errors[:20],
        success=error_count == 0,
    )


@router.get("/{module}/export")
def export_records(
    module: str,
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
) -> Response:
    """Export master data records as Excel or CSV."""
    cfg = _get_config(module)
    model = cfg["model"]
    code_field = cfg["code_field"]
    name_field = cfg["name_field"]
    symbol_field = cfg["symbol_field"]

    stmt = select(model)
    if not include_deleted:
        stmt = stmt.where(model.is_deleted == False)
    records = db.scalars(stmt).all()

    log_audit(
        db,
        user=current_user,
        action="EXPORT",
        module=cfg["name"],
        details=f"Exported {len(records)} {cfg['name']} records as {format}",
        request=request,
    )

    headers = [code_field, name_field]
    if symbol_field:
        headers.append(symbol_field)
    headers.extend(["description", "is_deleted", "created_at"])

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in records:
            row_data = [
                getattr(r, code_field, ""),
                getattr(r, name_field, ""),
            ]
            if symbol_field:
                row_data.append(getattr(r, symbol_field, ""))
            row_data.extend([
                r.description or "",
                r.is_deleted,
                r.created_at.isoformat() if r.created_at else "",
            ])
            writer.writerow(row_data)
        content = output.getvalue()
        return Response(content=content, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={module}.csv"})
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = cfg["name"]
        ws.append(headers)
        for r in records:
            row_data = [
                getattr(r, code_field, ""),
                getattr(r, name_field, ""),
            ]
            if symbol_field:
                row_data.append(getattr(r, symbol_field, ""))
            row_data.extend([
                r.description or "",
                r.is_deleted,
                r.created_at.isoformat() if r.created_at else "",
            ])
            ws.append(row_data)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return Response(content=bio.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={module}.xlsx"})
