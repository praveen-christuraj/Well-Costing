"""Vendor/Supplier and Purchase Order / Service Order API routes with common template features."""

# pyright: reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnknownArgumentType=false, reportArgumentType=false, reportOptionalMemberAccess=false, reportUnknownParameterType=false, reportMissingTypeArgument=false, reportAttributeAccessIssue=false, reportGeneralTypeIssues=false

import csv
import io
import re
import uuid
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.api.dependencies.auth import get_current_user
from app.db.session import get_db
from app.models.master_data import PurchaseOrderServiceOrder, VendorSupplier
from app.models.user import User
from app.schemas.master_data import (
    BulkAttachmentUploadResponse,
    BulkImportResponse,
    PurchaseOrderOut,
    VendorSupplierOut,
)
from app.services.audit import log_audit

router = APIRouter(prefix="/master-data", tags=["master-data-extended"])

# Upload directory
BASE_DIR = Path(__file__).resolve().parents[4]
UPLOAD_DIR = BASE_DIR / "app" / "uploads" / "po_attachments"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".doc",
    ".xlsx",
    ".xls",
    ".csv",
    ".jpg",
    ".jpeg",
    ".png",
}
MAX_FILE_SIZE = 15 * 1024 * 1024  # 15 MB

PO_TYPES = {"PO", "SO", "Callout", "Others"}


def _validate_file_extension(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"File extension '{ext}' not allowed. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
        )
    return ext


def _save_upload_file(file: UploadFile, record_id: int) -> dict[str, Any]:
    """Save uploaded file to disk and return metadata."""
    filename = file.filename or f"upload_{uuid.uuid4().hex}"
    _validate_file_extension(filename)

    # Read content to check size
    content = file.file.read()
    size = len(content)
    if size > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File size exceeds 15 MB limit (got {size / 1024 / 1024:.2f} MB)")
    if size == 0:
        raise HTTPException(status_code=400, detail="Empty file not allowed")

    file.file.seek(0)

    record_dir = UPLOAD_DIR / str(record_id)
    record_dir.mkdir(parents=True, exist_ok=True)

    stored_name = f"{uuid.uuid4().hex}_{Path(filename).name}"
    stored_path = record_dir / stored_name

    with open(stored_path, "wb") as f:
        f.write(content)

    return {
        "attachment_path": str(stored_path.relative_to(BASE_DIR)),
        "attachment_original_name": filename,
        "attachment_mime_type": file.content_type or "application/octet-stream",
        "attachment_size": size,
    }


def _parse_date_flexible(value: str | None) -> date | None:
    """Parse date with flexible formats for bulk import."""
    if not value or not str(value).strip():
        return None
    val = str(value).strip()
    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d.%m.%Y",
        "%Y/%m/%d",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%m/%d/%Y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    try:
        if re.match(r"^\d+(\.\d+)?$", val):
            serial = float(val)
            if 1 < serial < 100000:
                base = datetime(1899, 12, 30)
                dt = base + timedelta(days=serial)
                return dt.date()
    except Exception:
        pass
    try:
        val_clean = val.split()[0]
        parts = re.split(r"[-/\. ]", val_clean)
        if len(parts) == 3:
            if len(parts[0]) == 4:
                y, m, d = parts
            else:
                d, m, y = parts
                if len(y) == 2:
                    y = "20" + y if int(y) < 50 else "19" + y
            return date(int(y), int(m), int(d))
    except Exception:
        pass
    raise ValueError(f"Unrecognized date format: '{value}'. Use YYYY-MM-DD or DD/MM/YYYY etc.")


# ---------- Vendor/Supplier Endpoints ----------
# Order matters: static paths before dynamic {record_id}


@router.get("/vendors/dropdown")
def list_vendors_dropdown(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    """Lightweight dropdown list for PO/SO vendor selection (shows code & name)."""
    stmt = select(VendorSupplier).where(VendorSupplier.is_deleted == False).order_by(VendorSupplier.vendor_code)
    records = db.scalars(stmt).all()
    return [
        {
            "id": r.id,
            "vendor_code": r.vendor_code,
            "vendor_name": r.vendor_name,
            "display_name": f"{r.vendor_code} - {r.vendor_name}",
            "label": f"{r.vendor_code} - {r.vendor_name}",
            "value": r.id,
        }
        for r in records
    ]


@router.get("/vendors/deleted", response_model=list[VendorSupplierOut])
def list_vendors_deleted(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    stmt = select(VendorSupplier).where(VendorSupplier.is_deleted == True).order_by(VendorSupplier.deleted_at.desc())
    records = db.scalars(stmt).all()
    return [VendorSupplierOut.model_validate(r) for r in records]


@router.get("/vendors/export")
def export_vendors(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    stmt = select(VendorSupplier)
    if not include_deleted:
        stmt = stmt.where(VendorSupplier.is_deleted == False)
    records = db.scalars(stmt).all()

    log_audit(
        db,
        user=current_user,
        action="EXPORT",
        module="Vendors/Suppliers",
        details=f"Exported {len(records)} vendors as {format}",
        request=request,
    )

    headers = ["vendor_code", "vendor_name", "contact", "description", "is_deleted", "created_at"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in records:
            writer.writerow(
                [
                    r.vendor_code,
                    r.vendor_name,
                    r.contact or "",
                    r.description or "",
                    r.is_deleted,
                    r.created_at.isoformat() if r.created_at else "",
                ]
            )
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=vendors_export.csv"},
        )
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors"
        ws.append(headers)
        for r in records:
            ws.append(
                [
                    r.vendor_code,
                    r.vendor_name,
                    r.contact or "",
                    r.description or "",
                    r.is_deleted,
                    r.created_at.isoformat() if r.created_at else "",
                ]
            )
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return Response(
            content=bio.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=vendors_export.xlsx"},
        )


@router.get("/vendors", response_model=list[VendorSupplierOut])
def list_vendors(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    include_deleted: bool = Query(False),
    search: str | None = None,
):
    stmt = select(VendorSupplier).order_by(VendorSupplier.id.desc())
    if not include_deleted:
        stmt = stmt.where(VendorSupplier.is_deleted == False)
    if search:
        like = f"%{search}%"
        stmt = stmt.where(
            or_(
                VendorSupplier.vendor_code.ilike(like),
                VendorSupplier.vendor_name.ilike(like),
                VendorSupplier.contact.ilike(like),
            )
        )
    records = db.scalars(stmt).all()
    return [VendorSupplierOut.model_validate(r) for r in records]


@router.get("/vendors/{record_id}", response_model=VendorSupplierOut)
def get_vendor(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    instance = db.get(VendorSupplier, record_id)
    if not instance:
        raise HTTPException(status_code=404, detail="Vendor/Supplier not found")
    return VendorSupplierOut.model_validate(instance)


@router.post("/vendors", response_model=VendorSupplierOut)
def create_vendor(
    payload: dict[str, Any],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    code = (payload.get("vendor_code") or payload.get("code") or "").strip()
    name = (payload.get("vendor_name") or payload.get("name") or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="Vendor/Supplier Code is required")
    if not name:
        raise HTTPException(status_code=400, detail="Vendor/Supplier Name is required")

    existing = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code == code))
    if existing and not existing.is_deleted:
        raise HTTPException(status_code=400, detail=f"Vendor code '{code}' already exists")
    if existing and existing.is_deleted:
        existing.vendor_name = name
        existing.contact = payload.get("contact")
        existing.description = payload.get("description")
        existing.is_deleted = False
        existing.deleted_at = None
        existing.updated_by = current_user.id
        db.commit()
        db.refresh(existing)
        log_audit(
            db,
            user=current_user,
            action="RESTORE",
            module="Vendors/Suppliers",
            entity_id=existing.id,
            entity_code=code,
            details=f"Restored existing deleted vendor {code} on create",
            request=request,
        )
        return VendorSupplierOut.model_validate(existing)

    instance = VendorSupplier(
        vendor_code=code,
        vendor_name=name,
        contact=payload.get("contact"),
        description=payload.get("description"),
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(instance)
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="CREATE",
        module="Vendors/Suppliers",
        entity_id=instance.id,
        entity_code=code,
        details=f"Created vendor {code} - {name}",
        request=request,
    )
    return VendorSupplierOut.model_validate(instance)


@router.put("/vendors/{record_id}", response_model=VendorSupplierOut)
def update_vendor(
    record_id: int,
    payload: dict[str, Any],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(VendorSupplier, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="Vendor/Supplier not found")

    new_code = payload.get("vendor_code")
    if new_code and new_code != instance.vendor_code:
        existing = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code == new_code))
        if existing and existing.id != record_id:
            raise HTTPException(status_code=400, detail=f"Vendor code '{new_code}' already exists")
        instance.vendor_code = new_code

    if payload.get("vendor_name"):
        instance.vendor_name = payload["vendor_name"]
    if "contact" in payload:
        instance.contact = payload["contact"]
    if "description" in payload:
        instance.description = payload["description"]

    instance.updated_by = current_user.id
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="UPDATE",
        module="Vendors/Suppliers",
        entity_id=instance.id,
        entity_code=instance.vendor_code,
        details=f"Updated vendor {instance.vendor_code}",
        request=request,
    )
    return VendorSupplierOut.model_validate(instance)


@router.delete("/vendors/{record_id}")
def soft_delete_vendor(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(VendorSupplier, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="Vendor/Supplier not found")

    active_po = db.scalar(
        select(PurchaseOrderServiceOrder).where(
            PurchaseOrderServiceOrder.vendor_id == record_id,
            PurchaseOrderServiceOrder.is_deleted == False,
        )
    )
    if active_po:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete vendor '{instance.vendor_code}' - it is referenced by active PO/SO '{active_po.po_so_number}'. Soft-delete those first or reassign.",
        )

    instance.is_deleted = True
    instance.deleted_at = datetime.now(UTC)
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="SOFT_DELETE",
        module="Vendors/Suppliers",
        entity_id=instance.id,
        entity_code=instance.vendor_code,
        details=f"Soft deleted vendor {instance.vendor_code}",
        request=request,
    )
    return {"status": "success", "message": "Vendor moved to deleted entries"}


@router.post("/vendors/{record_id}/restore")
def restore_vendor(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(VendorSupplier, record_id)
    if not instance or not instance.is_deleted:
        raise HTTPException(status_code=404, detail="Deleted vendor not found")

    instance.is_deleted = False
    instance.deleted_at = None
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="RESTORE",
        module="Vendors/Suppliers",
        entity_id=instance.id,
        entity_code=instance.vendor_code,
        details=f"Restored vendor {instance.vendor_code}",
        request=request,
    )
    return {"status": "success", "message": "Vendor restored"}


@router.delete("/vendors/{record_id}/permanent")
def permanent_delete_vendor(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(VendorSupplier, record_id)
    if not instance:
        raise HTTPException(status_code=404, detail="Vendor not found")

    code = instance.vendor_code
    db.delete(instance)
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="PERMANENT_DELETE",
        module="Vendors/Suppliers",
        entity_id=record_id,
        entity_code=code,
        details=f"Permanently deleted vendor {code}",
        request=request,
    )
    return {"status": "success", "message": "Vendor permanently deleted"}


@router.post("/vendors/import", response_model=BulkImportResponse)
async def import_vendors(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    contents = await file.read()
    filename = file.filename or ""

    rows: list[tuple[int, dict[str, Any]]] = []
    errors: list[str] = []

    try:
        if filename.endswith(".csv"):
            text_data = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text_data))
            for i, row in enumerate(reader, start=1):
                norm_row = {str(k).strip().lower().replace(" ", "_"): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
                rows.append((i, norm_row))
        elif filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Excel workbook has no active sheet")
            header_row = [str(cell.value).strip().lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                row_dict = {}
                for h_name, val in zip(header_row, row, strict=False):
                    if h_name:
                        row_dict[h_name] = str(val).strip() if val is not None else ""
                rows.append((r_idx, row_dict))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Upload CSV or XLSX.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc!s}") from exc

    imported = 0
    err_count = 0

    for r_num, row in rows:
        code = row.get("vendor_code") or row.get("code") or row.get("vendorcode") or ""
        name = row.get("vendor_name") or row.get("name") or row.get("vendorname") or ""
        contact = row.get("contact") or ""
        desc = row.get("description") or row.get("desc") or ""

        if not code:
            err_count += 1
            errors.append(f"Row {r_num}: Missing vendor_code")
            continue
        if not name:
            name = code

        existing = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code == code))
        if existing:
            existing.vendor_name = name
            if contact:
                existing.contact = contact
            if desc:
                existing.description = desc
            existing.is_deleted = False
            existing.deleted_at = None
            existing.updated_by = current_user.id
            imported += 1
        else:
            try:
                inst = VendorSupplier(
                    vendor_code=code,
                    vendor_name=name,
                    contact=contact or None,
                    description=desc or None,
                    created_by=current_user.id,
                    updated_by=current_user.id,
                )
                db.add(inst)
                imported += 1
            except Exception as exc:
                err_count += 1
                errors.append(f"Row {r_num} ({code}): {exc!s}")

    db.commit()

    log_audit(
        db,
        user=current_user,
        action="BULK_IMPORT",
        module="Vendors/Suppliers",
        details=f"Imported {imported} vendors with {err_count} errors from {filename}",
        request=request,
    )

    return BulkImportResponse(imported_count=imported, error_count=err_count, errors=errors[:20], success=err_count == 0)


# ---------- Purchase Orders / Service Orders ----------

def _build_po_out(po: PurchaseOrderServiceOrder) -> PurchaseOrderOut:
    vendor_code = po.vendor.vendor_code if po.vendor else None
    vendor_name = po.vendor.vendor_name if po.vendor else None
    vendor_display = f"{vendor_code} - {vendor_name}" if vendor_code and vendor_name else vendor_code or vendor_name
    return PurchaseOrderOut(
        id=po.id,
        po_type=po.po_type,
        vendor_id=po.vendor_id,
        po_so_number=po.po_so_number,
        effective_date=po.effective_date,
        value=float(po.value) if po.value is not None else None,
        is_amendment=po.is_amendment,
        amendment_number=po.amendment_number,
        remarks=po.remarks,
        attachment_path=po.attachment_path,
        attachment_original_name=po.attachment_original_name,
        attachment_mime_type=po.attachment_mime_type,
        attachment_size=po.attachment_size,
        is_deleted=po.is_deleted,
        deleted_at=po.deleted_at,
        created_at=po.created_at,
        updated_at=po.updated_at,
        vendor_code=vendor_code,
        vendor_name=vendor_name,
        vendor_display=vendor_display,
    )


@router.get("/purchase-orders/deleted", response_model=list[PurchaseOrderOut])
def list_po_deleted(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    stmt = select(PurchaseOrderServiceOrder).where(PurchaseOrderServiceOrder.is_deleted == True).order_by(PurchaseOrderServiceOrder.deleted_at.desc())
    records = db.scalars(stmt).all()
    return [_build_po_out(r) for r in records]


@router.get("/purchase-orders/export")
def export_purchase_orders(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    stmt = select(PurchaseOrderServiceOrder)
    if not include_deleted:
        stmt = stmt.where(PurchaseOrderServiceOrder.is_deleted == False)
    records = db.scalars(stmt).all()

    log_audit(
        db,
        user=current_user,
        action="EXPORT",
        module="Purchase Orders/Service Orders",
        details=f"Exported {len(records)} PO/SO as {format}",
        request=request,
    )

    headers = [
        "po_type",
        "vendor_code",
        "vendor_name",
        "po_so_number",
        "effective_date",
        "value",
        "is_amendment",
        "amendment_number",
        "remarks",
        "attachment_original_name",
        "is_deleted",
        "created_at",
    ]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in records:
            writer.writerow(
                [
                    r.po_type,
                    r.vendor.vendor_code if r.vendor else "",
                    r.vendor.vendor_name if r.vendor else "",
                    r.po_so_number,
                    r.effective_date.isoformat() if r.effective_date else "",
                    str(r.value) if r.value is not None else "",
                    r.is_amendment,
                    r.amendment_number or "",
                    r.remarks or "",
                    r.attachment_original_name or "",
                    r.is_deleted,
                    r.created_at.isoformat() if r.created_at else "",
                ]
            )
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=purchase_orders_export.csv"},
        )
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Orders"
        ws.append(headers)
        for r in records:
            ws.append(
                [
                    r.po_type,
                    r.vendor.vendor_code if r.vendor else "",
                    r.vendor.vendor_name if r.vendor else "",
                    r.po_so_number,
                    r.effective_date.isoformat() if r.effective_date else "",
                    str(r.value) if r.value is not None else "",
                    r.is_amendment,
                    r.amendment_number or "",
                    r.remarks or "",
                    r.attachment_original_name or "",
                    r.is_deleted,
                    r.created_at.isoformat() if r.created_at else "",
                ]
            )
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return Response(
            content=bio.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=purchase_orders_export.xlsx"},
        )


@router.post("/purchase-orders/attachments/bulk", response_model=BulkAttachmentUploadResponse)
async def bulk_upload_attachments(
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
) -> BulkAttachmentUploadResponse:
    """Bulk upload attachments: filename should contain PO/SO Number."""
    uploaded = 0
    err_count = 0
    errors: list[str] = []

    for upload in files:
        fname = upload.filename or ""
        if not fname:
            err_count += 1
            errors.append("One file had no filename")
            continue
        try:
            _validate_file_extension(fname)
            content = await upload.read()
            size = len(content)
            if size > MAX_FILE_SIZE:
                raise ValueError(f"File {fname} exceeds 15MB")
            if size == 0:
                raise ValueError(f"File {fname} is empty")

            base = Path(fname).stem
            candidate_number = base
            amend_candidate = None
            if "__" in base:
                parts = base.split("__")
                candidate_number = parts[0].strip()
                if len(parts) > 1:
                    try:
                        amend_candidate = int(re.sub(r"[^\d]", "", parts[1]) or 0) or None
                    except Exception:
                        amend_candidate = None
            elif "_" in base:
                parts = base.rsplit("_", 1)
                if len(parts) == 2 and parts[1].isdigit():
                    candidate_number = parts[0].strip()
                    try:
                        amend_candidate = int(parts[1])
                    except Exception:
                        amend_candidate = None

            candidate_number = candidate_number.strip()

            stmt = select(PurchaseOrderServiceOrder).where(PurchaseOrderServiceOrder.is_deleted == False)
            if amend_candidate:
                stmt = stmt.where(
                    PurchaseOrderServiceOrder.po_so_number == candidate_number,
                    PurchaseOrderServiceOrder.amendment_number == amend_candidate,
                )
            else:
                stmt = stmt.where(PurchaseOrderServiceOrder.po_so_number == candidate_number)
                stmt = stmt.order_by(PurchaseOrderServiceOrder.is_amendment, PurchaseOrderServiceOrder.id.desc())

            po_record = db.scalar(stmt)

            if not po_record:
                all_pos = db.scalars(select(PurchaseOrderServiceOrder).where(PurchaseOrderServiceOrder.is_deleted == False)).all()
                for po in all_pos:
                    if po.po_so_number.lower() in fname.lower() or po.po_so_number.lower() in base.lower():
                        if amend_candidate:
                            if po.amendment_number == amend_candidate:
                                po_record = po
                                break
                        else:
                            po_record = po
                            break

            if not po_record:
                err_count += 1
                errors.append(f"{fname}: No matching PO/SO found for number '{candidate_number}'")
                continue

            record_dir = UPLOAD_DIR / str(po_record.id)
            record_dir.mkdir(parents=True, exist_ok=True)
            stored_name = f"{uuid.uuid4().hex}_{Path(fname).name}"
            stored_path = record_dir / stored_name
            with open(stored_path, "wb") as f:
                f.write(content)

            if po_record.attachment_path:
                try:
                    old_full = BASE_DIR / po_record.attachment_path
                    if old_full.exists():
                        old_full.unlink()
                except Exception:
                    pass

            po_record.attachment_path = str(stored_path.relative_to(BASE_DIR))
            po_record.attachment_original_name = fname
            po_record.attachment_mime_type = upload.content_type or "application/octet-stream"
            po_record.attachment_size = size
            po_record.updated_by = current_user.id
            uploaded += 1

        except Exception as exc:
            err_count += 1
            errors.append(f"{fname}: {exc!s}")

    db.commit()

    log_audit(
        db,
        user=current_user,
        action="BULK_IMPORT",
        module="Purchase Orders/Service Orders",
        details=f"Bulk uploaded {uploaded} attachments with {err_count} errors",
        request=request,
    )

    return BulkAttachmentUploadResponse(
        uploaded_count=uploaded, error_count=err_count, errors=errors[:20], success=err_count == 0
    )


@router.post("/purchase-orders/import", response_model=BulkImportResponse)
async def import_purchase_orders(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    contents = await file.read()
    filename = file.filename or ""

    rows: list[tuple[int, dict[str, Any]]] = []
    errors: list[str] = []

    try:
        if filename.endswith(".csv"):
            text_data = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text_data))
            for i, row in enumerate(reader, start=1):
                norm_row = {str(k).strip().lower().replace(" ", "_"): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
                rows.append((i, norm_row))
        elif filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Excel workbook has no active sheet")
            header_row = [str(cell.value).strip().lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                row_dict = {}
                for h_name, val in zip(header_row, row, strict=False):
                    if h_name:
                        row_dict[h_name] = str(val).strip() if val is not None else ""
                rows.append((r_idx, row_dict))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Upload CSV or XLSX.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc!s}") from exc

    imported = 0
    err_count = 0

    for r_num, row in rows:
        try:
            po_type = (
                row.get("po_type")
                or row.get("type")
                or row.get("po/type")
                or row.get("order_type")
                or ""
            )
            po_type = str(po_type).strip()
            if po_type.lower() == "po":
                po_type = "PO"
            elif po_type.lower() == "so":
                po_type = "SO"
            elif po_type.lower() in ("callout", "call out", "call-out"):
                po_type = "Callout"
            elif po_type.lower() in ("others", "other"):
                po_type = "Others"

            if po_type not in PO_TYPES:
                raise ValueError(f"Invalid Type '{po_type}'. Must be one of {', '.join(PO_TYPES)}")

            vendor_ref = (
                row.get("vendor_id")
                or row.get("vendor_code")
                or row.get("vendor")
                or row.get("supplier")
                or row.get("vendor/supplier")
                or ""
            )
            vendor_ref = str(vendor_ref).strip()
            if not vendor_ref:
                raise ValueError("Vendor/Supplier is mandatory")

            vendor_obj = None
            try:
                vid = int(vendor_ref)
                vendor_obj = db.get(VendorSupplier, vid)
            except Exception:
                pass
            if not vendor_obj:
                vendor_obj = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code == vendor_ref, VendorSupplier.is_deleted == False))
            if not vendor_obj:
                vendor_obj = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code.ilike(vendor_ref), VendorSupplier.is_deleted == False))
            if not vendor_obj:
                vendor_obj = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_name.ilike(f"%{vendor_ref}%"), VendorSupplier.is_deleted == False))
            if not vendor_obj:
                raise ValueError(f"Vendor '{vendor_ref}' not found")

            po_number = (
                row.get("po_so_number")
                or row.get("po_number")
                or row.get("so_number")
                or row.get("number")
                or row.get("po/so_number")
                or row.get("po_so_no")
                or ""
            )
            po_number = str(po_number).strip()
            if not po_number:
                raise ValueError("PO/SO Number is mandatory")

            eff_raw = row.get("effective_date") or row.get("effective") or row.get("date") or ""
            eff_date = None
            if eff_raw:
                try:
                    eff_date = _parse_date_flexible(str(eff_raw))
                except ValueError as ve:
                    raise ValueError(f"Effective Date error: {ve}")

            value_raw = row.get("value") or row.get("amount") or ""
            value_dec = None
            if value_raw and str(value_raw).strip():
                try:
                    cleaned = re.sub(r"[^\d\.\-]", "", str(value_raw))
                    if cleaned:
                        value_dec = Decimal(cleaned)
                except Exception:
                    raise ValueError(f"Invalid Value '{value_raw}'")

            is_amend_raw = row.get("is_amendment") or row.get("amendment") or row.get("is_amend") or ""
            is_amend = False
            if str(is_amend_raw).lower() in ("1", "true", "yes", "y", "checked", "x"):
                is_amend = True

            amend_num_raw = row.get("amendment_number") or row.get("if_yes") or row.get("amendment_no") or row.get("amend_no") or ""
            amend_num = None
            if amend_num_raw and str(amend_num_raw).strip():
                try:
                    amend_num = int(float(str(amend_num_raw).strip()))
                    if not (1 <= amend_num <= 200):
                        raise ValueError("Amendment number must be 1-200")
                    is_amend = True
                except ValueError as ve:
                    if "Amendment number must be" in str(ve):
                        raise
                    raise ValueError(f"Invalid amendment number '{amend_num_raw}'")

            if is_amend and not amend_num:
                raise ValueError("Amendment number mandatory when amendment is Yes/checked")

            remarks = row.get("remarks") or row.get("remark") or ""

            if is_amend:
                dup = db.scalar(
                    select(PurchaseOrderServiceOrder).where(
                        PurchaseOrderServiceOrder.po_so_number == po_number,
                        PurchaseOrderServiceOrder.amendment_number == amend_num,
                        PurchaseOrderServiceOrder.is_deleted == False,
                    )
                )
            else:
                dup = db.scalar(
                    select(PurchaseOrderServiceOrder).where(
                        PurchaseOrderServiceOrder.po_so_number == po_number,
                        PurchaseOrderServiceOrder.is_amendment == False,
                        PurchaseOrderServiceOrder.is_deleted == False,
                    )
                )

            if dup:
                dup.po_type = po_type
                dup.vendor_id = vendor_obj.id
                if eff_date:
                    dup.effective_date = eff_date
                if value_dec is not None:
                    dup.value = value_dec
                dup.is_amendment = is_amend
                dup.amendment_number = amend_num if is_amend else None
                if remarks:
                    dup.remarks = remarks
                dup.updated_by = current_user.id
                imported += 1
            else:
                inst = PurchaseOrderServiceOrder(
                    po_type=po_type,
                    vendor_id=vendor_obj.id,
                    po_so_number=po_number,
                    effective_date=eff_date,
                    value=value_dec,
                    is_amendment=is_amend,
                    amendment_number=amend_num if is_amend else None,
                    remarks=remarks or None,
                    created_by=current_user.id,
                    updated_by=current_user.id,
                )
                db.add(inst)
                imported += 1

        except Exception as exc:
            err_count += 1
            errors.append(f"Row {r_num}: {exc!s}")

    db.commit()

    log_audit(
        db,
        user=current_user,
        action="BULK_IMPORT",
        module="Purchase Orders/Service Orders",
        details=f"Imported {imported} PO/SO with {err_count} errors from {filename}",
        request=request,
    )

    return BulkImportResponse(imported_count=imported, error_count=err_count, errors=errors[:30], success=err_count == 0)


@router.get("/purchase-orders", response_model=list[PurchaseOrderOut])
def list_purchase_orders(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    include_deleted: bool = Query(False),
    search: str | None = None,
    vendor_id: int | None = None,
    po_type: str | None = None,
):
    stmt = select(PurchaseOrderServiceOrder).order_by(PurchaseOrderServiceOrder.id.desc())
    if not include_deleted:
        stmt = stmt.where(PurchaseOrderServiceOrder.is_deleted == False)
    if search:
        like = f"%{search}%"
        stmt = stmt.where(
            or_(
                PurchaseOrderServiceOrder.po_so_number.ilike(like),
                PurchaseOrderServiceOrder.remarks.ilike(like),
            )
        )
    if vendor_id:
        stmt = stmt.where(PurchaseOrderServiceOrder.vendor_id == vendor_id)
    if po_type:
        stmt = stmt.where(PurchaseOrderServiceOrder.po_type == po_type)

    records = db.scalars(stmt).all()
    return [_build_po_out(r) for r in records]


@router.get("/purchase-orders/{record_id}", response_model=PurchaseOrderOut)
def get_purchase_order(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance:
        raise HTTPException(status_code=404, detail="PO/SO not found")
    return _build_po_out(instance)


@router.post("/purchase-orders", response_model=PurchaseOrderOut)
def create_purchase_order(
    payload: dict[str, Any],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    po_type = payload.get("po_type") or payload.get("type")
    if not po_type or po_type not in PO_TYPES:
        raise HTTPException(status_code=400, detail=f"Type is mandatory and must be one of {', '.join(PO_TYPES)}")

    vendor_id = payload.get("vendor_id") or payload.get("vendor")
    if not vendor_id:
        raise HTTPException(status_code=400, detail="Vendor/Supplier is mandatory")
    try:
        vendor_id_int = int(vendor_id)
    except (ValueError, TypeError):
        vendor_code_search = str(vendor_id).strip()
        vendor_obj = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code == vendor_code_search, VendorSupplier.is_deleted == False))
        if not vendor_obj:
            raise HTTPException(status_code=400, detail=f"Vendor '{vendor_id}' not found")
        vendor_id_int = vendor_obj.id

    vendor = db.get(VendorSupplier, vendor_id_int)
    if not vendor or vendor.is_deleted:
        raise HTTPException(status_code=404, detail="Vendor/Supplier not found or deleted")

    po_so_number = (payload.get("po_so_number") or payload.get("po_number") or payload.get("number") or "").strip()
    if not po_so_number:
        raise HTTPException(status_code=400, detail="PO/SO Number is mandatory")

    is_amendment = bool(payload.get("is_amendment", False))
    amendment_number = payload.get("amendment_number")

    if is_amendment:
        if amendment_number is None or amendment_number == "":
            raise HTTPException(status_code=400, detail="Amendment number is mandatory when amendment checkbox is checked (1-200)")
        try:
            amendment_number_int = int(amendment_number)
            if not (1 <= amendment_number_int <= 200):
                raise HTTPException(status_code=400, detail="Amendment number must be between 1 and 200")
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Amendment number must be an integer 1-200")
    else:
        amendment_number_int = None
        if amendment_number:
            try:
                amendment_number_int = int(amendment_number)
            except Exception:
                amendment_number_int = None
            if amendment_number_int:
                is_amendment = True

    if is_amendment:
        dup = db.scalar(
            select(PurchaseOrderServiceOrder).where(
                PurchaseOrderServiceOrder.po_so_number == po_so_number,
                PurchaseOrderServiceOrder.amendment_number == amendment_number_int,
                PurchaseOrderServiceOrder.is_deleted == False,
            )
        )
        if dup:
            raise HTTPException(status_code=400, detail=f"PO/SO Number '{po_so_number}' with Amendment {amendment_number_int} already exists")
    else:
        dup = db.scalar(
            select(PurchaseOrderServiceOrder).where(
                PurchaseOrderServiceOrder.po_so_number == po_so_number,
                PurchaseOrderServiceOrder.is_amendment == False,
                PurchaseOrderServiceOrder.is_deleted == False,
            )
        )
        if dup:
            raise HTTPException(status_code=400, detail=f"PO/SO Number '{po_so_number}' already exists. Use amendment if this is a revision.")

    eff_date_raw = payload.get("effective_date")
    eff_date = None
    if eff_date_raw:
        if isinstance(eff_date_raw, date) and not isinstance(eff_date_raw, datetime):
            eff_date = eff_date_raw
        elif isinstance(eff_date_raw, datetime):
            eff_date = eff_date_raw.date()
        elif isinstance(eff_date_raw, str) and eff_date_raw.strip():
            try:
                eff_date = _parse_date_flexible(eff_date_raw)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=str(e))

    value_raw = payload.get("value")
    value_dec = None
    if value_raw is not None and str(value_raw).strip() != "":
        try:
            cleaned = re.sub(r"[^\d\.\-]", "", str(value_raw))
            if cleaned:
                value_dec = Decimal(cleaned)
        except (InvalidOperation, ValueError):
            raise HTTPException(status_code=400, detail=f"Invalid value format: {value_raw}")

    instance = PurchaseOrderServiceOrder(
        po_type=po_type,
        vendor_id=vendor_id_int,
        po_so_number=po_so_number,
        effective_date=eff_date,
        value=value_dec,
        is_amendment=is_amendment,
        amendment_number=amendment_number_int if is_amendment else None,
        remarks=payload.get("remarks"),
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(instance)
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="CREATE",
        module="Purchase Orders/Service Orders",
        entity_id=instance.id,
        entity_code=po_so_number,
        details=f"Created {po_type} {po_so_number} for vendor {vendor.vendor_code} amendment={amendment_number_int if is_amendment else 'No'}",
        request=request,
    )
    return _build_po_out(instance)


@router.put("/purchase-orders/{record_id}", response_model=PurchaseOrderOut)
def update_purchase_order(
    record_id: int,
    payload: dict[str, Any],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="PO/SO not found")

    if "po_type" in payload or "type" in payload:
        new_type = payload.get("po_type") or payload.get("type")
        if new_type not in PO_TYPES:
            raise HTTPException(status_code=400, detail=f"Type must be one of {', '.join(PO_TYPES)}")
        instance.po_type = new_type

    if "vendor_id" in payload or "vendor" in payload:
        vendor_id = payload.get("vendor_id") or payload.get("vendor")
        if vendor_id:
            try:
                vendor_id_int = int(vendor_id)
            except Exception:
                vendor_code_search = str(vendor_id).strip()
                vendor_obj = db.scalar(select(VendorSupplier).where(VendorSupplier.vendor_code == vendor_code_search, VendorSupplier.is_deleted == False))
                if not vendor_obj:
                    raise HTTPException(status_code=400, detail=f"Vendor '{vendor_id}' not found")
                vendor_id_int = vendor_obj.id
            vendor = db.get(VendorSupplier, vendor_id_int)
            if not vendor or vendor.is_deleted:
                raise HTTPException(status_code=404, detail="Vendor not found")
            instance.vendor_id = vendor_id_int

    if "po_so_number" in payload:
        new_number = str(payload["po_so_number"]).strip()
        if new_number and new_number != instance.po_so_number:
            if instance.is_amendment:
                dup = db.scalar(
                    select(PurchaseOrderServiceOrder).where(
                        PurchaseOrderServiceOrder.po_so_number == new_number,
                        PurchaseOrderServiceOrder.amendment_number == instance.amendment_number,
                        PurchaseOrderServiceOrder.id != record_id,
                        PurchaseOrderServiceOrder.is_deleted == False,
                    )
                )
            else:
                dup = db.scalar(
                    select(PurchaseOrderServiceOrder).where(
                        PurchaseOrderServiceOrder.po_so_number == new_number,
                        PurchaseOrderServiceOrder.is_amendment == False,
                        PurchaseOrderServiceOrder.id != record_id,
                        PurchaseOrderServiceOrder.is_deleted == False,
                    )
                )
            if dup:
                raise HTTPException(status_code=400, detail=f"PO/SO Number '{new_number}' already exists with same amendment")
            instance.po_so_number = new_number

    if "effective_date" in payload:
        eff_raw = payload["effective_date"]
        if eff_raw in (None, ""):
            instance.effective_date = None
        elif isinstance(eff_raw, date) and not isinstance(eff_raw, datetime):
            instance.effective_date = eff_raw
        elif isinstance(eff_raw, datetime):
            instance.effective_date = eff_raw.date()
        elif isinstance(eff_raw, str):
            try:
                instance.effective_date = _parse_date_flexible(eff_raw)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=str(e))

    if "value" in payload:
        val_raw = payload["value"]
        if val_raw in (None, ""):
            instance.value = None
        else:
            try:
                cleaned = re.sub(r"[^\d\.\-]", "", str(val_raw))
                instance.value = Decimal(cleaned) if cleaned else None
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid value: {val_raw}")

    if "is_amendment" in payload:
        is_amend = bool(payload["is_amendment"])
        instance.is_amendment = is_amend
        if not is_amend:
            instance.amendment_number = None

    if "amendment_number" in payload:
        amend_raw = payload["amendment_number"]
        if amend_raw in (None, ""):
            if instance.is_amendment:
                raise HTTPException(status_code=400, detail="Amendment number mandatory when amendment is checked")
            instance.amendment_number = None
        else:
            try:
                amend_int = int(amend_raw)
                if not (1 <= amend_int <= 200):
                    raise HTTPException(status_code=400, detail="Amendment number 1-200")
                if not instance.is_amendment:
                    instance.is_amendment = True
                instance.amendment_number = amend_int
            except HTTPException:
                raise
            except Exception:
                raise HTTPException(status_code=400, detail="Amendment number must be 1-200")

    if "remarks" in payload:
        instance.remarks = payload["remarks"]

    if instance.is_amendment and not instance.amendment_number:
        raise HTTPException(status_code=400, detail="Amendment number mandatory when amendment checked")

    instance.updated_by = current_user.id
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="UPDATE",
        module="Purchase Orders/Service Orders",
        entity_id=instance.id,
        entity_code=instance.po_so_number,
        details=f"Updated {instance.po_type} {instance.po_so_number}",
        request=request,
    )
    return _build_po_out(instance)


@router.delete("/purchase-orders/{record_id}")
def soft_delete_po(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="PO/SO not found")

    instance.is_deleted = True
    instance.deleted_at = datetime.now(UTC)
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="SOFT_DELETE",
        module="Purchase Orders/Service Orders",
        entity_id=instance.id,
        entity_code=instance.po_so_number,
        details=f"Soft deleted {instance.po_type} {instance.po_so_number}",
        request=request,
    )
    return {"status": "success", "message": "PO/SO moved to deleted entries"}


@router.post("/purchase-orders/{record_id}/restore")
def restore_po(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance or not instance.is_deleted:
        raise HTTPException(status_code=404, detail="Deleted PO/SO not found")

    instance.is_deleted = False
    instance.deleted_at = None
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="RESTORE",
        module="Purchase Orders/Service Orders",
        entity_id=instance.id,
        entity_code=instance.po_so_number,
        details=f"Restored {instance.po_type} {instance.po_so_number}",
        request=request,
    )
    return {"status": "success", "message": "PO/SO restored"}


@router.delete("/purchase-orders/{record_id}/permanent")
def permanent_delete_po(
    record_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    request: Request,
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance:
        raise HTTPException(status_code=404, detail="PO/SO not found")

    if instance.attachment_path:
        try:
            full_path = BASE_DIR / instance.attachment_path
            if full_path.exists():
                full_path.unlink()
        except Exception:
            pass

    code = instance.po_so_number
    db.delete(instance)
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="PERMANENT_DELETE",
        module="Purchase Orders/Service Orders",
        entity_id=record_id,
        entity_code=code,
        details=f"Permanently deleted {code}",
        request=request,
    )
    return {"status": "success", "message": "PO/SO permanently deleted"}


@router.post("/purchase-orders/{record_id}/attachment")
async def upload_po_attachment(
    record_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance or instance.is_deleted:
        raise HTTPException(status_code=404, detail="PO/SO not found")

    if instance.attachment_path:
        try:
            old_full = BASE_DIR / instance.attachment_path
            if old_full.exists():
                old_full.unlink()
        except Exception:
            pass

    meta = _save_upload_file(file, record_id)
    instance.attachment_path = meta["attachment_path"]
    instance.attachment_original_name = meta["attachment_original_name"]
    instance.attachment_mime_type = meta["attachment_mime_type"]
    instance.attachment_size = meta["attachment_size"]
    instance.updated_by = current_user.id
    db.commit()
    db.refresh(instance)

    log_audit(
        db,
        user=current_user,
        action="UPDATE",
        module="Purchase Orders/Service Orders",
        entity_id=instance.id,
        entity_code=instance.po_so_number,
        details=f"Uploaded attachment {meta['attachment_original_name']} for {instance.po_so_number}",
        request=request,
    )

    return {"status": "success", "message": "Attachment uploaded", "data": _build_po_out(instance)}


@router.get("/purchase-orders/{record_id}/attachment")
def download_po_attachment(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance or not instance.attachment_path:
        raise HTTPException(status_code=404, detail="Attachment not found")

    full_path = BASE_DIR / instance.attachment_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Attachment file missing on server")

    return FileResponse(
        path=str(full_path),
        filename=instance.attachment_original_name or "attachment",
        media_type=instance.attachment_mime_type or "application/octet-stream",
    )


@router.delete("/purchase-orders/{record_id}/attachment")
def delete_po_attachment(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    request: Request = None,
):
    instance = db.get(PurchaseOrderServiceOrder, record_id)
    if not instance or not instance.attachment_path:
        raise HTTPException(status_code=404, detail="Attachment not found")

    try:
        full_path = BASE_DIR / instance.attachment_path
        if full_path.exists():
            full_path.unlink()
    except Exception:
        pass

    instance.attachment_path = None
    instance.attachment_original_name = None
    instance.attachment_mime_type = None
    instance.attachment_size = None
    instance.updated_by = current_user.id
    db.commit()

    log_audit(
        db,
        user=current_user,
        action="UPDATE",
        module="Purchase Orders/Service Orders",
        entity_id=instance.id,
        entity_code=instance.po_so_number,
        details=f"Deleted attachment for {instance.po_so_number}",
        request=request,
    )

    return {"status": "success", "message": "Attachment deleted"}


